import os
import re
import time
import shutil
import zipfile
import logging
import subprocess
import builtins
import argparse
import psycopg2
import pytz
from datetime import datetime
from shutil import copyfile
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, PatternFill
from pyspark.sql import SparkSession, Row
from pyspark.sql import functions as F
from pyspark.sql.functions import from_unixtime,greatest,udf,col, lit, expr, when, trim, lower, upper, date_format, concat, concat_ws,substring,current_timestamp, row_number, monotonically_increasing_id, nanvl,abs,ltrim, rtrim,regexp_replace,size, split,nth_value, length,count,max,collect_list,explode,array, array_contains,transform,array_sort,struct
from pyspark.sql import DataFrame as SparkDataFrame
from pyspark.sql.utils import AnalysisException
from pyspark.sql.types import StructType, StructField, StringType, IntegerType,DecimalType,LongType,TimestampType,ArrayType,Row
import numpy as np
from pyspark.sql.window import Window
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathvalidate import sanitize_filepath
import argparse
from delta.tables import DeltaTable
from pyspark.sql.functions import array, struct
from pyspark.sql import DataFrame
import random
from functools import wraps
import fnmatch

def get_spark_session():
    return SparkSession.builder.appName("Tradepanel").getOrCreate()

def get_logger():

    logger = logging.getLogger('TP_logger')
    logger.setLevel(logging.INFO)
    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    stream_handler.setFormatter(formatter)
    if logger.hasHandlers():
        logger.handlers.clear()
    logger.addHandler(stream_handler)
    return logger

def read_run_params():
    parser = argparse.ArgumentParser()
    parser.add_argument("--FILE_NAME",type=str)
    parser.add_argument("--CNTRT_ID",type=str)
    parser.add_argument("--RUN_ID",type=str)
    parser.add_argument("--ENV",type=str)
    parser.add_argument("--refresh_type",type=str)
    parser.add_argument("--job_run_id",type=str)
    parser.add_argument("--retention_days",type=str)
    parser.add_argument("--frequency",type=str)
    parser.add_argument("--mode",type=str)
    parser.add_argument("--databricks_wkf_name",type=str)
    parser.add_argument("--postgres_wkf_name",type=str)
    parser.add_argument("--wkf_job_id",type=str)
    args = parser.parse_args()
    return args

def get_dbutils(spark):
    try:
        from pyspark.dbutils import DBUtils
        dbutils = DBUtils(spark)
    except ImportError:
        import IPython
        ipy = IPython.get_ipython()
        if ipy:
            dbutils = ipy.user_ns["dbutils"]
    return dbutils

def derive_base_path(spark):
    dbutils = get_dbutils(spark)
    storage_name= dbutils.secrets.get('tp_dpf2cdl', 'storageName') 
    base_storage_path=f"abfss://tp-source-data@{storage_name}.dfs.core.windows.net"
    return base_storage_path
    
def derive_publish_path(spark):
    dbutils = get_dbutils(spark)
    storage_name= dbutils.secrets.get('tp_dpf2cdl', 'storageName') 
    publish_storage_path=f"abfss://tp-publish-data@{storage_name}.dfs.core.windows.net"
    return publish_storage_path

def materialise_path(spark):
    base_path = derive_base_path(spark)
    mat_path = f"{base_path}/temp/materialised"
    return mat_path


def get_database_config(dbutils):
    """Retrieve all database configuration secrets."""
    return {
        'ref_db_jdbc_url': dbutils.secrets.get('tp_dpf2cdl', 'refDBjdbcURL'),
        'ref_db_name': dbutils.secrets.get('tp_dpf2cdl', 'refDBname'),
        'ref_db_user': dbutils.secrets.get('tp_dpf2cdl', 'refDBuser'),
        'ref_db_pwd': dbutils.secrets.get('tp_dpf2cdl', 'refDBpwd'),
        'ref_db_hostname': dbutils.secrets.get('tp_dpf2cdl', 'refDBhostname'),
        'catalog_name': dbutils.secrets.get('tp_dpf2cdl', 'databricks-catalog-name'),
        'postgres_schema': dbutils.secrets.get('tp_dpf2cdl', 'database-postgres-schema'),
        'consol_postgres_schema': dbutils.secrets.get('tp_dpf2cdl', 'database-postgres-schema-console')
    }
def lst_database_tables(spark, catalog_name):
    df_lst_tables_internal = spark.sql(f"show tables from {catalog_name}.internal_tp")
    df_lst_tables_gold = spark.sql(f"show tables from {catalog_name}.gold_tp")
    df_lst_tables = df_lst_tables_internal.unionByName(df_lst_tables_gold,True).withColumn('table_name', concat('database',lit('.'),'tableName' ))
    return df_lst_tables

def sanitize_variable(var):
    if not re.match(r'^[a-zA-Z0-9_.-]+$', str(var)): # Only allows letters, numbers, and underscores
        raise ValueError(f"Unsafe input detected: {var}")
    return var

# retry decorator
def retry_with_backoff(max_retries=6, initial_delay=5, max_delay=80, backoff_factor=2, jitter=True):
    """
    Decorator to retry functions with exponential backoff for concurrency-related conflicts.

    Args:
        max_retries: Maximum number of retry attempts
        initial_delay: Initial delay in seconds
        max_delay: Maximum delay between retries
        backoff_factor: Multiplier for delay after each retry
        jitter: Add randomization to avoid thundering herd
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            delay = initial_delay
            last_exception = None

            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    last_exception = e
                    error_msg = str(e).lower()

                    # Check for concurrency-related keywords
                    if any(keyword in error_msg for keyword in [
                        'concurrent', 'conflict', 'concurrent update',
                        'concurrent transaction', 'concurrent write',
                        'version mismatch', 'optimistic transaction',
                        'concurrentappendexception', 'concurrentmodificationexception'
                    ]):
                        if attempt < max_retries - 1:
                            current_delay = min(delay * (backoff_factor ** attempt), max_delay)
                            if jitter:
                                current_delay *= (0.5 + random.random())

                            logger = get_logger()
                            logger.warning(
                                f"Concurrency conflict detected in {func.__name__}. "
                                f"Retry {attempt + 1}/{max_retries} after {current_delay:.2f}s. "
                            )
                            time.sleep(current_delay)
                        else:
                            logger = get_logger()
                            logger.error(
                                f"Max retries ({max_retries}) reached for {func.__name__}. "
                                f"Last error: {str(e)}"
                            )
                            raise
                    else:
                        # Re-raise if it's not a concurrency issue
                        raise

            # If we've exhausted retries, raise the last exception
            if last_exception:
                raise last_exception

        return wrapper
    return decorator

@retry_with_backoff()
def safe_delete_with_retry(spark,target_table,run_id,cntrt_id,srce_sys_id):
    logger = get_logger()
    logger.info(f"Executing Delete from {target_table}")
    spark.sql(f"DELETE FROM {target_table} WHERE RUN_ID = :run_id and cntrt_id=:cntrt_id and srce_sys_id=:srce_sys_id",{"run_id":run_id,"cntrt_id":cntrt_id,"srce_sys_id":srce_sys_id})
    logger.info(f"Successfully deleted from {target_table}")

def safe_merge_with_retry(spark, df, target_table, merge_condition):
    logger = get_logger()

    @retry_with_backoff()
    def execute_merge():
        df.createOrReplaceTempView('src_view')
        merge_sql = f"""
        MERGE INTO {target_table} tgt
        USING src_view src
        ON {merge_condition}
        WHEN MATCHED THEN UPDATE SET *
        WHEN NOT MATCHED THEN INSERT *
        """
        logger.info(f"Executing MERGE into {target_table}")
        print(merge_sql)
        spark.sql(merge_sql)
        logger.info(f"Successfully merged into {target_table}")

    execute_merge()

def safe_write_with_retry(df, table_name, mode, partition_by=None, options=None):
    logger = get_logger()

    # Initialize options dict
    options = options or {}

    # If mode is overwrite, set dynamic as default unless user overrides
    if mode == "overwrite":
        options.setdefault("partitionOverwriteMode", "dynamic")

    @retry_with_backoff()
    def execute_write():
        logger.info(f"Writing to {table_name} in {mode} mode")
        writer = df.write.format("delta").mode(mode)

        if partition_by:
            writer = writer.partitionBy(partition_by)

        if options:
            for key, value in options.items():
                writer = writer.option(key, value)

        writer.saveAsTable(table_name)
        logger.info(f"Successfully written to {table_name}")

    execute_write()

# Publishing delta logs 
@retry_with_backoff()
def publish_delta_logs(spark,physical_table_name,catalog_name,df_partition_list):
    logger = get_logger()
    logger.info("Starting to publish delta logs")

    def get_partition_values_from_json_dict(
        json_partition_values_dict, partition_columns
    ) -> list:
        """The method prepares partition_values with the appropriate nested schema based on partitionValues element read from the Delta Log JSON file.

        :param json_partition_values_dict: partitionValues element read from the Delta Log JSON file.
        :param list partition_columns: List of all columns used for partitioning the table.
        :returns list: partition_values list with Rows with the appropriate nested schema.
        """
        json_partition_values_dict = json_partition_values_dict.asDict()

        partition_values = []

        for partition_column in partition_columns:
            partition_values.append(
                Row(
                    partition_column_name=partition_column,
                    partition_column_value=json_partition_values_dict[partition_column],
                )
            )
        return partition_values


    def get_partition_values_from_json_dict_udf(partition_columns: list):
        return udf(
            lambda col1: get_partition_values_from_json_dict(col1, partition_columns),
            returnType=ArrayType(
                StructType(
                    [
                        StructField("partition_column_name", StringType()),
                        StructField("partition_column_value", StringType()),
                    ]
                )
            ),
        )


    class UCPartitionChangeNotificationService:
        PARTITION_VALUES_TYPE = ArrayType(
            StructType(
                [
                    StructField("partition_column_name", StringType()),
                    StructField("partition_column_value", StringType()),
                ]
            )
        )
        LOG_TABLE_SUFFIX = "_partition_update_log"

        def _convert_deltalog_t_to_timestamp(self, deltalog_timestamp: str) -> str:
            """The method converts a Unix timestamp from Delta Log JSON to a Timestamp.

            :param str deltalog_timestamp: Unix time value.
            :returns str: Formatted timestamp as string.
            """
            return from_unixtime(
                expr(
                    "substring("
                    + deltalog_timestamp
                    + ", 1, length("
                    + deltalog_timestamp
                    + ")-3)"
                )
            )

        def _get_last_commit_version(self, full_uc_table_name: str) -> int:
            """The method finds the last commit version in the history for provided table.

            :param str full_uc_table_name: Full Unity Catalog path (catalog.schema.table) of the table.
            :returns int: The last commit version in the history.
            """
            return (
                spark.sql(f"describe history {full_uc_table_name}")
                .orderBy(col("version").desc())
                .select("version")
                .first()[0]
            )

        def _add_partition_values_column(
            self, input_df: DataFrame, partition_columns
        ) -> DataFrame:
            """The method adds a partition_values column with the appropriate nested schema based on the provided DataFrame with partitions columns in a flat form.

            :param DataFrame input_df: DataFrame with partitions columns in a flat form.
            :param list partition_columns: List of all columns used for partitioning the table.
            :returns DataFrame: DataFrame with the added partition_values column.
            """
            return input_df.withColumn(
                "partition_values",
                array(
                    *[
                        struct(
                            lit(partition_column).alias("partition_column_name"),
                            col(partition_column)
                            .cast(StringType())
                            .alias("partition_column_value"),
                        )
                        for partition_column in partition_columns
                    ]
                ),
            )

        def _generate_partitions_update_logs_using_publisher_df(
            self,
            publisher_partitions_update_logs_df: DataFrame,
            partition_columns: list,
            logger=None,
        ) -> DataFrame:
            """The method generates partitions update logs for the provided table based on the DataFrame provided by the Publisher.

            :param DataFrame publisher_partitions_update_logs_df: DataFrame provided by the Publisher indicating which partitions were changed in the given run.
            :param list partition_columns: List of all columns used for partitioning the table.
            :param Logger logger: Publisher's logger. If not provided, messages will be displayed in the console output.
            :returns DataFrame: DataFrame with generated partitions update logs.
            """
            expected_columns = partition_columns.copy()
            expected_columns.extend(["run_id", "update_timestamp"])

            for column in expected_columns:
                if column not in publisher_partitions_update_logs_df.columns:
                    self._log_message(
                        f"Expected column '{column}' is missing in the provided DataFrame.",
                        logger,
                    )
                    self._log_message(
                        "The process of inserting partitions update logs - FAILED.",
                        logger,
                    )
                    return None

            return self._add_partition_values_column(
                publisher_partitions_update_logs_df, partition_columns
            ).select(
                col("run_id").cast(StringType()), "partition_values", "update_timestamp"
            )

        def _generate_partitions_update_logs_using_cdf(
            self, full_uc_table_name: str, partition_columns: list, commit_version: int
        ) -> DataFrame:
            """The method generates partitions update logs for the provided table based on the CDF.

            :param str full_uc_table_name: Full Unity Catalog path (catalog.schema.table) of the table for which partition logs should be generated.
            :param list partition_columns: List of all columns used for partitioning the table.
            :param int commit_version: Delta history commit version for which partition logs should be generated.
            :returns DataFrame: DataFrame with generated partitions update logs.
            """
            partition_columns_select_str = ",".join(partition_columns)

            cdf_df = spark.sql(
                f"select distinct _commit_version, _commit_timestamp, {partition_columns_select_str} FROM table_changes('{full_uc_table_name}', {commit_version},{commit_version})"
            )

            final_df = self._add_partition_values_column(cdf_df, partition_columns).select(
                col("_commit_version").cast(StringType()).alias("run_id"),
                "partition_values",
                col("_commit_timestamp").alias("update_timestamp"),
            )

            return final_df

        def _generate_partitions_update_logs_using_delta_log(
            self, full_uc_table_name: str, partition_columns: list, commit_version: int
        ) -> DataFrame:
            """The method generates partitions update logs for the provided table based on the Delta Log JSON.

            :param str full_uc_table_name: Full Unity Catalog path (catalog.schema.table) of the table for which partition logs should be generated.
            :param list partition_columns: List of all columns used for partitioning the table.
            :param int commit_version: Delta history commit version for which partition logs should be generated.
            :returns DataFrame: DataFrame with generated partitions update logs.
            """
            table_external_location = (
                DeltaTable.forName(spark, full_uc_table_name)
                .detail()
                .select("location")
                .collect()[0][0]
            )

            delta_log_name = str(commit_version).zfill(20) + ".json"

            delta_log_df = spark.read.json(
                table_external_location + "/_delta_log/" + delta_log_name
            )

            schema = StructType(
                [
                    StructField("partition_values", self.PARTITION_VALUES_TYPE),
                    StructField("update_timestamp", StringType()),
                ]
            )
            partitions_add_df = spark.createDataFrame(data=[], schema=schema)
            partitions_remove_df = spark.createDataFrame(data=[], schema=schema)

            partition_values_from_json_dict_udf = get_partition_values_from_json_dict_udf(
                partition_columns
            )
            if "add" in delta_log_df.columns:
                if "partitionValues" in str(delta_log_df.select("add").schema):
                    delta_log_df = delta_log_df.withColumn(
                        "partition_values",
                        partition_values_from_json_dict_udf(col("add.partitionValues")),
                    )
                else:
                    delta_log_df = delta_log_df.withColumn(
                        "partition_values", lit(None).cast(self.PARTITION_VALUES_TYPE)
                    )

                partitions_add_df = (
                    delta_log_df.where("add is not null")
                    .groupBy(col("partition_values"))
                    .agg(max("add.modificationTime").alias("modificationTime"))
                    .select(
                        "partition_values",
                        self._convert_deltalog_t_to_timestamp("modificationTime")
                        .cast(TimestampType())
                        .alias("update_timestamp"),
                    )
                )

            if "remove" in delta_log_df.columns:
                if "partitionValues" in str(delta_log_df.select("remove").schema):
                    delta_log_df = delta_log_df.withColumn(
                        "partition_values",
                        partition_values_from_json_dict_udf(col("remove.partitionValues")),
                    )
                else:
                    delta_log_df = delta_log_df.withColumn(
                        "partition_values", lit(None).cast(self.PARTITION_VALUES_TYPE)
                    )

                partitions_remove_df = (
                    delta_log_df.where("remove is not null")
                    .groupBy(col("partition_values"))
                    .agg(max("remove.deletionTimestamp").alias("deletionTimestamp"))
                    .select(
                        "partition_values",
                        self._convert_deltalog_t_to_timestamp("deletionTimestamp")
                        .cast(TimestampType())
                        .alias("update_timestamp"),
                    )
                )

            partitions_changes_df = (
                partitions_add_df.join(
                    partitions_remove_df, on=["partition_values"], how="full_outer"
                )
                .withColumn(
                    "greatest_update_timestamp",
                    greatest(
                        partitions_add_df.update_timestamp,
                        partitions_remove_df.update_timestamp,
                    ),
                )
                .select("partition_values", col("greatest_update_timestamp"))
                .groupBy("partition_values")
                .agg(max("greatest_update_timestamp").alias("update_timestamp"))
            )

            partitions_changes_df = partitions_changes_df.withColumn(
                "run_id", lit(commit_version).cast(StringType())
            )

            return partitions_changes_df.select(
                "run_id", "partition_values", "update_timestamp"
            )

        def _log_message(self, message: str, logger=None):
            """The method logs the status of the process.

            :param str message: The message to be logged.
            :param Logger logger: Publisher's logger. If not provided, messages will be displayed in the console output.
            """
            if logger != None:
                logger.info(message)
            else:
                print(message)

        def insert_partitions_update_logs(
            self,
            full_uc_table_name: str,
            commit_version: int = None,
            publisher_partitions_update_logs_df: DataFrame = None,
            logger=None,
        ) -> int:
            """The method generates partitions update logs for the provided table and inserts these logs into [full_uc_table_name]_partition_update_log.
            The order of methods for inserting partitions update logs:
                1) Using the DataFrame provided by the Publisher, indicating which partitions were changed in the given run.
                2) Using Change Data Feed (CDF), if enabled.
                3) Using the Delta Log JSON file.

            :param str full_uc_table_name: Full Unity Catalog path (catalog.schema.table) of the table for which partition logs should be generated.
            :param int commit_version: Delta history commit version for which partition logs should be generated.
                It is used in CDF and Delta Log scenarios.
                If not provided, logs will be generated for the last commit version in history.
            :param DataFrame publisher_partitions_update_logs_df: DataFrame provided by the Publisher indicating which partitions were changed in the given run.
                Required columns: run_id, update_timestamp and all columns used for partitioning the table.
                If not provided, logs will be generated based on CDF or the Delta Log.
            :param Logger logger: Publisher's logger. If not provided, messages will be displayed in the console output.
            :returns int: Status of logs generation: 0 - Success, 1 - Failure.
            """
            self._log_message(
                f"The process of inserting partitions update logs for the {full_uc_table_name} table - STARTED.",
                logger,
            )

            if spark.catalog.tableExists(full_uc_table_name) == False:
                self._log_message(f"Table {full_uc_table_name} does not exist.", logger)
                self._log_message(
                    f"The process of inserting partitions update logs for the {full_uc_table_name} table - FAILED.",
                    logger,
                )
                return 1

            partition_columns = (
                DeltaTable.forName(spark, full_uc_table_name)
                .detail()
                .select("partitionColumns")
                .collect()[0][0]
            )

            if publisher_partitions_update_logs_df != None:
                self._log_message(
                    "Partitions update logs are being generated based on the DataFrame input from the Publisher.",
                    logger,
                )
                partitions_update_logs_df = (
                    self._generate_partitions_update_logs_using_publisher_df(
                        publisher_partitions_update_logs_df, partition_columns
                    )
                )
                if partitions_update_logs_df == None:
                    return 1
            else:
                table_format = (
                    DeltaTable.forName(spark, full_uc_table_name)
                    .detail()
                    .select("format")
                    .collect()[0][0]
                )
                if table_format != "delta":
                    self._log_message(
                        f"ChangeDataFeed and Transactional Logs are only available for delta objects. {full_uc_table_name} has {table_format} format.",
                        logger,
                    )
                    self._log_message(
                        f"The process of inserting partitions update logs for the {full_uc_table_name} table - FAILED.",
                        logger,
                    )
                    return 1

                if commit_version == None:
                    commit_version = self._get_last_commit_version(full_uc_table_name)

                is_cdf_enabled = (
                    spark.sql(
                        f"SHOW TBLPROPERTIES {full_uc_table_name}(delta.enableChangeDataFeed)"
                    )
                    .select("value")
                    .collect()[0][0]
                )

                if is_cdf_enabled == "true":
                    self._log_message(
                        f"Partitions update logs are being generated based on the Change Data Feed for the version {commit_version}.",
                        logger,
                    )

                    partitions_update_logs_df = (
                        self._generate_partitions_update_logs_using_cdf(
                            full_uc_table_name, partition_columns, commit_version
                        )
                    )
                else:
                    self._log_message(
                        f"Partitions update logs are being generated based on the Delta Log file for the version {commit_version}",
                        logger,
                    )

                    partitions_update_logs_df = (
                        self._generate_partitions_update_logs_using_delta_log(
                            full_uc_table_name, partition_columns, commit_version
                        )
                    )

            partition_update_log_table_name = full_uc_table_name + self.LOG_TABLE_SUFFIX
            partitions_update_logs_df.write.format("delta").mode("append").saveAsTable(
                partition_update_log_table_name
            )
            self._log_message(
                f"The number of log entries inserted: {partitions_update_logs_df.count()}",
                logger,
            )
            self._log_message(
                f"The process of inserting partitions update logs for the {full_uc_table_name} table - FINISHED.",
                logger,
            )
            return 0
    
    service = UCPartitionChangeNotificationService()
    service.insert_partitions_update_logs(full_uc_table_name=f"{catalog_name}.gold_tp.{physical_table_name}", publisher_partitions_update_logs_df=df_partition_list)

def load_cntrt_col_assign(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ):
    query = f"{postgres_schema}.mm_col_asign_lkp"
    df_dpf_col_asign_vw=read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ).filter(f"cntrt_id = {cntrt_id}")
    return df_dpf_col_asign_vw

def load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ):
    query = f"{postgres_schema}.mm_cntrt_lkp"
    # Read the data from PostgreSQL into a DataFrame
    df_cntrt_lkp=read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ).filter(f"cntrt_id= {cntrt_id}")
    return df_cntrt_lkp

def load_cntrt_file_lkp(cntrt_id, dmnsn_name, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ):
    query = f"""{postgres_schema}.mm_cntrt_file_lkp"""
    # Read the data from PostgreSQL into a DataFrame
    df_cntrt_file_lkp=read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ).filter(f"cntrt_id={cntrt_id} AND dmnsn_name='{dmnsn_name}' ").select('file_patrn')
    return df_cntrt_file_lkp

def load_cntrt_dlmtr_lkp(cntrt_id, dmnsn_name, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ):
    df_cntrt_file_lkp = read_from_postgres(f"{postgres_schema}.mm_cntrt_file_lkp", spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ).filter(f"cntrt_id={cntrt_id} AND dmnsn_name='{dmnsn_name}'")
    df_cntrt_file_lkp.createOrReplaceTempView('mm_cntrt_file_lkp')
    
    df_col_dlmtr_lkp = read_from_postgres(f"{postgres_schema}.mm_col_dlmtr_lkp", spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    df_col_dlmtr_lkp.createOrReplaceTempView('mm_col_dlmtr_lkp')

    df_cntrt_dlmtr_lkp = spark.sql(""" SELECT dt.dlmtr_val FROM mm_cntrt_file_lkp ct join mm_col_dlmtr_lkp dt on ct.dlmtr_id = dt.dlmtr_id """)

    return df_cntrt_dlmtr_lkp

def load_mm_cntrt_categ_assoc(spark,postgres_schema,cntrt_id,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    df_cntrt_categ_assoc_query = f"{postgres_schema}.mm_cntrt_categ_assoc"
    df_cntrt_categ_assoc=read_from_postgres(df_cntrt_categ_assoc_query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id = {cntrt_id}").select('categ_id')
    return df_cntrt_categ_assoc

def load_cntry_lkp(spark,postgres_schema,cntry_name,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    df = read_from_postgres(f"{postgres_schema}.mm_cntry_lkp",spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntry_name='{cntry_name}'").select('cntry_id')
    return df

def acn_prod_trans_materialize(df,run_id):
    spark = get_spark_session()
    path=f'{materialise_path(spark)}/{run_id}/product_transformation_df_prod_stgng_vw'
    df.write.format("parquet").options(header=True).mode('overwrite').save(path)
def acn_prod_trans(srce_sys_id, run_id, catalog_name , spark):
    logger = get_logger()
    # Read PROD_DIM schema from Delta Table
    df_sch_prod_dim = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_prod_dim limit 0")

    # Create a temporary view from the complemented DataFrame
    df_sch_prod_dim.createOrReplaceTempView("df_sch_prod_dim")
    query = """ select a.*except(a.run_id,a.prod_skid,a.cntrt_id, a.srce_sys_id, a.part_srce_sys_id, a.part_cntrt_id) from df_sch_prod_dim a"""
    df_sch_prod_dim = spark.sql(query)
    # Read the Product parquet file into DataFrame
    df_rawfile_input = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/load_product_df_prod_extrn")

    # Complement the columns of df_rawfile_input with those in tp_prod_dim
    df_rawfile_input = df_rawfile_input.unionByName(df_sch_prod_dim,allowMissingColumns=True)

    df_rawfile_input = df_rawfile_input.withColumn('last_sellg_txt',date_format(when(trim(col('last_sellg_txt')) == 'NOT APPLICABLE', '1991-04-15').otherwise(trim(col('last_sellg_txt'))), 'yyyy-MM-dd'))

    df_rawfile_input.createOrReplaceTempView("df_rawfile_input")
    

    # Remove the duplicates from the df_rawfile_input DataFrame
    df_final = spark.sql("""
    WITH row_deduplicated AS (
        SELECT *,
            ROW_NUMBER() OVER (
                PARTITION BY pg_categ_txt, pg_super_categ_txt, upc_txt 
                ORDER BY last_sellg_txt DESC
            ) AS row_number_deduplicator_column
        FROM df_rawfile_input
    )
    SELECT *
    FROM row_deduplicated
    WHERE row_number_deduplicator_column = 1
    AND pg_categ_txt IS NOT NULL""")

    # Drop the unwanted columns
    df = df_final.drop("row_number_deduplicator_column")


    # Grouping the Product Data
    # Aggregate the Product Data and Grouping by pg_categ_txt, pg_super_categ_txt
    df_agg = df.groupBy("pg_categ_txt", "pg_super_categ_txt").agg(
        expr(
            "CASE WHEN COUNT(DISTINCT pg_brand_txt) = 1 THEN MIN(pg_brand_txt) ELSE NULL END"
        ).alias("pg_brand_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_conc_txt) =1 THEN MIN( pg_conc_txt) ELSE NULL END"
        ).alias("pg_conc_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_grp_size_txt) =1 THEN MIN( pg_grp_size_txt) ELSE NULL END"
        ).alias("pg_grp_size_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_mfgr_txt) =1 THEN MIN( pg_mfgr_txt) ELSE NULL END"
        ).alias("pg_mfgr_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_mega_categ_txt) =1 THEN MIN( pg_mega_categ_txt) ELSE NULL END"
        ).alias("pg_mega_categ_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_sectr_txt) =1 THEN MIN( pg_sectr_txt) ELSE NULL END"
        ).alias("pg_sectr_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_seg_txt) =1 THEN MIN( pg_seg_txt) ELSE NULL END"
        ).alias("pg_seg_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_sku_base_desc_txt) =1 THEN MIN( pg_sku_base_desc_txt) ELSE NULL END"
        ).alias("pg_sku_base_desc_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_sku_num_txt) =1 THEN MIN( pg_sku_num_txt) ELSE NULL END"
        ).alias("pg_sku_num_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_sub_brand_txt) =1 THEN MIN( pg_sub_brand_txt) ELSE NULL END"
        ).alias("pg_sub_brand_txt"),
        expr(
            "CASE WHEN COUNT(DISTINCT pg_sub_mfgr_txt) =1 THEN MIN( pg_sub_mfgr_txt) ELSE NULL END"
        ).alias("pg_sub_mfgr_txt")
    )

    # Complementing df with df_agg
    df = df.unionByName(df_agg,allowMissingColumns=True)
    logger.info(f"Row Count after aggregation: {df.count()}")


    # Standardize the product data and add the key columns run_id and srce_sys_id
    df.createOrReplaceTempView("df")
    df = spark.sql(f"""
    SELECT 
    ref_df.prod_skid,
    in_df.* 
    FROM (
    SELECT 
        CASE 
        WHEN upc_txt IS NOT NULL AND pg_super_categ_txt IS NOT NULL 
            THEN upc_txt || ';' || pg_categ_txt || ';' || pg_super_categ_txt
        WHEN upc_txt IS NOT NULL AND pg_super_categ_txt IS NULL 
            THEN upc_txt || ';' || pg_categ_txt || ';NOT APPLICABLE'
        WHEN upc_txt IS NULL AND pg_super_categ_txt IS NOT NULL 
            THEN pg_categ_txt || ';' || pg_super_categ_txt
        ELSE pg_categ_txt 
        END AS extrn_prod_id,
        :run_id AS run_id,
        0 AS cntrt_id,
        :srce_sys_id AS srce_sys_id,
        :srce_sys_id AS part_srce_sys_id,
        0 AS part_cntrt_id,
        * 
    FROM df
    ) AS in_df
    LEFT OUTER JOIN (
    SELECT * 
    FROM {catalog_name}.internal_tp.tp_prod_sdim
    WHERE part_srce_sys_id = :srce_sys_id 
    ) AS ref_df
    ON in_df.extrn_prod_id == ref_df.extrn_prod_id  and in_df.srce_sys_id =ref_df.srce_sys_id
    """, {"srce_sys_id" : srce_sys_id, "run_id" : run_id })
    return df

@retry_with_backoff()
def t2_publish_product(df,catalog_name,schema_name, prod_dim, spark):
    df_connect_mm_prod_dim_vw = spark.sql(f"SELECT * FROM {catalog_name}.{schema_name}.{prod_dim} limit 0")
    df = df.unionByName(df_connect_mm_prod_dim_vw,allowMissingColumns=True)
    df.createOrReplaceTempView("df_mm_prod_sdim_promo_vw")

    merge_sql_sdim = f"""
    MERGE INTO {catalog_name}.{schema_name}.{prod_dim} tgt
    USING df_mm_prod_sdim_promo_vw src
    ON src.prod_skid = tgt.prod_skid
    AND src.part_srce_sys_id = tgt.part_srce_sys_id 
    WHEN MATCHED THEN
    UPDATE SET *
    WHEN NOT MATCHED THEN
    INSERT *
    """
    spark.sql(merge_sql_sdim)

@retry_with_backoff()
def release_semaphore(catalog_name,run_id,lock_path, spark ):
    lock_paths_sani=sanitize_filepath(lock_path)

    query = f"""
        DELETE FROM {catalog_name}.internal_tp.tp_run_lock_plc
        WHERE run_id = {run_id}
        AND lock_path IN ({lock_paths_sani})
    """

    print(query)
    spark.sql(query)   
   


def read_from_postgres(object_name, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ):

    """
    Reads data from a PostgreSQL table into a Spark DataFrame.
    
    Parameters:
    object_name (str): The name of the table to read from.
    
    Returns:
    DataFrame: A Spark DataFrame containing the data from the specified table.
    """
    df = spark.read.format("jdbc") \
        .option("driver", "org.postgresql.Driver") \
        .option("url", f"{ref_db_jdbc_url}/{ref_db_name}") \
        .option("dbtable", f'''{object_name}''') \
        .option("user", f"{ref_db_user}") \
        .option("password", f"{ref_db_pwd}") \
        .option("ssl", True) \
        .option("sslmode", "require") \
        .option("sslfactory", "org.postgresql.ssl.NonValidatingFactory") \
        .load()
    return df

def read_query_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):

    """
    Reads data from a PostgreSQL table using a query into a Spark DataFrame.
    
    Parameters:
    query (str): The SQL query to execute.
    
    Returns:
    DataFrame: A Spark DataFrame containing the data using the query.
    """
    df = spark.read.format("jdbc") \
        .option("driver", "org.postgresql.Driver") \
        .option("url", f"{ref_db_jdbc_url}/{ref_db_name}") \
        .option("query", query) \
        .option("user", ref_db_user) \
        .option("password", ref_db_pwd) \
        .option("ssl", True) \
        .option("sslmode", "require") \
        .option("sslfactory", "org.postgresql.ssl.NonValidatingFactory") \
        .load()
    return df

def write_to_postgres(df, object_name, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):

    """
    Writes data from a Spark DataFrame to a PostgreSQL table.
    
    Parameters:
    df (DataFrame): The Spark DataFrame to write.
    object_name (str): The name of the table to write to.
    """
    df.write.format("jdbc") \
        .option("url", f"{ref_db_jdbc_url}/{ref_db_name}") \
        .option("dbtable", f'''{object_name}''') \
        .option("user", f"{ref_db_user}") \
        .option("password", f"{ref_db_pwd}") \
        .mode("append") \
        .save()

def trunc_and_write_to_postgres(df, object_name, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):

    """
    Writes data from a Spark DataFrame to a PostgreSQL table.
    
    Parameters:
    df (DataFrame): The Spark DataFrame to write.
    object_name (str): The name of the table to write to.
    """
    df.write.format("jdbc") \
        .option("url", f"{ref_db_jdbc_url}/{ref_db_name}") \
        .option("dbtable", f'''{object_name}''') \
        .option("user", f"{ref_db_user}") \
        .option("password", f"{ref_db_pwd}") \
        .option("truncate", True) \
        .mode("overwrite") \
        .save()

def update_to_postgres(query,params, ref_db_name, ref_db_user, ref_db_pwd,ref_db_hostname):
    conn = psycopg2.connect(
        dbname=ref_db_name,
        user=ref_db_user,
        password=ref_db_pwd,
        host=ref_db_hostname,
        sslmode='require'
    )
    cursor = conn.cursor()
    cursor.execute(query,params)
    conn.commit()
    cursor.close()
    conn.close()

import os
import re
import zipfile
import shutil

def unzip(file_name, vol_path):
    """
    Unzips a file using the system 'unzip' command.
    - Removes .zip extension to create target folder
    - Deletes the folder if it already exists (overwrite)
    - Extracts only the files from inside the inner folder (flatten to root)
    """
    logger = get_logger()
    
    # Remove .zip extension to create target folder
    filename_wo_extn = re.sub(r'\.zip$', '', file_name, flags=re.IGNORECASE)
    target_dir = os.path.join(vol_path, filename_wo_extn)

    # Overwrite: delete folder if exists
    if os.path.exists(target_dir):
        logger.info(f"Target directory '{target_dir}' exists. Removing it for overwrite.")
        shutil.rmtree(target_dir)
    os.makedirs(target_dir, exist_ok=True)
    logger.info(f"Created target directory: {target_dir}")

    # Use system unzip command
    zip_path = os.path.join(vol_path, file_name)
    logger.info(f"Extracting '{zip_path}' to '{target_dir}' using system unzip...")
    subprocess.run(["unzip", zip_path, "-d", target_dir], check=True)
    logger.info("Extraction completed successfully.")

    # Flatten inner folder if present
    entries = os.listdir(target_dir)
    if len(entries) == 1 and os.path.isdir(os.path.join(target_dir, entries[0])):
        inner_folder = os.path.join(target_dir, entries[0])
        logger.info(f"Flattening inner folder: {inner_folder}")
        for item in os.listdir(inner_folder):
            shutil.move(os.path.join(inner_folder, item), target_dir)
        shutil.rmtree(inner_folder)
        logger.info("Inner folder flattened successfully.")

    logger.info(f"Files extracted to: {target_dir}")


def add_secure_group_key(df, cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ):
    logger = get_logger()

    # SQL query to fetch the secure_group_key for the given contract ID
    sgk_cntrt_id = f"""
    SELECT secure_group_key
    FROM {postgres_schema}.mm_cntrt_secur_grp_key_assoc_vw
    WHERE cntrt_id = '{cntrt_id}'
    """
    # Execute the query and collect the secure_group_key
    sgk_cntrt_id = read_query_from_postgres(sgk_cntrt_id, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ).collect() # type: ignore
    logger.info(f"secure_grp_key_id: {sgk_cntrt_id}")

    sgk_default = f"""
        SELECT secure_group_key
        FROM {postgres_schema}.mm_secur_grp_key_lkp
        WHERE secur_grp_key_id IS NULL
        """
    sgk_default = read_query_from_postgres(sgk_default, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ).collect() # type: ignore

    # Check if a secure_group_key was found for the given contract ID
    if sgk_cntrt_id and sgk_cntrt_id[0]['secure_group_key'] is not None:
        secure_group_key = sgk_cntrt_id[0]['secure_group_key']
        logger.info(f"special secure group key added: {secure_group_key}")
    else:
        # If no secure_group_key found for cntrt_id, fetch the default secure_group_key from the mm_secur_grp_key_lkp table
        secure_group_key = sgk_default[0]['secure_group_key']
        logger.info(f"default secure group key added :{secure_group_key}")
    df.createOrReplaceTempView("temp")
    query = """ select *except(secure_group_key),cast(:secure_group_key as long) as secure_group_key from temp"""
    df = spark.sql(query,{"secure_group_key": secure_group_key})

    return df

################################################################################

def work_to_arch(file_name, dbutils):
    spark = get_spark_session()
    paths=dbutils.fs.ls(f'{derive_base_path(spark)}/WORK') # type: ignore

    for filename in paths:
        if (file_name==filename.name): 
            print(filename.name)
            dbutils.fs.mv(f'{derive_base_path(spark)}/WORK/'+filename.name,f'{derive_base_path(spark)}/ARCH/'+filename.name) # type: ignore
            folder_name = file_name.strip()[:-4]
            dbutils.fs.rm(f'{derive_base_path(spark)}/WORK/'+folder_name, recurse=True)
            return True
    return False

def get_file_name(spark, raw_path, run_id, zip_file_name):
    dbutils = get_dbutils(spark)
    file_name = []
    for entry in dbutils.fs.ls(raw_path):
        if str(run_id) in entry.path and entry.isDir():
            for f in dbutils.fs.ls(entry.path):
                if fnmatch.fnmatch(f.name.lower(), zip_file_name.lower()):
                    file_name.append(f.name)

    if not file_name:
        raise FileNotFoundError("No file matched pattern '{}' under '{}'".format(zip_file_name, f"{raw_path}/{run_id}"))
    if len(file_name) > 1:
        raise ValueError("Multiple files matched pattern '{}' under '{}'".format(zip_file_name, f"{raw_path}/{run_id}"))

    return file_name[0]

###################################################################################
def load_file(file_type, run_id, cntrt_id, step_file_pattern, notebook_name, delimiter, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ):
    logger = get_logger()
    # Define paths for raw and light refined data
    base_path = derive_base_path(spark)
    raw_path = f"{base_path}/WORK/"
    
    # Read column mappings from parquet file
    df_dpf_col_asign_vw = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/column_mappings_df_dpf_col_asign_vw")
    
    # Prepare file name patterns
    s = step_file_pattern
    file_name_for_zip = s.replace('%', '*')
    file_name_for_zip = get_file_name(spark,raw_path,run_id,file_name_for_zip)
    
    logger.info(f"File to be loaded: {raw_path}/*{run_id}*/{file_name_for_zip}")

    df_extrn = spark.read.format('csv').option('header', True).option('delimiter', delimiter).load(f"{raw_path}/*{run_id}*/{file_name_for_zip}")

    if file_type == 'fact':
        df_extrn_raw = spark.read.format('csv').option('header', True).option('delimiter', delimiter).load(f'{raw_path}/*{run_id}*/{file_name_for_zip}*')
        count_df_extrn_raw=df_extrn_raw.count()
        #Add Row_count to mm_run_detl_plc
        add_row_count(count_df_extrn_raw,postgres_schema,run_id,cntrt_id,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    # Strip whitespace from column names
    for i in df_extrn.columns:
        df_extrn = df_extrn.withColumnRenamed(i, i.strip())

    if file_type == 'fact':
        for i in df_extrn_raw.columns:
            df_extrn_raw = df_extrn_raw.withColumnRenamed(i, i.strip())
    if file_type != 'fact':
        df_extrn_raw = df_extrn
    
    # Define list of columns based on file type
    if file_type == 'prod':
        list_cols = ['PROD', 'product', 'prod']
    elif file_type == 'mkt':
        list_cols = ['MKT', 'market', 'mkt']
    elif file_type == 'fact':
        list_cols = ['FACT', 'fact','FCT','fct']
    elif file_type == 'time':
        list_cols = ['TIME', 'time']
    else:
        list_cols = []

    # Filter column mappings based on contract ID and table type
    df_dpf_col_asign_vw = df_dpf_col_asign_vw.filter(col('cntrt_id') == cntrt_id).where(df_dpf_col_asign_vw['dmnsn_name'].rlike("|".join(["(" + column + ")" for column in list_cols])))
    df_dpf_col_asign_vw = df_dpf_col_asign_vw.select("file_col_name", "db_col_name").distinct()
    df_dpf_col_asign_vw.show(100)

    df_dpf_col_asign_vw1 = df_dpf_col_asign_vw

    # Create a temporary view for column assignments
    df_dpf_col_asign_vw.createOrReplaceTempView("col_asign")
    
    df_dpf_col_asign_vw = spark.sql('''
        SELECT t.file_col_name, array_join(collect_list(t.db_col_name), ',') AS db_col_name
        FROM (
            SELECT file_col_name, db_col_name
            FROM col_asign
            ORDER BY file_col_name, db_col_name
        ) t
        GROUP BY t.file_col_name
    ''')
    

    # Create a dictionary for column mappings
    dictionary = {row['file_col_name']: row['db_col_name'] for row in df_dpf_col_asign_vw.collect()}
    
    for key, value in dictionary.items():
        if "," in value:
            lst_values = value.split(",")
            for v in lst_values:
                df_extrn = df_extrn.withColumn(v, col(f'`{key}`'))
            df_extrn = df_extrn.drop(key)
        else:
            df_extrn = df_extrn.withColumnRenamed(key, value)

    # Drop extra columns not in the user-mapped columns
    df_cols = df_extrn.columns
    column = []
    for item in dictionary:
        if "," in dictionary[item]:
            lst_items = dictionary[item].split(",")
            for v in lst_items:
                column.append(v)
        else:
            column.append(dictionary[item])
    column = [n for n in column if len(n) != 0]

    drop_cols = [col for col in df_cols if col not in column]
    df_extrn = df_extrn.drop(*drop_cols)
    lst_drop_row_cols = df_extrn.columns
    
    # Drop rows with all measure columns null for non-fact files
    if file_type != 'fact':

        df_measr=read_query_from_postgres(f"SELECT *  FROM {postgres_schema}.mm_measr_lkp", spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
        
        lst_measrs = [row['measr_phys_name'] for row in df_measr.select('measr_phys_name').distinct().collect()]
        drop_null_measr_cols = [col for col in lst_drop_row_cols if col in lst_measrs]
        df_extrn = df_extrn.dropna(thresh=len(drop_null_measr_cols), subset=(drop_null_measr_cols))

    # Handle columns with special characters
    df1 = df_dpf_col_asign_vw1.filter("db_col_name like '%#%'").select("db_col_name")
    df2 = df1.withColumn("db_col_name", F.regexp_replace(F.col("db_col_name"), "[0-9#\\s]", "")).distinct()

    l1 = df1.select("db_col_name").collect()
    l1 = [row["db_col_name"] for row in l1]
    l2 = df2.select("db_col_name").collect()
    l2 = [row["db_col_name"] for row in l2]

    for z in l2:
        lst = []
        for x in l1:
            if x.startswith(z):
                lst.append(x)
        lst_cols1 = sorted(lst)

        for d in lst_cols1:
            
            df_extrn = df_extrn.withColumn(d, when(col(d).isNull(), '').otherwise(col(d)))

            if file_type == 'mkt':
                df_extrn =df_extrn.withColumn(z, regexp_replace(regexp_replace(concat_ws(':',*lst_cols1),'^:+|:+$' ,'' ),':+',':'))
                
        if file_type != 'mkt':
            df_extrn = df_extrn.withColumn(z, concat_ws(':', *lst_cols1))

        df_extrn = df_extrn.drop(*lst_cols1)
    df_extrn = df_extrn.drop(*l1)

    # Convert column names to lowercase
    for i in df_extrn.columns:
       df_extrn = df_extrn.withColumnRenamed(i, i.lower())

    # Remove duplicate rows for non-fact files
    if file_type =='fact':
        df_extrn = df_extrn.withColumn('extrn_mkt_id', regexp_replace(regexp_replace(col('extrn_mkt_id'),'^:+|:+$' ,'' ),':+',':'))

    else:
        df_extrn = df_extrn.distinct()

    # Save the processed DataFrame to parquet files
  
 
    df_extrn.write.mode("overwrite").format('parquet').save(f"{materialise_path(spark)}/{run_id}/{notebook_name}_df_{file_type}_extrn")
    logger.info(f"Write completed to path: {materialise_path(spark)}/{run_id}/{notebook_name}_df_{file_type}_extrn")
    df_extrn_raw.write.mode("overwrite").format('parquet').save(f"{materialise_path(spark)}/{run_id}/{notebook_name}_df_{file_type}_extrn_raw")
    logger.info(f"Write completed to path: {materialise_path(spark)}/{run_id}/{notebook_name}_df_{file_type}_extrn_raw")
    if file_type == 'fact':
        df_extrn.write.mode("overwrite").format('parquet').save(f"{materialise_path(spark)}/{run_id}/{notebook_name}_df_{file_type}_extrn_dvm")
        logger.info(f"Write completed to path: {materialise_path(spark)}/{run_id}/{notebook_name}_df_{file_type}_extrn_dvm")
    
    return 'Success'

############################################################################


def t2_load_file(cntrt_id, dmnsn_name, file_type, run_id, notebook_name, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    logger = get_logger()
    # Filter the DataFrame to get the vendor file pattern and step file pattern for the specified contract ID
    df_cntrt_lkp= load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).select('file_patrn').collect()
    vendor_file_pattern = df_cntrt_lkp[0].file_patrn if df_cntrt_lkp else None
    logger.info(f"Vendor File Pattern: {vendor_file_pattern}")

    df_cntrt_file_lkp= load_cntrt_file_lkp(cntrt_id, dmnsn_name, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).select('file_patrn').collect()
    step_file_pattern = df_cntrt_file_lkp[0].file_patrn if df_cntrt_file_lkp else None
    logger.info(f"Step File Pattern: {step_file_pattern}")

    # Define the query to get the file delimiter from the PostgreSQL database
    df_cntrt_dlmtr_lkp= load_cntrt_dlmtr_lkp(cntrt_id, dmnsn_name, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).collect()
    delimiter = df_cntrt_dlmtr_lkp[0].dlmtr_val if df_cntrt_dlmtr_lkp else None
    logger.info(f"File Delimiter: {delimiter}")

    missing = []
    if not step_file_pattern:
        missing.append("step_file_pattern")
    if not delimiter:
        missing.append("delimiter")
    if not vendor_file_pattern:
        missing.append("vendor_file_pattern")

    if missing:
        raise AttributeError(f"Missing: {', '.join(missing)} for {cntrt_id}")

    # Use the UDF to Load the file using the specified parameters
    load_file(file_type,run_id,cntrt_id,step_file_pattern,notebook_name, delimiter, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

def semaphore_generate_path(dim_type, srce_sys_id, cntrt_id):
    if dim_type in ('TP_MKT_SDIM','TP_MKT_DIM'):
        mkt_path = [
            f"/mnt/tp-publish-data/{dim_type}/part_srce_sys_id={srce_sys_id}"
            ]
        return mkt_path
    
    elif dim_type in ('TP_PROD_SDIM','TP_PROD_DIM') and srce_sys_id!=3:
        prod_path = [
            f"/mnt/tp-publish-data/{dim_type}/part_srce_sys_id={srce_sys_id}"
            ]
        return prod_path
        
    elif dim_type in ('TP_PROD_SDIM','TP_PROD_DIM') and srce_sys_id==3:
        prod_path = [
            f"/mnt/tp-publish-data/{dim_type}/part_srce_sys_id={srce_sys_id}/part_cntrt_id={cntrt_id}"
            ]
        return prod_path
    
    # RUN_PLC PATHS
    elif dim_type in ('TP_RUN_MKT_PLC','TP_RUN_PROD_PLC','TP_RUN_TIME_PERD_PLC','TP_RUN_MEASR_PLC','TP_RUN_PRTTN_PLC'):
        run_plc_path = [
            f"/mnt/tp-publish-data/{dim_type}/part_srce_sys_id={srce_sys_id}/part_cntrt_id={cntrt_id}"
            ]
        return run_plc_path
    
    elif dim_type in ('mkt_skid_seq','prod_skid_seq'):
        skid_path = [
            f"/mnt/tp-publish-data/{dim_type}"
            ]
        return skid_path
    
    elif dim_type in ('TP_VALDN_AGG_FCT'):
        valdn_agg_fct_path = [
            f"/mnt/tp-publish-data/{dim_type}/part_cntrt_id={cntrt_id}"
            ]
        return valdn_agg_fct_path

    elif dim_type in ('TP_VALDN_PROD_DIM'):
        valdn_prod_dim_path = [
            f"/mnt/tp-publish-data/{dim_type}/part_srce_sys_id={srce_sys_id}/part_cntrt_id={cntrt_id}"
            ]
        return valdn_prod_dim_path


def semaphore_queue(run_id,paths, catalog_name, spark):
  
    check_path = ', '.join(f"'{i}'" for i in paths)
    # Create a DataFrame with the paths
    paths_df = spark.createDataFrame([Row(lock_path=path) for path in paths])
    
    # Add additional columns to the DataFrame
    paths_df = paths_df.withColumn("run_id", lit(run_id).cast("bigint")) \
                       .withColumn("lock_sttus", lit(False)) \
                       .withColumn("creat_date", current_timestamp())
    paths_df.createOrReplaceTempView("paths_df")
    # Insert data into the tp_run_lock_plc table
    table_name = f"{catalog_name}.internal_tp.tp_run_lock_plc"
    #retry 
    safe_write_with_retry(paths_df, table_name, "append",["run_id","lock_path"])

    print("Data inserted successfully into", table_name)
    return check_path

#############################################################################

def check_lock(run_id, check_path, catalog_name, spark):
    # Check for existing locks in the tp_run_lock_plc table
    df = spark.sql(f"""
        SELECT * 
        FROM paths_df curr 
        JOIN (
            SELECT run_id, lock_path 
            FROM {catalog_name}.internal_tp.tp_run_lock_plc 
            QUALIFY ROW_NUMBER() OVER(PARTITION BY lock_path ORDER BY creat_date)=1
        ) tbl 
        ON curr.run_id != tbl.run_id AND tbl.lock_path = curr.lock_path
    """)

    # If no locks are found, update the lock status to true
    if df.count() == 0:
        spark.sql(f"""
            UPDATE {catalog_name}.internal_tp.tp_run_lock_plc 
            SET lock_sttus = true 
            WHERE run_id = {run_id} AND lock_path IN ({check_path})
        """)
        print("Semaphore Acquired by the current process")
        return check_path
    else:
        # Wait for the lock to be released and retry
        print('Waiting for lock')
        time.sleep(30)
        return check_lock(run_id, check_path,catalog_name, spark)
    
###############################################################################

def semaphore_acquisition(run_id, paths, catalog_name, spark):
    # Acquire semaphore by queuing and checking locks
    check_path = semaphore_queue(run_id, paths, catalog_name, spark)
    check_path = check_lock(run_id, check_path, catalog_name, spark)
    return check_path

################################################################################

def assign_skid(df, run_id, skid_type, catalog_name,spark,srce_sys_id, cntrt_id):
    logger = get_logger()
    try:

        # Convert the skid_type to lowercase for consistency
        skid_type = skid_type.lower()

        # Proceed only if the type is either 'prod' or 'mkt'
        if skid_type in ('prod', 'mkt'):
            # Register the input DataFrame as a temporary SQL view
            df.createOrReplaceTempView('input_df')

            # Store original column names for later use
            cols = df.columns

            # Select run_id and external ID (based on skid_type) as key
            query = f"""
                SELECT CAST(run_id AS BIGINT), extrn_{skid_type}_id AS key 
                FROM input_df
            """

            df_sel = spark.sql(query)

            # Check if the run_id already exists in the target table
            query = f"""
                SELECT * FROM {catalog_name}.internal_tp.tp_{skid_type}_skid_seq 
                WHERE run_id = :run_id_param
                LIMIT 1
            """
            if spark.sql(query,{"run_id_param": run_id }).count() == 0:
                # Define the path to the PROD_SDIM data based on source system and contract ID
                dim_type=f"{skid_type}_skid_seq"
                print(dim_type)
                skid_path = semaphore_generate_path(dim_type, srce_sys_id, cntrt_id)

                print(skid_path)

                # Acquire semaphore to ensure safe access to the PROD_SDIM path for the given run ID
                skid_check_path = semaphore_acquisition(run_id, skid_path,catalog_name,spark)
                try:
                    print(f'Started writing to tp_{skid_type}_skid_seq')
                    # If not, append the selected data to the target table
                    safe_write_with_retry(df_sel, f'{catalog_name}.internal_tp.tp_{skid_type}_skid_seq', 'append')

                finally:
                    # Release the semaphore for the current run
                    release_semaphore(catalog_name,run_id,skid_check_path, spark)
                    logger.info(f"Semaphore released for run_id={run_id} and paths={skid_check_path}")

            # Read the skid mapping for the given run_id
            read_query = f"""
                SELECT {skid_type}_skid, run_id, key 
                FROM {catalog_name}.internal_tp.tp_{skid_type}_skid_seq 
                WHERE run_id = :run_id_param
            """
            df_sel = spark.sql(read_query,{"run_id_param": run_id })

            # Register both input and skid DataFrames as temporary views
            df.createOrReplaceTempView('input_df')
            df_sel.createOrReplaceTempView('skid_df')

            # Join input data with skid mapping to assign skid values
            query = f"""
                SELECT a.* EXCEPT(a.{skid_type}_skid), b.{skid_type}_skid 
                FROM input_df a 
                JOIN skid_df b 
                ON a.extrn_{skid_type}_id = b.key 
                AND a.run_id = b.run_id
            """
            df = spark.sql(query).select(*cols)

            return df

    except Exception as e:
        print(f"Error occurred: {e}")
        raise


#####################################################################

def cdl_publishing(logical_table_name,physical_table_name,unity_catalog_table_name,partition_definition_value, dbutils,configuration,metapsclient):
    config = configuration.load_for_default_environment_notebook(dbutils=dbutils)
    meta_client = metapsclient.configure(config).get_client()
    p_file_type_code = "delta"
    # get the tables from the config and iterate over them calling the publish_table method
    meta_client \
    .mode(publish_mode="update") \
    .publish_table(
        logical_table_name=logical_table_name,     #TP_WK_FCT
        physical_table_name=physical_table_name,    #TP_WK_FCT
        unity_catalog_table_name=unity_catalog_table_name,   #TP_WK_FCT
        unity_catalog_schema="gold_tp", 
        data_signals=[6888],
        partition_definition_value=partition_definition_value,
        file_type_code=p_file_type_code,
        data_type_code="TP",
        data_provider_code="TP",
        secure_group_key=0,
        test_only=False
        )
    # start the publishing process, which will trigger the publication of all tables defined above 
    meta_client.start_publishing()

def load_time_exprn_id(time_exprn_id,postgres_schema,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    # Construct the SQL query to select relevant columns from the time expression lookup table
    query = f"{postgres_schema}.mm_time_exprn_lkp"
    # Execute the query and read the result into a DataFrame
    df_time_exprn = read_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"time_exprn_id= {time_exprn_id}").select('start_date_val', 'end_date_val', 'time_perd_type_val')
    return df_time_exprn

def validation(catalog_name,value,validation_name,run_id,spark,validation_type):

    logger = get_logger()
    
    if isinstance(value, str): 
        df_val = spark.sql(value) #Check if the value is of type str
        # Need to perform checks for validation conditions
        validation_name_check = validation_name.lower()

        if validation_name_check == 'time period of weeks not properly generated':
            df_val = df_val.filter(lower(col("STATUS")).like("%missing%"))
        elif validation_name_check == 'missing/delivered areas':
            df_val = df_val.filter(lower(col("LAST_RUN_STATUS")).like(f"%missing%"))
        elif validation_name_check == 'missing/delivered hierarchies product' or validation_name_check == 'missing/delivered measures':
            df_val = df_val.filter(
                (lower(col("LAST_RUN_STATUS")).like("%missing%")) | (lower(col("LAST_PERIOD_STATUS")).like("%missing%"))
            )
        elif validation_name_check == 'missing/delivered product level':
            df_val = df_val.filter(
                (lower(col("sttus_vs_last_time_perd")).like("%new%")) |
                (lower(col("sttus_vs_last_time_perd")).like("%missing%"))
            )
    else: 
        df_val = value #check if the value of type dataframe, eg: Bad Fact Data Format (FYI Validations), value is of type DataFrame
        
    # Validation Check for Validation Report
    if df_val.count()==0:
        dq_val=(int(run_id),validation_name, 'PASSED', '',validation_type,'')
    else:
        print(df_val.count(), "-- DF COUNT")
        tp_tab_name_df = spark.sql(f"""select * from {catalog_name}.internal_tp.tp_tab_name where COL_NAME = '{validation_name}'""")
        hyperlink_name = tp_tab_name_df.collect()[0]['TAB_NAME']
        print(hyperlink_name, " --hyperlink name")
        light_refined_path = derive_publish_path(spark)
        cleaned_validation_type = validation_type.replace(" ", "_")
        hyperlink_val = f'=HYPERLINK("#\'{hyperlink_name}\'!A1","click_here")'
        path=f'{light_refined_path}/tp_dvm_rpt/{run_id}/{cleaned_validation_type}/{hyperlink_name}.parquet'
        df_val.write.format("parquet").options(header=True).mode('overwrite').save(path)
        dq_val=(int(run_id),validation_name, 'FAILED', hyperlink_val,validation_type,path)
    return dq_val

def dq_query_retrieval(catalog_name,validation_name,tier,description,spark):
    
    df = spark.sql(f"select * from {catalog_name}.internal_tp.tp_valdn_detl_lkp ")
    df = df.filter(f" valdn_grp_name = '{validation_name}' and valdn_name = '{description}' and data_tier = '{tier}'")
    return df

def dq_append_rows(results,output_rows):
    # output_rows = []
    # Format rows for grouped display by validation_type
    current_type = None
    for row in results:
        validation_type = row[3]  # Access validation_type using index
        if validation_type != current_type:
            current_type = validation_type
            # Insert group header
            output_rows.append(
                Row(Validation_Type="", Validation=current_type, Result="", Details="")
            )
        # Append actual data
        output_rows.append(
            Row(Validation_Type="", Validation=row[0], Result=row[1], Details=row[2])
        )
    return output_rows

def time_period_of_weeks_not_generated(run_id, cntrt_id, file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema):
    
    # Fetch contract lookup details from PostgreSQL
    df_cntrt_lkp = load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    row = df_cntrt_lkp.collect()[0]
    cntrt_code=row.cntrt_code
    country_name=row.cntry_name
    category_name=row.categ_name
    srce_sys_id = row.srce_sys_id
    time_perd_type_code = row.time_perd_type_code
    tier1_vendr_id = row.vendr_id
    df_cntrt_lkp.createOrReplaceTempView('mm_cntrt_lkp')

    time_perd_class_code = time_perd_class_codes(cntrt_id,catalog_name, postgres_schema,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    tier1_time_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Time_df_srce_mtime")
    tier1_time_mtrlz_tbl.createOrReplaceTempView('tier1_time_mtrlz_tbl')
    mm_time_perd_id_lkp=read_from_postgres(f'{postgres_schema}.mm_time_perd_id_lkp', spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_time_perd_id_lkp.createOrReplaceTempView('mm_time_perd_id_lkp')
    # mm_time_perd_id_lkp

    mm_time_perd_fdim = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_fdim")
    mm_time_perd_fdim.createOrReplaceTempView('mm_time_perd_fdim')
    # mm_time_perd_fdim
    
    tp_run_time_perd_plc = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_run_time_perd_plc")
    tp_run_time_perd_plc.createOrReplaceTempView('mm_run_time_perd_plc')
    # mm_run_time_perd_plc
    
    dpf_all_run_vw=read_from_postgres(f'{postgres_schema}.mm_run_plc', spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    dpf_all_run_vw.createOrReplaceTempView('dpf_all_run_vw')
    # dpf_all_run_vw
    
    mm_time_perd=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_fdim')
    mm_time_perd.createOrReplaceTempView('mm_time_perd')
    
    mm_time_perd_assoc=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_assoc')
    mm_time_perd_assoc.createOrReplaceTempView('mm_time_perd_assoc')
    
    mm_time_perd_assoc_type=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_assoc_type')
    mm_time_perd_assoc_type.createOrReplaceTempView('mm_time_perd_assoc_type')

    df_mm_time_perd_assoc_tier1_vw = mm_time_perd_assoc_tier1_vw(spark, mm_time_perd_assoc, mm_time_perd_fdim, mm_time_perd_assoc_type)
    df_mm_time_perd_assoc_tier1_vw.createOrReplaceTempView("mm_time_perd_assoc_tier1_vw")
    # mm_time_perd
    if time_perd_class_code == 'WK':
        df = dq_query_retrieval(catalog_name,'Reference Data Validations','Tier1','Time period of weeks not properly generated',spark)
        
    else:
        df = dq_query_retrieval(catalog_name,'Reference Data Validations','Tier1','Time period of weeks not properly generated not week',spark)

    query = df.collect()[0]['qry_txt']
    
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    return df_val

# Function to generate validation report
def generate_report(run_id, cntrt_id, file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema):
    
    logger = get_logger()
    logger.info("Starting the report generation")
    # Define paths for workspace and volume
    WORKSPACE_PATH=f'/Workspace/Application_Pipelines/Validation_Reports/tp_dvm_rprt_{run_id}.xlsx'

    # Fetch contract lookup details from PostgreSQL
    df_cntrt_lkp = load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    row = df_cntrt_lkp.collect()[0]
    cntrt_code=row.cntrt_code
    country_name=row.cntry_name
    category_name=row.categ_name
    srce_sys_id = row.srce_sys_id
    time_perd_type_code = row.time_perd_type_code
    tier1_vendr_id = row.vendr_id

    time_perd_class_code = time_perd_class_codes(cntrt_id,catalog_name, postgres_schema,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    # Create Excel writer object
    writer = pd.ExcelWriter(WORKSPACE_PATH, engine="xlsxwriter")
    
    df_report_creation = spark.sql(f"""
                    SELECT valdn_name,reslt,hyperlink_txt,valdn_type
                    FROM {catalog_name}.internal_tp.tp_valdn_rprt 
                    WHERE run_id=:run_id
                    QUALIFY (ROW_NUMBER() OVER(PARTITION BY valdn_name ORDER BY creat_time DESC) =1)
                    """, {'run_id':run_id}
            )
    # display(df_report_creation)

    list_df = df_report_creation.agg(collect_list("valdn_type").alias("valdn_types_list"))

    # Extract the list from the DataFrame
    valdn_name_list = list_df.collect()[0]["valdn_types_list"]
    # logger.info(valdn_name_list)
    
    DQ_Validation_tier1_list = ['File Structure Validation','Reference Data Vendors Validations','For Your Information Validations','Reference Data Validations','Business Validations']
    
    DQ_Validation_tier2_list = ['File Structure Validation','Reference Data Validations','For Your Information Validations','Business Validation']
    output_rows = []
    data_pds =[]
    empty_schema = StructType([])
    # Create an empty DataFrame
    redeliverd_pds = spark.createDataFrame([], schema=empty_schema)
    reference_data_validation_switch = 0
    if srce_sys_id == 3:
        for dq_item in DQ_Validation_tier1_list:
            logger.info(f"----dq_item---- {dq_item}")
            logger.info(f"----data_pds---- {data_pds}")
            
            if dq_item in valdn_name_list:
                df_report_creation_new = df_report_creation.select("*").filter(col("valdn_type") == dq_item)            
                results = df_report_creation_new.collect()
                output_rows =dq_append_rows(results,output_rows)
            if dq_item == 'Reference Data Validations':

                logger.info("--generating reference data validations--")
                reference_data_validation_switch = 1
                # # ------------------Time period of weeks not properly generated Validation----------------
                df_val = time_period_of_weeks_not_generated(run_id, cntrt_id, file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
                logger.info(f"time_perd_class_code is {time_perd_class_code}")
                # New Periods
                new_pds = []
                redeliverd_pds=[]
                if time_perd_class_code == 'WK':
                    
                    for i in df_val.filter("STATUS = 'new'").select('WK_NAME').collect():
                        new_pds.append(i[0])
                    
                    df_time_period_wks_nt_gen=df_val.filter("STATUS = 'redelivered'").select('WK_NAME')
                    rows_list = [row.asDict() for row in df_time_period_wks_nt_gen.collect()]
                    for i in rows_list:
                        redeliverd_pds.append(i['WK_NAME'])
                else:
                    
                    for i in df_val.filter("STATUS = 'new'").select('TIME_PERD_NAME').collect():
                        new_pds.append(i[0])
                    df_time_period_wks_nt_gen=df_val.filter("STATUS = 'redelivered'").select('TIME_PERD_NAME')
                    rows_list = [row.asDict() for row in df_time_period_wks_nt_gen.collect()]
                    for i in rows_list:
                        redeliverd_pds.append(i['TIME_PERD_NAME'])
                redeliverd_pds=','.join(redeliverd_pds)
                ds_redeliverd_pds = ('Redelivered Periods', redeliverd_pds)
                ds_new_pds = ('New Periods', new_pds)
                
                data_pds.append(ds_new_pds)
                ##Prepare Redelivered Periods Tab
                
                schema_for_vld_dtls = StructType([ 
                    StructField("redeliverd_pds",StringType(),True)
                ])
                data_rp=[]
                data_rp.append(ds_redeliverd_pds)
                schema_for_vld_dtls = StructType([ 
                    StructField("Validation",StringType(),True),
                    StructField("Details",StringType(),True)
                ])
                df_redv_pd = spark.createDataFrame(data_rp, schema_for_vld_dtls)
                df_redv_pd=df_redv_pd.withColumn("redlv_pds", split(df_redv_pd["Details"], ","))
                df_redv_pd =df_redv_pd.select(explode(df_redv_pd["redlv_pds"]).alias("REDELIVERD_PDS"))
                redeliverd_pds = df_redv_pd.toPandas()
    else:
        for dq_item in DQ_Validation_tier2_list:
            if dq_item in valdn_name_list:
                df_report_creation_new = df_report_creation.select("*").filter(col("valdn_type") == dq_item)            
                results = df_report_creation_new.collect()
                output_rows =dq_append_rows(results,output_rows)
    
    formatted_df = spark.createDataFrame(output_rows)

    # Validation Details Tab Generation and stored to excel
    CET = pytz.timezone("CET")
    datetime_cet = datetime.now(CET)
    ts = datetime_cet.strftime("%Y:%m:%d %H:%M:%S") + " CET"

    # Prepare metadata for validation details tab
    if formatted_df.filter(" Result = 'FAILED' ").count() > 0:
        validation_result = "FAILED"
    else:
        validation_result = "PASSED"
    overall_validation_result = ("Overall Validation Result", validation_result)
    ds_validation_report_time = ("Validation Report Time", ts)

    ds_file_name = ("File name for given report", file_name)
    ds_file_prefix = ("File prefix", cntrt_code)
    ds_dlvry_id = ("Delivery ID", run_id)
    ds_category = ("Category", category_name)
    ds_cntry = ("Country", country_name)

    # Compile metadata into a list
    data_dvm_dtls = []
    data_dvm_dtls.append(overall_validation_result)
    data_dvm_dtls.append(ds_validation_report_time)
    data_dvm_dtls.append(ds_file_name)
    data_dvm_dtls.append(ds_file_prefix)
    data_dvm_dtls.append(ds_dlvry_id)
    data_dvm_dtls.append(ds_category)
    data_dvm_dtls.append(ds_cntry)

    # Define schema for validation details
    schema_for_vld_dtls = StructType(
        [
            StructField("Validation", StringType(), True),
            StructField("Details", StringType(), True),
        ]
    )
    # Creation of Validation Details Tab
    if data_pds:
        df_overall_result = spark.createDataFrame(data_dvm_dtls, schema_for_vld_dtls)
        df_period_results = spark.createDataFrame(data_pds, schema_for_vld_dtls)
        df_vld_dtls = df_overall_result.unionByName(df_period_results, True)
    else:
        df_vld_dtls = spark.createDataFrame(data_dvm_dtls, schema_for_vld_dtls)
    vldtn_dtls = df_vld_dtls.toPandas()
    vldtn_dtls.to_excel(writer, sheet_name="VALIDATION_DETAILS", index=False)

    # Summary Tab Data stored to Excel
    SUMMARY = formatted_df.select(col("Validation"), col("Result"), col("Details"))
    SUMMARY_DF = SUMMARY.toPandas()
    SUMMARY_DF.to_excel(writer, sheet_name="SUMMARY", index=False)

    print(reference_data_validation_switch, "---reference_data_validation_switch")

    if srce_sys_id ==3 and reference_data_validation_switch == 1:
        if redeliverd_pds['REDELIVERD_PDS'].dropna().apply(lambda x: isinstance(x, str) and x.strip() != '').any():
        
            print(redeliverd_pds.head(10), "---redeliverd_pds")

            redeliverd_pds.to_excel(writer,sheet_name="REDELIVERED_PERDS",index=False)
    
    # Other Tab Data stored to Excel
    # Fetch failed validations for detailed tab creation
    df_tap_report_creation = spark.sql(f"""
                            SELECT *
                             FROM {catalog_name}.internal_tp.tp_valdn_rprt 
                             WHERE run_id=:run_id AND reslt = 'FAILED' 
                            """, {'run_id':run_id}
    )
    df_tab_name = spark.sql(f"""SELECT * FROM {catalog_name}.internal_tp.tp_tab_name""")

    # Iterate through each failed validation path
    for row in df_tap_report_creation.select('valdn_path','valdn_name').collect():
        file_path = row['valdn_path']
        df_tab_creation = spark.read.parquet(file_path)
        valdn_name = row['valdn_name']
        logger.info(f"valdn_name is {valdn_name}")

        if df_tab_creation.count() > 0:

            # # Join with tab name to get sheet name
            tab_name_value = df_tab_name.where(df_tab_name.COL_NAME == valdn_name).select("TAB_NAME").collect()[0][0]
            logger.info(f"tab_name_value is {tab_name_value}")
            
            df_tab_creation.show()
            # Write tab data to Excel
            df = df_tab_creation.toPandas().astype(str)
            tab = tab_name_value
            df.to_excel(writer, sheet_name=f"{tab}", index=False)

    writer.close()
    
    logger.info(f"Formatted workbook saved to {WORKSPACE_PATH}")
    SAFE_BASE = "/Workspace/Application_Pipelines/Validation_Reports"
    WORKSPACE_PATH = f"{SAFE_BASE}/tp_dvm_rprt_{run_id}.xlsx"
    DSTN_BASE = f"/Volumes/{catalog_name}/internal_tp/tp-source-data/Validation_Reports"
    VOLUME_PATH = f'{DSTN_BASE}/tp_dvm_rprt_{run_id}.xlsx'
    
    src_path = os.path.abspath(WORKSPACE_PATH)
    dst_path = os.path.abspath(VOLUME_PATH)
    safe_base = os.path.abspath(SAFE_BASE)
    dstn_base = os.path.abspath(DSTN_BASE)
    
    if src_path.startswith(safe_base) and dst_path.startswith(dstn_base):
        copyfile(WORKSPACE_PATH, VOLUME_PATH)
    logger.info(f"Report written to WORKSPACE_PATH {WORKSPACE_PATH}")
    logger.info(f"WORKSPACE_PATH copied to WORKSPACE_PATH {VOLUME_PATH}")
    
    logger.info("Report generation completed successfully")

# Function to apply formatting to the validation report
def generate_formatting_report(run_id,catalog_name):
    logger = get_logger()
    logger.info("Starting formatting of the report.")
    # Define paths for workspace and volume
    WORKSPACE_PATH=f'/Workspace/Application_Pipelines/Validation_Reports/tp_dvm_rprt_{run_id}.xlsx'
    VOLUME_PATH = f'/Volumes/{catalog_name}/internal_tp/tp-source-data/Validation_Reports/tp_dvm_rprt_{run_id}.xlsx'
    
    # Load the workbook from volume path
    wb = openpyxl.load_workbook(VOLUME_PATH)
    logger.info(f"{wb.sheetnames},--wb.sheetnames")
    # Check for Redelivered Perd Sheet
    if 'REDELIVERED_PERDS' in wb.sheetnames:
        
        # Apply formatting for redelivered periods 
        ws_summary = wb['REDELIVERED_PERDS']
        ws_summary['A1'].font = Font(color='000000', bold=True, size=12)
        ws_summary['A1'].fill = PatternFill('solid', start_color='8DB4E2')

 
    # Access the SUMMARY sheet
    ws_summary = wb['SUMMARY']
	
    # Apply header formatting to SUMMARY sheet
    ws_summary['A1'].font = Font(color = '000000',bold=True, size=12) ## black
    ws_summary['B1'].font = Font(color = '000000',bold=True, size=12) ## black
    ws_summary['C1'].font = Font(color = '000000',bold=True, size=12) ## black
    ws_summary['A1'].fill = PatternFill('solid', start_color="8DB4E2") ## blue
    ws_summary['B1'].fill = PatternFill('solid', start_color="8DB4E2") ## blue
    ws_summary['C1'].fill = PatternFill('solid', start_color="8DB4E2") ## blue
	
    # Apply conditional formatting for PASSED results
	
    fill = PatternFill(
		start_color='C6EFCE',
		end_color='C6EFCE',fill_type='solid') # specify background color
	
    ws_summary.conditional_formatting.add(
		'B2:B16594', CellIsRule(operator='equal', formula=['"PASSED"'], fill=fill)) # include formatting rule
	
    # Apply conditional formatting for FAILED results
    fill_failed = PatternFill(
		start_color='FFC7CE',
		end_color='FFC7CE',fill_type='solid') # specify background color
	
    ws_summary.conditional_formatting.add(
		'B2:B16594', CellIsRule(operator='equal', formula=['"FAILED"'], fill=fill_failed)) # include formatting rule
	
    fill_val_highlight = Font(color = '000000',bold=True, size=12) ## black
    excel_cells_limit = 'A2:A16594'
    
    ws_summary.conditional_formatting.add(
		excel_cells_limit, CellIsRule(operator='equal', formula=['"File Structure Validation"'], font=fill_val_highlight)) # include formatting rule
    
    ws_summary.conditional_formatting.add(
		excel_cells_limit, CellIsRule(operator='equal', formula=['"Reference Data Vendors Validations"'], font=fill_val_highlight)) # include formatting rule
    ws_summary.conditional_formatting.add(
		excel_cells_limit, CellIsRule(operator='equal', formula=['"For Your Information Validations"'], font=fill_val_highlight)) # include formatting rule
    ws_summary.conditional_formatting.add(
		excel_cells_limit, CellIsRule(operator='equal', formula=['"Reference Data Validations"'], font=fill_val_highlight)) # include formatting rule
    ws_summary.conditional_formatting.add(
		excel_cells_limit, CellIsRule(operator='equal', formula=['"Business Validations"'], font=fill_val_highlight)) # include formatting rule
	
	# Auto Size columns from all the workbooks
    print(wb)
    for sheet_name in wb.sheetnames:
        dims = {}
        for row in wb[sheet_name].rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = builtins.max(dims.get(cell.column, 0), len(str(cell.value)))
        for col, value in dims.items():
            wb[sheet_name].column_dimensions[get_column_letter(col)].width = value+2
	
	
	# Updating Validation Details
    ws_vd = wb['VALIDATION_DETAILS']
    ws_vd['A1'].font = Font(color = '000000',bold=True, size=12) ## black
    ws_vd['B1'].font = Font(color = '000000',bold=True, size=12) ## black
	
    ws_vd['A1'].fill = PatternFill('solid', start_color="8DB4E2") ## blue
    ws_vd['B1'].fill = PatternFill('solid', start_color="8DB4E2") ## blue
	
    wb.save(WORKSPACE_PATH)
    logger.info(f"Formatted workbook saved to {WORKSPACE_PATH}")
    GR_SAFE_BASE = "/Workspace/Application_Pipelines/Validation_Reports"
    GR_WORKSPACE_PATH = f"{GR_SAFE_BASE}/tp_dvm_rprt_{run_id}.xlsx"
    GR_DSTN_BASE = f"/Volumes/{catalog_name}/internal_tp/tp-source-data/Validation_Reports"
    GR_VOLUME_PATH = f'{GR_DSTN_BASE}/tp_dvm_rprt_{run_id}.xlsx'
    
    src_path = os.path.abspath(GR_WORKSPACE_PATH)
    dst_path = os.path.abspath(GR_VOLUME_PATH)
    safe_base = os.path.abspath(GR_SAFE_BASE)
    dstn_base = os.path.abspath(GR_DSTN_BASE)
    
    if src_path.startswith(safe_base) and dst_path.startswith(dstn_base):
        copyfile(GR_WORKSPACE_PATH, GR_VOLUME_PATH)
    logger.info(f"Report written to WORKSPACE_PATH {GR_WORKSPACE_PATH}")
    logger.info(f"WORKSPACE_PATH copied to WORKSPACE_PATH {GR_VOLUME_PATH}")

    # Simple path traversal protection
    if os.path.abspath(GR_WORKSPACE_PATH).startswith(os.path.abspath(GR_SAFE_BASE)):
        os.remove(GR_WORKSPACE_PATH)


def merge_into_tp_data_vldtn_rprt_tbl(run_id,spark,data,valdn_grp_id,catalog_name,postgres_schema, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    logger = get_logger()
    logger.info("Starting merge into tp_valdn_rprt.")
    schema_for_kpi = StructType([ 
            StructField("run_id",LongType(),True),
            StructField("valdn_name",StringType(),True),
            StructField("reslt",StringType(),True),
            StructField("hyperlink_txt",StringType(),True),
            StructField("valdn_type",StringType(),True),
            StructField("valdn_path",StringType(),True)])

    # Creation of Summary Tab
    df_file_summary = spark.createDataFrame(data, schema_for_kpi)
    df_file_summary=df_file_summary.withColumn("creat_time", current_timestamp())
    df_file_summary= df_file_summary.orderBy('reslt')
    df_file_summary.createOrReplaceTempView('df_file_summary_view')
    logger.info("df_file_summary")
    df_file_summary.show()
   
    #retry for Merge operation
    safe_merge_with_retry(spark,df_file_summary, f'{catalog_name}.internal_tp.tp_valdn_rprt', f"tgt.run_id={run_id} AND src.valdn_name= tgt.valdn_name")
    logger.info(f"Final Validation Summary written successfully to {catalog_name}.internal_tp.tp_data_vldtn_rprt")
    
    if df_file_summary.filter(df_file_summary["reslt"]=="FAILED").count()>0:
        flag='Y'
    else:
        flag='N'
    logger.info(f"FLAG: {flag}")
    table_name=f'{postgres_schema}.mm_run_valdn_plc'

    df_valdn=read_from_postgres(table_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    
    data=[{"run_id": int(run_id),"valdn_grp_id": valdn_grp_id,"fail_ind": flag}]
    df=spark.createDataFrame(data)

    if df_valdn.filter((df_valdn["run_id"]==int(run_id)) & (df_valdn["valdn_grp_id"]==valdn_grp_id)).count()==0:
        write_to_postgres(df,table_name, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
        print('Appended to the table mm_run_valdn_plc')
        df.show()
    else:
        params=(flag,run_id,valdn_grp_id)
        query=f"UPDATE {postgres_schema}.mm_run_valdn_plc SET fail_ind=%s WHERE run_id=%s AND valdn_grp_id=%s"
        dbutils = get_dbutils(spark)
        db_config = get_database_config(dbutils)
        ref_db_hostname = db_config['ref_db_hostname']
        update_to_postgres(query,params, ref_db_name, ref_db_user, ref_db_pwd,ref_db_hostname)
        print('Updated table mm_run_valdn_plc')

# Loading the reference data variable
def reference_data_load_variables(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd ):
    df_cntrt_lkp=load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    cntrt_code=df_cntrt_lkp.collect()[0].cntrt_code
    srce_sys_id = df_cntrt_lkp.collect()[0].srce_sys_id
    time_perd_type_code = df_cntrt_lkp.collect()[0].time_perd_type_code

    df_mm_cntrt_categ_assoc_query = f"SELECT * FROM {postgres_schema}.mm_cntrt_categ_assoc WHERE cntrt_id={cntrt_id}"
    df_mm_cntrt_categ_assoc=read_query_from_postgres(df_mm_cntrt_categ_assoc_query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    return cntrt_code,srce_sys_id,time_perd_type_code,df_mm_cntrt_categ_assoc,df_cntrt_lkp

def business_validation(run_id,cntrt_id,srce_sys_id,file_name,validation_name,postgres_schema,catalog_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    logger = get_logger()
    df_mm_dvm_run_strct_lvl_plc=spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_valdn_run_strct_lvl_plc WHERE run_id=:run_id",{'run_id':run_id})
    df_mm_dvm_run_strct_plc=spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_valdn_run_strct_plc WHERE run_id=:run_id",{'run_id':run_id})
    mm_mkt_dim=spark.sql(f"""SELECT * FROM {catalog_name}.gold_tp.tp_mkt_dim WHERE srce_sys_id=:srce_sys_id""",{'srce_sys_id':srce_sys_id})

    mm_mkt_xref=spark.sql(f"""
        SELECT 
            extrn_mkt_id,
            mkt_skid,
            srce_sys_id,
            cntrt_id,
            part_srce_sys_id
        FROM {catalog_name}.internal_tp.tp_mkt_sdim 
        WHERE srce_sys_id = :srce_sys_id
    """,{'srce_sys_id':srce_sys_id})
    mm_prod_dim=spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_prod_dim WHERE srce_sys_id=:srce_sys_id",{'srce_sys_id':srce_sys_id})
    mm_prod_xref=spark.sql(f"""
        SELECT 
            extrn_prod_id,
            extrn_prod_attr_val_list,
            prod_match_attr_list,
            prod_skid,
            srce_sys_id,
            cntrt_id,
            part_srce_sys_id
        FROM {catalog_name}.internal_tp.tp_prod_sdim 
        WHERE srce_sys_id = :srce_sys_id
        """,{'srce_sys_id':srce_sys_id})

    query = f"{postgres_schema}.mm_cntrt_valdn_assoc"
    mm_cntrt_checks_assoc = read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id = {cntrt_id} and use_ind = 'Y'")
    checks = mm_cntrt_checks_assoc.select('valdn_id').collect()

    lst_of_checks=[]
    for i in checks:
        lst_of_checks.append(i[0].upper())

    chk_dq1 = 'CHK_DQ1' in lst_of_checks
    chk_dq2 = 'CHK_DQ2' in lst_of_checks
    chk_dq3 = 'CHK_DQ3' in lst_of_checks

    logger.info(f"Validation checks enabled: CHK_DQ1={chk_dq1}, CHK_DQ2={chk_dq2},  CHK_DQ3={chk_dq3}")

    # Load external fact/mkt/time data
    tier2_fact_stgng_tbl = spark.read.parquet(f'{materialise_path(spark)}/{run_id}/load_fact_df_fact_extrn')
    tier2_mkt_mtrlz_tbl = spark.read.parquet(f'{materialise_path(spark)}/{run_id}/market_transformation_df_mkt_stgng_vw')
    df_time_stgng_vw=spark.read.parquet(f'''{materialise_path(spark)}/{run_id}/time_transformation_df_time_stgng_vw''')
    df_fct_time_perd=df_time_stgng_vw
    # Load materialized product dimension data for the given run
    tier2_prod_mtrlz_tbl = spark.read.parquet(f'{materialise_path(spark)}/{run_id}/product_transformation_df_prod_stgng_vw')
    df_calc_index=spark.read.parquet(f'{materialise_path(spark)}/{run_id}/DQ_Pre_Business_Validation_df_calc_index')
    df_tier2_dqm_calc_index=df_calc_index
    df_time_pd = spark.sql(f"select * from {catalog_name}.gold_tp.tp_time_perd_fdim")

    tier2_fact_stgng_tbl.createOrReplaceTempView('TIER2_FACT_STGNG_TBL')
    tier2_mkt_mtrlz_tbl.createOrReplaceTempView('TIER2_MKT_MTRLZ_TBL')
    tier2_prod_mtrlz_tbl.createOrReplaceTempView('TIER2_PROD_MTRLZ_TBL')
    df_fct_time_perd.createOrReplaceTempView('fct_time_perd')
    df_tier2_dqm_calc_index.createOrReplaceTempView('TIER2_DQM_CALC_INDEX')
    df_mm_dvm_run_strct_lvl_plc.createOrReplaceTempView('MM_DVM_RUN_STRCT_LVL_PLC')
    df_mm_dvm_run_strct_plc.createOrReplaceTempView('MM_DVM_RUN_STRCT_PLC')
    df_time_pd.createOrReplaceTempView('mm_time_perd_fdim')
    mm_mkt_dim.createOrReplaceTempView('mm_mkt_dim')
    mm_mkt_xref.createOrReplaceTempView('mm_mkt_xref')
    mm_prod_dim.createOrReplaceTempView('mm_prod_dim')
    mm_prod_xref.createOrReplaceTempView('mm_prod_xref')

    # Validations
    #1 - Unexpected change vs. previous period
    df = dq_query_retrieval(catalog_name,validation_name,"Tier2",'Unexpected change vs previous period',spark)
    query1= df.collect()[0]["qry_txt"]
    logger.info(f"Started CHK_DQ1 validation: {query1}")
    df_val1 = spark.sql(query1)

    #2 - Unexpected backdata difference
    df = dq_query_retrieval(catalog_name,validation_name,"Tier2",'Unexpected backdata difference',spark)
    query2 = df.collect()[0]["qry_txt"]
    query2 = f"f'''{query2}'''"
    e=eval
    query2 = e(query2)

    logger.info(f"Started CHK_DQ2 validation: {query2}")
    df_val2 = spark.sql(query2)
    df_val2.show(5)

    #3 - Negative fact values    
    query_vars = {
            "run_id": run_id,
            "srce_sys_id": srce_sys_id,
            "tier2_fact_stgng_tbl": tier2_fact_stgng_tbl,
            "df_mm_dvm_run_strct_plc": df_mm_dvm_run_strct_plc,
            "spark": spark,
            "row_number":row_number,
            "monotonically_increasing_id": monotonically_increasing_id,
            "Window": Window
        }
    df = dq_query_retrieval(catalog_name,validation_name,"Tier2",'Negative fact values',spark)
    query3= df.collect()[0]["qry_txt"]
    logger.info(f"Started CHK_DQ3 validation: {query3}")
    try:
        exec(query3,query_vars)
        df_val3=query_vars.get("df_val3")
        df_val3.show(5)
    except Exception as e:
        print(" Error during exec of CHK_DQ3 validation:", e)

    data = []
    if chk_dq1:
        dq_val1=validation(catalog_name,df_val1,'Unexpected change vs previous period',run_id,spark,validation_name)
        data.append(dq_val1)

    if chk_dq2:
        dq_val2=validation(catalog_name,df_val2,'Unexpected backdata difference',run_id,spark,validation_name)
        data.append(dq_val2)

    if chk_dq3:
        dq_val3=validation(catalog_name,df_val3,'Negative fact values',run_id,spark,validation_name)
        data.append(dq_val3)
    valdn_grp_id=4
    logger.info("Creating summary report for validations.")    
    merge_into_tp_data_vldtn_rprt_tbl(run_id,spark,data,valdn_grp_id,catalog_name,postgres_schema, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    logger.info("Starting report generation")
    #Use generate_report definition to generate an excel 
    generate_report(run_id, cntrt_id, file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
    
    #Use generate_formatting_report definition to format the excel 
    generate_formatting_report(run_id,catalog_name)  

def fact_multipliers_trans(run_id,cntrt_id,spark):
        logger = get_logger()
        # Multipliers
        df_dpf_col_asign_vw = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/column_mappings_df_dpf_col_asign_vw")
        df_fact_extrn = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/load_fact_df_fact_extrn")
        logger.info("df_fact_extrn")
        df_fact_extrn.show(5)

        # Filter and select relevant rows and columns from df_dpf_col_asign_vw
        rows_measr_mul = df_dpf_col_asign_vw.where('dmnsn_name LIKE "%FCT%" OR dmnsn_name LIKE "%fct%"') \
        .where('multr_val is not null and length(multr_val)!=0') \
        .filter(col('cntrt_id') == f'{cntrt_id}') \
        .select('db_col_name', 'multr_val') \
        .select(concat_ws('*', df_dpf_col_asign_vw.db_col_name, nanvl(df_dpf_col_asign_vw.multr_val, lit(1))).alias('fct_trans'), df_dpf_col_asign_vw.db_col_name) \
        .collect()
        logger.info(f"Multiplier mappings found: {rows_measr_mul}")

        # Get the list of columns from df_fact_extrn
        lst_inp_cols = df_fact_extrn.columns
        str_inp_cols = ','.join(lst_inp_cols)

        # Create a dictionary for columns with multr_vals
        dict_measr_mul = {}
        for row in rows_measr_mul:
            if row.db_col_name.lower() in str_inp_cols:  # Check if the column is present in the input file
                if "," in row.db_col_name:
                    mult_factor = row.fct_trans.split("*")[1]
                    lst_keys = row.db_col_name.split(",")
                    for key2 in lst_keys:
                        dict_measr_mul[key2.lower()] = '*'.join([key2.lower(), mult_factor])
                else:
                    dict_measr_mul[row.db_col_name.lower()] = row.fct_trans.lower()

        # Create a list of transformed columns with multr_vals
        lst_measr_mul = [f"({dict_measr_mul[x]}) as {x}," for x in dict_measr_mul]
        str_measr_mul = ''.join(lst_measr_mul)

        # Create a list of remaining source columns
        lst_rem_source_cols = [col for col in lst_inp_cols if col not in [x for x in dict_measr_mul]]
        str_rem_source_cols = ','.join(lst_rem_source_cols)

        # Combine the transformed and remaining columns into a query string
        str_qry = str_measr_mul + str_rem_source_cols
        logger.info(f"[{run_id}] SQL query for Fact Multipliers Transformation: {str_qry}")

        # Create a temporary view from df_fact_extrn
        df_fact_extrn.createOrReplaceTempView('temp_input')

        # Execute the SQL query to transform the DataFrame
        df_fact_extrn2 = spark.sql(f""" select {str_qry} from temp_input """)
        df_fact_extrn2.show(5)

        # Write the transformed DataFrame to a parquet file
        df_fact_extrn2.write.mode("overwrite").format('parquet').save(f"{materialise_path(spark)}/{run_id}/load_file_df_fact_extrn")
        logger.info(f"Writing transformed DataFrame to {materialise_path(spark)}/{run_id}/load_file_df_fact_extrn with {df_fact_extrn2.count()} rows")

#df_input is the dataframe that needs to be complemented with the schema of df_schema dataframe 
def column_complementer(df_input, df_schema):

    # Get the columns of the input DataFrame
    raw_input_cols = df_input.columns
    input_cols=[col.lower() for col in raw_input_cols]

    # Get the columns of the schema DataFrame
    raw_schema_cols=df_schema.columns
    schema_cols = [col.lower() for col in raw_schema_cols]

    # Determine the columns that need to be added to the input DataFrame
    add_cols = list(set(schema_cols)-set(input_cols))

    # Add the missing columns to the input DataFrame with null values
    for i in add_cols:
        df_input = df_input.withColumn(i,lit(None).cast('string'))

    # Reorder the columns of the input DataFrame to match the schema DataFrame
    df_input = df_input.select(*schema_cols)
    cols = df_input.columns

    # Ensure the data types of the columns in the input DataFrame match the schema DataFrame
    for j in cols:
        if (dict(df_input.dtypes)[j]!=dict(df_schema.dtypes)[j]):
            df_input = df_input.withColumn(j, col(j).cast(dict(df_schema.dtypes)[j]))

    # Return the modified input DataFrame
    return df_input

def dynamic_expression(df_exprn_lkp, cntrt_id, exprn_name, dmnsn_name,spark):
    # Create or replace a temporary view from the DataFrame
    df_exprn_lkp.createOrReplaceTempView('mm_exprn_lkp')
    
    # Execute a parameterized SQL query to fetch the expression based on the provided contract ID, operation type, and dimension type
    df_parameterised_qry = spark.sql('''SELECT * FROM mm_exprn_lkp WHERE cntrt_id= :cntrt_id AND exprn_name= :exprn_name AND dmnsn_name= :dmnsn_name ''',{"cntrt_id" : cntrt_id,"dmnsn_name" : dmnsn_name, "exprn_name" : exprn_name })
    
    # Select the first row of the result and get the 'exprn' column
    row_exprn = df_parameterised_qry.select('exprn_val').first()
    
    # If no contract specific expression is found, execute a query with a default contract ID (101)
    if not row_exprn:
        df_parameterised_qry = spark.sql('''SELECT * FROM mm_exprn_lkp WHERE cntrt_id=101 AND exprn_name= :exprn_name AND dmnsn_name= :dmnsn_name ''',{"dmnsn_name" : dmnsn_name, "exprn_name" : exprn_name })
        row_exprn = df_parameterised_qry.select('exprn_val').first()
    
    # Extract the expression from the row
    query = row_exprn['exprn_val']
    
    # Format the query string to be evaluated later
    query = f"f'''{query}'''"

    #Return the formatted query string
    return query

def add_partition_columns(df, cntrt_id):
    return df.withColumn('part_srce_sys_id', df["srce_sys_id"].cast(IntegerType())) \
             .withColumn('part_cntrt_id', lit(int(cntrt_id))) \
             .withColumn('part_mm_time_perd_end_date', df["mm_time_perd_end_date"].cast("date"))
                        
def time_perd_class_codes(cntrt_id,catalog_name, postgres_schema,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    logger = get_logger()
    # Fetch contract lookup data from the Postgres table for the given contract ID
    df_cntrt_lkp=load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )

    # Extract the time period type code from the first row of the result
    time_perd_type_code = df_cntrt_lkp.collect()[0].time_perd_type_code
    logger.info(f"time_perd_type_code: {time_perd_type_code}")
    # Map the time period type code to a corresponding class code
    query = f"""SELECT time_perd_class_code FROM {catalog_name}.gold_tp.tp_time_perd_type WHERE time_perd_type_code=:time_perd_type_code"""

    time_perd_class_code_df=spark.sql(query,{'time_perd_type_code':time_perd_type_code})
    time_perd_class_code = time_perd_class_code_df.collect()[0][0]
    return time_perd_class_code

def time_perd_class_codes_new(time_perd_type_code,catalog_name,spark):
    logger = get_logger()
    # # Fetch contract lookup data from the Postgres table for the given contract ID
    logger.info(f"time_perd_type_code: {time_perd_type_code}")
    # Map the time period type code to a corresponding class code
    query = f"""SELECT time_perd_class_code FROM {catalog_name}.gold_tp.tp_time_perd_type WHERE time_perd_type_code=:time_perd_type_code"""

    time_perd_class_code_df=spark.sql(query,{'time_perd_type_code':time_perd_type_code})
    time_perd_class_code = time_perd_class_code_df.collect()[0][0]
    return time_perd_class_code

def your_information_validations(run_id,cntrt_id,srce_sys_id,file_name,validation_name,catalog_name,postgres_schema,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    logger = get_logger()
    mmc_mkt_source=spark.read.parquet(f"{materialise_path(spark)}/{run_id}/load_market_df_mkt_extrn")
    mmc_mkt_source_raw=spark.read.parquet(f"{materialise_path(spark)}/{run_id}/load_market_df_mkt_extrn_raw")
    tier2_fact_mtrlz_tbl_raw=spark.read.parquet(f"{materialise_path(spark)}/{run_id}/load_fact_df_fact_extrn_raw")
    df_db_cols_map=spark.read.parquet(f"{materialise_path(spark)}/{run_id}/column_mappings_df_dpf_col_asign_vw")
    mmc_prod_source=spark.read.parquet(f"{materialise_path(spark)}/{run_id}/load_product_df_prod_extrn")
    mmc_prod_source_raw=spark.read.parquet(f"{materialise_path(spark)}/{run_id}/load_product_df_prod_extrn_raw")
    df_mkt_xref = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_mkt_sdim WHERE part_srce_sys_id=:srce_sys_id",{'srce_sys_id':srce_sys_id})
    df_mkt_dim_sel = spark.sql(f"select * from {catalog_name}.gold_tp.tp_mkt_dim WHERE part_srce_sys_id=:srce_sys_id",{'srce_sys_id':srce_sys_id})
    #Perform the join operation
    df_mkt_dim_sel = df_mkt_dim_sel.alias("df_mkt_dim_sel")
    df_mkt_xref = df_mkt_xref.alias("df_mkt_xref")
    df_mkt_dim_sel.createOrReplaceTempView("mkt_dim_sel")
    df_mkt_xref.createOrReplaceTempView("mkt_xref")
    df_mkt_dim_sel= spark.sql(""" SELECT mkt_dim_sel.* FROM mkt_dim_sel INNER JOIN mkt_xref ON mkt_dim_sel.srce_sys_id = mkt_xref.srce_sys_id """)

    query = f"{postgres_schema}.mm_cntrt_valdn_assoc"
    mm_cntrt_checks_assoc = read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id = {cntrt_id} and use_ind = 'Y'")
    checks = mm_cntrt_checks_assoc.select('valdn_id').collect()

    lst_of_checks=[]
    for i in checks:
        lst_of_checks.append(i[0].upper())

    chk_dq10 = 'CHK_DQ10' in lst_of_checks
    chk_dq13 = 'CHK_DQ13' in lst_of_checks

    logger.info(f"Validation checks enabled: CHK_DQ10={chk_dq10}, CHK_DQ13={chk_dq13}")

    if 'mkt_name' not in mmc_mkt_source.columns:
        mmc_mkt_source = mmc_mkt_source.withColumn('mkt_name', lit(None).cast(StringType()))
    mmc_mkt_source.createOrReplaceTempView('mmc_mkt_source')
    df_db_cols_map.createOrReplaceTempView('df_db_cols_map')
    mmc_prod_source_raw.createOrReplaceTempView('mmc_prod_source_raw')
    mmc_mkt_source_raw.createOrReplaceTempView('mmc_mkt_source_raw')
    tier2_fact_mtrlz_tbl_raw.createOrReplaceTempView('tier2_fact_mtrlz_tbl_raw')
    df_mkt_xref.createOrReplaceTempView('mm_prod_xref')

    #-----------------Generate df_val10 value-------
    query_vars= {
        'df_db_cols_map': df_db_cols_map,
        'tier2_fact_mtrlz_tbl_raw': tier2_fact_mtrlz_tbl_raw,
        'mmc_mkt_source_raw': mmc_mkt_source_raw,
        'mmc_prod_source_raw': mmc_prod_source_raw,
        'StructType': StructType,
        'StructField': StructField,
        'StringType': StringType,
        'spark': spark,
        'cntrt_id': cntrt_id
    }
    df = dq_query_retrieval(catalog_name,validation_name,"Tier2",'Unmapped and missing optional columns',spark)
    query_df_val10= df.collect()[0]["qry_txt"]
    try:
        logger.info(f"Started CHK_DQ10 validation: {query_df_val10}")
        exec(query_df_val10,query_vars)
        df_val10 = query_vars.get("df_val10")
        logger.info("DF_VAL10")
        df_val10.show(5)
    except Exception as e:
        print(" Error during exec of CHK_DQ10 validation:", e)

        #-----------------Generate df_val13 value-------

    query_vars.update({
        'df_mkt_dim_sel': df_mkt_dim_sel,
        'df_mkt_xref': df_mkt_xref,
        'mmc_mkt_source': mmc_mkt_source,
        'col': col,
        'expr': expr,
        'lit': lit,
    })
    df = dq_query_retrieval(catalog_name,validation_name,"Tier2",'Modified area description',spark)
    query_df_val13= df.collect()[0]["qry_txt"]
    try:
        logger.info(f"Started CHK_DQ13 validation: {query_df_val13}")
        exec(query_df_val13,query_vars)
        df_val13 = query_vars.get('df_val13')
        logger.info("DF_VAL13")
        df_val13.show(5)
    except Exception as e:
        print(" Error during exec of CHK_DQ13 validation:", e)

    #---------------chk_dq10 & chk_dq13------------------------
    data = []
    if chk_dq10:
        dq_val1=validation(catalog_name,df_val10,'Unmapped and missing optional columns',run_id,spark,validation_name)
        data.append(dq_val1)
    
    if chk_dq13:
        dq_val2=validation(catalog_name,df_val13,'Modified area description',run_id,spark,validation_name)
        data.append(dq_val2)
    
    # Using merge condition Insert and update the data to table
    valdn_grp_id=3
    logger.info("Creating summary report for validations.")
    merge_into_tp_data_vldtn_rprt_tbl(run_id,spark,data,valdn_grp_id,catalog_name,postgres_schema, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    logger.info("Starting report generation")
    generate_report(run_id, cntrt_id, file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
    generate_formatting_report(run_id,catalog_name)

def reference_validation(materialised_path,validation_name,file_name,cntrt_id,run_id,spark, ref_db_jdbc_url,ref_db_name, ref_db_user, ref_db_pwd,postgres_schema,catalog_name):
    logger = get_logger()
    valdn_grp_id = 2

    cntrt_code,srce_sys_id,time_perd_type_code,df_mm_cntrt_categ_assoc,df_cntrt_lkp = reference_data_load_variables(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )

    tier2_mkt_mtrlz_tbl=spark.read.parquet(f'''{materialised_path}/{run_id}/market_transformation_df_mkt_stgng_vw''')
    TIER2_PROD_MTRLZ_TBL=spark.read.parquet(f'{materialised_path}/{run_id}/product_transformation_df_prod_stgng_vw')
    tier2_fact_stgng =  spark.read.parquet(f"{materialised_path}/{run_id}/load_fact_df_fact_extrn")
    tier2_fact_mtrlz_tbl = spark.read.parquet(f"{materialised_path}/{run_id}/load_fact_df_fact_extrn")
    
    
    df_fact = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_wk_fct WHERE cntrt_id=:cntrt_id", {'cntrt_id':cntrt_id})
    df_time = spark.read.parquet(f"{materialised_path}/{run_id}/time_transformation_df_time_stgng_vw")
    fact_w_time_perd_id = df_time.join(df_fact, df_time.time_perd_id == df_fact.time_perd_id, 'leftouter').select(df_time['*'])
    df_mkt_dim = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_mkt_dim WHERE part_srce_sys_id=:srce_sys_id",{'srce_sys_id':srce_sys_id})
    df_mkt_sdim = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_mkt_sdim WHERE part_srce_sys_id=:srce_sys_id",{'srce_sys_id':srce_sys_id})

    df_run_plc_query = f"{postgres_schema}.mm_run_plc"
    df_run_plc = read_from_postgres(df_run_plc_query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id = {cntrt_id}")

    df_mm_run_mkt_plc = spark.read.parquet(f'''{materialised_path}/{run_id}/market_transformation_df_mkt_stgng_vw''')
    mm_run_prttn_plc = spark.sql(f"select * from {catalog_name}.gold_tp.tp_run_prttn_plc where cntrt_id = :cntrt_id",{"cntrt_id":cntrt_id})
    df_mkt_xref = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_mkt_sdim WHERE cntrt_id = :cntrt_id",{"cntrt_id":cntrt_id})
    df_prod_dim = spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_prod_dim WHERE part_srce_sys_id=:srce_sys_id',{'srce_sys_id':srce_sys_id})
    df_prod_sdim = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_prod_sdim WHERE part_srce_sys_id=:srce_sys_id",{"srce_sys_id":srce_sys_id})
    df_mm_run_prod_plc =spark.read.parquet(f'''{materialised_path}/{run_id}/product_transformation_df_prod_stgng_vw''')

    mm_cntrt_checks_assoc_query = f"{postgres_schema}.mm_cntrt_valdn_assoc"
    mm_cntrt_checks_assoc = read_from_postgres(mm_cntrt_checks_assoc_query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id = {cntrt_id} AND use_ind = 'Y'")
    #============================= temp_view ===================================
    tier2_mkt_mtrlz_tbl.createOrReplaceTempView("TIER2_MKT_MTRLZ_TBL")
    TIER2_PROD_MTRLZ_TBL.createOrReplaceTempView("TIER2_PROD_MTRLZ_TBL")
    tier2_fact_mtrlz_tbl.createOrReplaceTempView('TIER2_FACT_MTRLZ_TBL')
    fact_w_time_perd_id.createOrReplaceTempView("fact_w_time_perd_id")
    df_mkt_dim.createOrReplaceTempView("MM_MKT_DIM")
    df_mkt_sdim.createOrReplaceTempView("MM_MKT_SDIM")
    df_run_plc.createOrReplaceTempView('MM_RUN_PLC')
    df_cntrt_lkp.createOrReplaceTempView('MM_CNTRT_LKP')

    mm_run_prttn_plc.createOrReplaceTempView('MM_RUN_PRTTN_PLC')
    df_mm_run_mkt_plc.createOrReplaceTempView('MM_RUN_MKT_PLC')
    df_mkt_xref.createOrReplaceTempView("MM_MKT_XREF")
    df_mm_run_prod_plc.createOrReplaceTempView('MM_RUN_PROD_PLC')

    df_prod_dim.createOrReplaceTempView("MM_PROD_DIM")

    df_fact.createOrReplaceTempView('MM_TP_MTH_FCT')
    tier2_fact_stgng.createOrReplaceTempView('tier2_fact_stgng')
    df_prod_sdim.createOrReplaceTempView("MM_PROD_SDIM")
    #====================Checks============================================
    checks = mm_cntrt_checks_assoc.select('valdn_id').collect()
    lst_of_checks=[]
    for i in checks:
        lst_of_checks.append(i[0].upper())

    # Need to check whether we have any lookup table defining the chk_dq in postgres, to inner join
    chk_dq16 = 'CHK_DQ16' in lst_of_checks
    chk_dq17 = 'CHK_DQ17' in lst_of_checks
    print(chk_dq17, "-----chk_dq17---")
    chk_dq18 = True  # Always true based on ADO 947773
    chk_dq19 = False 
    logger.info("Starting Validation")
    data = []
    if chk_dq16:
        df = dq_query_retrieval(catalog_name,validation_name,"Tier2","Missing/delivered areas",spark)
        query16 = df.collect()[0]['qry_txt']
        print("performing df_val------")
        query16 = f"f'''{query16}'''"
        e=eval
        query16 = e(query16)
        
        df_val = spark.sql(query16)
        print("-----printing df_val----")

        print("-----df_val------")

        dq_val = validation(catalog_name,df_val,"Missing/delivered areas",run_id,spark,validation_name)
        data.append(dq_val)
    if chk_dq17:
        df = dq_query_retrieval(catalog_name,validation_name,"Tier2","Missing/delivered hierarchies",spark)
        query17 = df.collect()[0]['qry_txt']
        query17 = f"f'''{query17}'''"
        e=eval
        query17 = e(query17)
        df_val = spark.sql(query17)
        dq_val = validation(catalog_name,df_val,"Missing/delivered hierarchies",run_id,spark,validation_name)
        data.append(dq_val)
    if chk_dq18:
        df = dq_query_retrieval(catalog_name,validation_name,"Tier2","Unknown time period",spark)
        query18 = df.collect()[0]['qry_txt']
        print("--queries--")
        print(query18)
        print("--queries--")
        dq_val = validation(catalog_name,query18,"Unknown time period",run_id,spark,validation_name)
        data.append(dq_val)        
    if chk_dq19:
        df = dq_query_retrieval(catalog_name,validation_name,"Tier2","Modified product description",spark)
        query19= df.collect()[0]['qry_txt']
        query19 = f"f'''{query19}'''"
        e=eval
        query19 = e(query19)
        df_val = spark.sql(query19)
        dq_val = validation(catalog_name,df_val,"Modified product description",run_id,spark,validation_name)
        data.append(dq_val)

    #  Create Dataframe Schema and Merge into tp_data_vldtn_rprt table
    logger.info("merging dataframe")
    merge_into_tp_data_vldtn_rprt_tbl(run_id,spark,data,valdn_grp_id,catalog_name,postgres_schema, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    # # Report Formation
    logger.info("generate report")
    generate_report(run_id,cntrt_id,file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
    
    # # Report Generation
    generate_formatting_report(run_id,catalog_name)


def file_structure_validation(notebook_name,validation_name,file_name,cntrt_id,run_id,spark, ref_db_jdbc_url,ref_db_name, ref_db_user, ref_db_pwd,postgres_schema,catalog_name):
    logger = get_logger()
    # Filter the DataFrame to get the 'Source System ID' and TIME_PERD_TYPE_CODE for the specified contract ID
    df_cntrt_lkp=load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    time_perd_type_code = df_cntrt_lkp.collect()[0].time_perd_type_code
    cntry_name=df_cntrt_lkp.collect()[0].cntry_name

    # Load various parquet files into DataFrames for further processing
    tier2_tmp_extrnl_tbl = spark.read.parquet(
        f"{materialise_path(spark)}/{run_id}/load_fact_df_fact_extrn_raw"
    )
    mmc_mkt_source = spark.read.parquet(
        f"{materialise_path(spark)}/{run_id}/load_market_df_mkt_extrn"
    )
    mmc_prod_source = spark.read.parquet(
        f"{materialise_path(spark)}/{run_id}/load_product_df_prod_extrn"
    )
    tier2_fact_mtrlz_tbl = spark.read.parquet(
        f"{materialise_path(spark)}/{run_id}/load_fact_df_fact_extrn_dvm"
    )
    mmc_mkt_source_raw = spark.read.parquet(
        f"{materialise_path(spark)}/{run_id}/load_market_df_mkt_extrn_raw"
    )
    mmc_prod_source_raw = spark.read.parquet(
        f"{materialise_path(spark)}/{run_id}/load_product_df_prod_extrn_raw"
    )
    tier2_fact_mtrlz_tbl_raw = spark.read.parquet(
        f"{materialise_path(spark)}/{run_id}/load_fact_df_fact_extrn_raw"
    )
    df_db_cols_map = spark.read.parquet(
        f"{materialise_path(spark)}/{run_id}/column_mappings_df_dpf_col_asign_vw"
    )

    # Initialize data quality check flags
    chk_dq4 = True
    chk_dq5 = True
    chk_dq6 = True
    chk_dq7 = True
    chk_dq8 = True
    chk_dq9 = True

    mmc_mkt_source.show()

    # Register DataFrames as temporary views for SQL querying
    mmc_mkt_source.createOrReplaceTempView('mmc_mkt_source')
    mmc_prod_source.createOrReplaceTempView('mmc_prod_source')


    mmc_mkt_source1 = mmc_mkt_source.withColumn("line_num", row_number().over(Window.orderBy(monotonically_increasing_id()))
	)
	
	# Ensure 'mkt_name' column exists; if not, add it with null values
    if 'mkt_name' not in mmc_mkt_source1.columns:
        mmc_mkt_source1 = mmc_mkt_source1.withColumn('mkt_name', lit(None).cast('string'))
	
	# Register the updated DataFrame as a temporary view
    mmc_mkt_source1.createOrReplaceTempView('mmc_mkt_source1')

    mmc_mkt_source1.show()

    
    # Add a unique line number to each row in the product source DataFrame
    mmc_prod_source1 = mmc_prod_source.withColumn(
		"line_num", row_number().over(Window.orderBy(monotonically_increasing_id()))
	)
	
	# Ensure 'prod_lvl_name' column exists; if not, add it with null values
    if 'prod_lvl_name' not in mmc_prod_source1.columns:
	    mmc_prod_source1 = mmc_prod_source1.withColumn('prod_lvl_name', lit(None).cast('string'))
	
	# Register the updated DataFrame as a temporary view
    mmc_prod_source1.createOrReplaceTempView('mmc_prod_source1')

    # Add a unique line number to each row in the materialized fact table
    # This is useful for tracking or ordering rows uniquely
    tier2_fact_mtrlz_tbl = tier2_fact_mtrlz_tbl.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id()))
    )

    # Register the updated DataFrame as a temporary view for SQL operations
    tier2_fact_mtrlz_tbl.createOrReplaceTempView('tier2_fact_mtrlz_tbl')

    # Register the temporary external table as a view for SQL operations
    tier2_tmp_extrnl_tbl.createOrReplaceTempView('tier2_tmp_extrnl_tbl')

    # Validation Starts
    
    data = []

    # Bad Fact Validation
    df = dq_query_retrieval(catalog_name,validation_name,'Tier2','Bad fact data format',spark)
    query = df.collect()[0]['qry_txt']
    exec_env = {
        "tier2_fact_mtrlz_tbl": tier2_fact_mtrlz_tbl,
        "expr": expr,
        "lit": lit,
        "when": when,
        "print": print
    }
    try:
        exec(query, exec_env)
        fact_bad_df = exec_env.get("fact_bad_df")
        print("✅ fact_bad_df retrieved:", fact_bad_df is not None)
    except Exception as e:
        print("❌ Error during exec:", e)
    dq_val4 = validation(catalog_name,fact_bad_df,'Bad fact data format',run_id,spark,validation_name)
    data.append(dq_val4)

    # Missing Mandatory columns
    
    exec_env = {
        "cntrt_id": cntrt_id,
        "df_db_cols_map": df_db_cols_map,
        "tier2_fact_mtrlz_tbl_raw": tier2_fact_mtrlz_tbl_raw,
        "mmc_mkt_source_raw": mmc_mkt_source_raw,
        "mmc_prod_source_raw": mmc_prod_source_raw,
        "col": col,
        "lower": lower,
        "upper": upper,
        "trim": trim,
        "lit": lit,
        "when": when,
        "print": print
    }

    df = dq_query_retrieval(catalog_name,validation_name,'Tier2','Missing Mandatory columns',spark)
    query = df.collect()[0]['qry_txt']

    exec(query, exec_env)
    df_val5 = exec_env.get('df_val5')
    dq_val5 = validation(catalog_name,df_val5,'Missing Mandatory columns',run_id,spark,validation_name)
    data.append(dq_val5)
    logger.info("Completed Missing Mandatory columns validation")
    print("Completed dq_val_5")
    # SQL Validations


    # ---------------Fact uses supplier tag not defined in the reference part Validation-------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier2','Fact uses supplier tag not defined in the reference part',spark)
    query = df.collect()[0]['qry_txt']
    dq_val6 = validation(catalog_name,query,'Fact uses supplier tag not defined in the reference part',run_id,spark,validation_name)
    data.append(dq_val6)

    # --------------------Duplicated fact in the input file Validation----------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier2','Duplicated fact in the input file',spark)
    query = df.collect()[0]['qry_txt']
    dq_val7 = validation(catalog_name,query,'Duplicated fact in the input file',run_id,spark,validation_name)
    data.append(dq_val7)

    # --------------------Duplicated market code in the input file---------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier2','Duplicated market code in the input file',spark)
    query = df.collect()[0]['qry_txt']
    dq_val8 = validation(catalog_name,query,'Duplicated market code in the input file',run_id,spark,validation_name)
    data.append(dq_val8)

    # --------------------------Duplicated product code in the input file--------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier2','Duplicated product code in the input file',spark)
    query = df.collect()[0]['qry_txt']
    dq_val9 = validation(catalog_name,query,'Duplicated product code in the input file',run_id,spark,validation_name)
    data.append(dq_val9)

    # Prepare Summary Report dataframe
    
    logger.info("Starting merge into tp_valdn_rprt.")
    schema_for_kpi = StructType([ 
        StructField("run_id",LongType(),True),
        StructField("valdn_name",StringType(),True),
        StructField("reslt",StringType(),True),
        StructField("hyperlink_txt",StringType(),True),
        StructField("valdn_type",StringType(),True),
        StructField("valdn_path",StringType(),True)
    ])
    # Creation of Summary Tab
    
    df_file_struct_summary = spark.createDataFrame(data, schema_for_kpi)
    df_file_struct_summary=df_file_struct_summary.withColumn("creat_time", current_timestamp())
    df_file_struct_summary= df_file_struct_summary.orderBy('reslt')
    
    #Display the Summary Report
    df_file_struct_summary.show()
    
    #Materialised the summary report to storage
    df_file_struct_summary.write.mode("overwrite").format('parquet').save(f"{materialise_path(spark)}/{run_id}/{notebook_name}_df_file_struct_summary")

    # Append the file structure validation summary to the table
    safe_write_with_retry(df_file_struct_summary, f'{catalog_name}.internal_tp.tp_valdn_rprt', 'append', 'run_id')
    
    # Determine if any validation result has failed
    if df_file_struct_summary.filter(df_file_struct_summary["reslt"]=="FAILED").count()>0:
        flag='Y'
    else:
        flag='N'

    table_name=f'{postgres_schema}.mm_run_valdn_plc'

    df_valdn=read_from_postgres(table_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    
    # Prepare new validation result data
    valdn_grp_id=1
    data=[{"run_id":int(run_id),"valdn_grp_id":1,"fail_ind":flag}]
    df=spark.createDataFrame(data)

    
    # Check if a record for the current run already exists
    if df_valdn.filter((df_valdn["run_id"]==int(run_id)) & (df_valdn["valdn_grp_id"]==1)).count()==0:
        
        # If not, insert the new record
        write_to_postgres(df,table_name, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    
    else:
        
        # If it exists, update the failure indicator
        params=(run_id,valdn_grp_id)
        query=f"UPDATE {postgres_schema}.mm_run_valdn_plc SET fail_ind='{flag}' WHERE run_id=%s AND valdn_grp_id=%s"
        dbutils = get_dbutils(spark)
        db_config = get_database_config(dbutils)
        ref_db_hostname = db_config['ref_db_hostname']
        update_to_postgres(query,params, ref_db_name, ref_db_user, ref_db_pwd,ref_db_hostname)
    
    df.show()

    # # Report Formation
    generate_report(run_id,cntrt_id,file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
    
    # # Report Generation
    generate_formatting_report(run_id,catalog_name)

def run_prttn_dtls(cntrt_id,catalog_name,postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):

    query=f'{postgres_schema}.mm_run_plc'
    df_run_plc=read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id={cntrt_id}")
    df_run_plc.createOrReplaceTempView("mm_run_plc")


    query=f"""
        SELECT 
            plc.time_perd_class_code,
            plc.srce_sys_id,
            plc.cntrt_id,
            plc.mm_time_perd_end_date,
            plc.run_id,
            run.run_sttus_name
        FROM 
            {catalog_name}.gold_tp.tp_run_prttn_plc plc
        JOIN 
            mm_run_plc run ON run.run_id = plc.run_id
        
        WHERE 
                plc.cntrt_id = {cntrt_id}

        """

    df_run_prttn_run_sttus_vw=spark.sql(query)
    
    return df_run_prttn_run_sttus_vw



def publish_valdn_run_strct_lvl(cntrt_id, run_id, srce_sys_id, postgres_schema, catalog_name, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    logger = get_logger()
    if srce_sys_id == 3:
        query=f'{postgres_schema}.mm_valdn_cntrt_strct_lvl_prc_vw'
        df_dqm_dvm_rslp = read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id={cntrt_id}")
    else:
        query=f'{postgres_schema}.mm_valdn_tier2_strct_lvl_prc_vw'
        df_dqm_dvm_rslp = read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id={cntrt_id}")

    df_dqm_dvm_rslp=df_dqm_dvm_rslp.withColumn('run_id', lit(run_id)).withColumn('strct_lvl_id', abs('strct_lvl_id'))

    df_dqm_dvm_rslp=df_dqm_dvm_rslp.withColumn('secure_group_key',lit(0))
    df_dqm_dvm_rslp=add_secure_group_key(df_dqm_dvm_rslp, cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    df_dqm_dvm_rslp.createOrReplaceTempView('mm_dqm_dvm_rslp')

    df_dvm_run_strct_lvl_plc=spark.sql(f'SELECT * FROM {catalog_name}.internal_tp.tp_valdn_run_strct_lvl_plc where run_id= :run_id',{"run_id":run_id})

    df_dqm_dvm_rslp.createOrReplaceTempView("dqm_dvm_rslp")
    df_dvm_run_strct_lvl_plc.createOrReplaceTempView("dvm_run_strct_lvl_plc")
    df_dqm_dvm_rslp_1 = spark.sql("""
    SELECT a.*
    FROM dqm_dvm_rslp a
    LEFT ANTI JOIN dvm_run_strct_lvl_plc b
    ON a.run_id = b.run_id
    """)

    df_dqm_dvm_rslp_1=column_complementer(df_dqm_dvm_rslp_1,df_dvm_run_strct_lvl_plc)

    safe_write_with_retry(df_dqm_dvm_rslp_1, f'{catalog_name}.internal_tp.tp_valdn_run_strct_lvl_plc', "append","run_id")

    logger.info("TP_VALDN_RUN_STRCT_LVL_PLC")

    df_dqm_dvm_rslp_1.show()

    return df_dqm_dvm_rslp

def publish_valdn_run_strct(run_id, cntrt_id, postgres_schema, catalog_name, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):

    logger = get_logger()

    query=f"{postgres_schema}.mm_valdn_cntrt_strct_prc_vw"
    df_dqm_dvm_rsp = read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id={cntrt_id}")
    

    df_dqm_dvm_rsp=df_dqm_dvm_rsp.withColumn('run_id', lit(run_id)).withColumn('strct_id', abs('strct_id'))
    
    df_dqm_dvm_rsp=df_dqm_dvm_rsp.withColumn('secure_group_key',lit(0))
    df_dqm_dvm_rsp=add_secure_group_key(df_dqm_dvm_rsp, cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    
    df_dvm_run_strct_plc=spark.sql(f'SELECT * FROM {catalog_name}.internal_tp.tp_valdn_run_strct_plc where run_id=:run_id',{"run_id":run_id})
    
    df_dqm_dvm_rsp.createOrReplaceTempView("dqm_dvm_rsp")
    df_dvm_run_strct_plc.createOrReplaceTempView("dvm_run_strct_plc")
    df_dqm_dvm_rsp_1 = spark.sql("""
        SELECT a.*
        FROM dqm_dvm_rsp a
        LEFT ANTI JOIN dvm_run_strct_plc b
        ON a.run_id = b.run_id""")

    df_dqm_dvm_rsp_1=column_complementer(df_dqm_dvm_rsp_1,df_dvm_run_strct_plc)
    safe_write_with_retry(df_dqm_dvm_rsp_1, f'{catalog_name}.internal_tp.tp_valdn_run_strct_plc', "append","run_id")

    logger.info("TP_VALDN_RUN_STRCT_PLC")
    df_dqm_dvm_rsp_1.show()



def publish_valdn_agg_fct(df,cntrt_id, postgres_schema, catalog_name, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,srce_sys_id,run_id):

    logger = get_logger()

    # Ensure required columns exist; add with default values if missing
    if 'time_factr_pp' not in df.columns:
        df = df.withColumn('time_factr_pp', lit(0))

    if 'share_su_pct' not in df.columns:
        df = df.withColumn('share_su_pct', lit(0))


    df_valdn_agg_fct_schema = spark.table(f"{catalog_name}.gold_tp.tp_valdn_agg_fct").limit(0)
    
    df=df.withColumn('part_cntrt_id',lit(cntrt_id))
    df=column_complementer(df,df_valdn_agg_fct_schema)
    df.createOrReplaceTempView('temp')

    df_mm_tp_valdn_agg_fct=spark.sql('SELECT * FROM temp WHERE cntrt_id IS NOT NULL AND cntrt_id > 0')

    df_mm_tp_valdn_agg_fct = add_secure_group_key(df_mm_tp_valdn_agg_fct,cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    dim_type='TP_VALDN_AGG_FCT'
    valdn_agg_fct_path = semaphore_generate_path(dim_type, srce_sys_id, cntrt_id)
    # Print the constructed path for verification
    logger.info(f"valdn_agg_fct_path: {valdn_agg_fct_path}")
    # Acquire semaphore for writing to the specified path
    logger.info("Acquiring semaphore for TP_RUN_PRTTN_PLC")
    valdn_agg_fct_check_path = semaphore_acquisition(run_id, valdn_agg_fct_path,catalog_name, spark)
    try:
        logger.info(f"Started Writing to tp_valdn_agg_fct")
        safe_write_with_retry(df_mm_tp_valdn_agg_fct, f'{catalog_name}.gold_tp.tp_valdn_agg_fct', 'append', 'part_cntrt_id')

    finally:
        # Release the semaphore for the current run
        release_semaphore(catalog_name,run_id,valdn_agg_fct_check_path , spark)
        logger.info("TP_VALDN_AGG_FCT Semaphore released")
    logger.info("TP_VALDN_AGG_FCT")
    df_mm_tp_valdn_agg_fct.show()

def calculate_retention_date(run_id,cntrt_id, srce_sys_id,retention_period,catalog_name,postgres_schema, time_perd_class_code,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):

    if retention_period and retention_period > 0:
        latst_time_perd_id_df = read_query_from_postgres(f"SELECT latst_time_perd_id from {postgres_schema}.mm_run_detl_plc where run_id={int(run_id)}", spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
        latst_time_perd_id = latst_time_perd_id_df.first()[0]
        end_date_df = spark.sql(f"SELECT time_perd_end_date FROM {catalog_name}.gold_tp.tp_time_perd_fdim WHERE time_perd_id= {int(latst_time_perd_id)}")
        end_date = end_date_df.first()[0]

        df_with_retention = spark.createDataFrame([(end_date, time_perd_class_code.upper(), retention_period)], ["end_date", "time_perd_class_code", "retention_period"])

        # Apply retention_date logic only to filtered rows
        df_with_retention = df_with_retention.withColumn(
            "retention_date",
            when(col("time_perd_class_code") == "WK", expr("date_sub(end_date, CAST(retention_period * 7 AS INT))"))
            .when(col("time_perd_class_code") == "MTH", expr("add_months(end_date, -CAST(retention_period AS INT))"))
            .when(col("time_perd_class_code") == "BIMTH", expr("add_months(end_date, -CAST(retention_period * 2 AS INT))"))
            .when(col("time_perd_class_code") == "QTR", expr("add_months(end_date, -CAST(retention_period * 3 AS INT))"))
            .otherwise(expr("date_sub(end_date, CAST(retention_period * 7 AS INT))"))
        )

        min_retention_date = df_with_retention.first()["retention_date"]

    else:
        min_retention_date = None

    return min_retention_date


def recursive_delete(path,dbutils):
    try:
        # Try listing the contents of the path
        items = dbutils.fs.ls(path)
        for item in items:
            if item.isDir():
                recursive_delete(item.path,dbutils)
            else:
                dbutils.fs.rm(item.path)
        dbutils.fs.rm(path)
        print(f"Successfully deleted: {path}")
    except Exception as e:
        # Ignore FileNotFoundException and continue
        if "FileNotFoundException" in str(e):
            print(f"Path not found, skipping: {path}")
        else:
            print(f"Error deleting {path}: {e}")

def t1_load_file(file_type,run_id,file_type2,fileformat,spark):
    logger = get_logger()
    mft_path = f"{derive_base_path(spark)}/WORK"
    logger.info(f"[t1_load_file] Loading file with format: {fileformat}")
    path = ''
    if fileformat == 'SFF3':
        file_name = get_file_name(spark,mft_path,run_id,f"*{file_type}*")
        path = f'{mft_path}/*{run_id}*/{file_name}'
        logger.info(f"[t1_load_file] Reading file from path: {path}")
        df_raw1 = spark.read.format('csv').option('header', True).option('delimiter', ',').load(path)

    elif fileformat in ('SFF','SFF2'):
        file_name = get_file_name(spark,mft_path,run_id,f"*{file_type2}*")
        path = f'{mft_path}/*{run_id}*/{file_name}'
        logger.info(f"[t1_load_file] Reading file from path: {path}")
        df_raw1 = spark.read.format('csv').option('header', True).option('delimiter', '|').load(path)
        
    elif fileformat in ('FFS','FFS2',"Tape2","Tape3"):
        path=f'{mft_path}/*{run_id}*/*'
        logger.info(f"[t1_load_file] Reading file from path: {path}")
        df_raw1=spark.read.text(path)
    else:
        raise ValueError(f"No file matched pattern '{path}' under '{mft_path}' for run_id '{run_id}'.")
        
    logger.info("[t1_load_file] File loaded successfully into df_raw1")

    return df_raw1

def load_fact_sfffile(file_type, run_id, file_type2, fileformat,spark):
    logger = get_logger()
    mft_path = f"{derive_base_path(spark)}/WORK"
    logger.info(f"[load_fact_sfffile] Starting file load process | run_id: {run_id}, fileformat: {fileformat}")
    schema = StructType([StructField(f"col{i}", StringType(), True) for i in range(1, 104)])
    path = ''
    
    if fileformat==('SFF3'):
        file_name = get_file_name(spark,mft_path,run_id,f"*{file_type2}*")
        path=f'{mft_path}/*{run_id}*/{file_name}'
        logger.info(f"[load_fact_sfffile] Reading file from path: {path}")
        df_raw1=spark.read.option('header','False').schema(schema).option('delimiter',',').csv(path)
        df_raw1=df_raw1.withColumn('col1', regexp_replace('col1','market_id', 'MKT_TAG')).withColumn('col2',regexp_replace('col2', 'product_id', 'PROD_TAG')).withColumn('col3', regexp_replace('col3', 'period_id', 'PER_TAG'))
        
    elif fileformat =='SFF' or fileformat =='SFF2':
        file_name = get_file_name(spark,mft_path,run_id,f"*{file_type}*")
        path=f'{mft_path}/*{run_id}*/{file_name}'
        logger.info(f"[load_fact_sfffile] Reading file from path: {path}")
        df_raw1 = spark.read.format('csv').option('delimiter', '|').schema(schema).load(path)
    
    else:
        raise ValueError(f"No file matched pattern '{path}' under '{mft_path}' for run_id '{run_id}'.")
    
    logger.info("[load_fact_sfffile] File loaded successfully into df_raw1")
    return df_raw1

def materialize(df, df_name, run_id):
    logger = get_logger()
    spark = get_spark_session()
    raw_path = materialise_path(spark)
    final_path = f"{raw_path}/{run_id}/{df_name}" 
    try:
        df.write.mode("overwrite").format("parquet").option("compression", "snappy").save(final_path)
        logger.info(f"[Materialize] {df_name} saved to path: {final_path}")
    
    except Exception as e:
        logger.error(f"[Materialize] Failed to write DataFrame '{df_name}' to {final_path}. Error: {str(e)}")
        raise

def extract_fact_ffs(df_raw1,run_id,cntrt_id,dbutils,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,postgres_schema):
    logger = get_logger()
    raw_path = materialise_path(spark)
    logger.info("[extract_fact_ffs] Filtering rows starting with 'U'")
    df_fact_u=df_raw1.filter(col("value").startswith("U"))
    logger.info("[extract_fact_ffs] Applying column transformations")
    df_fact=df_fact_u.withColumn("line_num", row_number().over(Window.orderBy(monotonically_increasing_id()))).withColumn("dummy_separator_1",substring(col("value"),1,1)).withColumn("mkt_extrn_code",substring(col("value"),2,9)).withColumn("prod_extrn_code",substring(col("value"),11,9)).withColumn("time_extrn_code",substring(col("value"),20,9)).withColumn("dummy_separator_2",substring(col("value"),28,1)).withColumn("mkt_extrn_code2",substring(col("value"),29,1)).withColumn("dummy_separator_3",substring(col("value"),30,23)).withColumn("fact_amt_1",substring(col("value"),53,10)).withColumn("fact_amt_2",substring(col("value"),63,10)).withColumn("fact_amt_3",substring(col("value"),73,10)).withColumn("fact_amt_4",substring(col("value"),83,10)).withColumn("fact_amt_5",substring(col("value"),93,10)).withColumn("fact_amt_6",substring(col("value"),103,10)).withColumn("fact_amt_7",substring(col("value"),113,10)).withColumn("fact_amt_8",substring(col("value"),123,10)).withColumn("fact_amt_9",substring(col("value"),133,10)).withColumn("fact_amt_10",substring(col("value"),143,10)).withColumn("fact_amt_11",substring(col("value"),153,10)).withColumn("fact_amt_12",substring(col("value"),163,10)).withColumn("fact_amt_13",substring(col("value"),173,10)).withColumn("fact_amt_14",substring(col("value"),183,10)).withColumn("fact_amt_15",substring(col("value"),193,10)).withColumn("fact_amt_16",substring(col("value"),203,10)).withColumn("fact_amt_17",substring(col("value"),213,10)).withColumn("fact_amt_18",substring(col("value"),223,10)).withColumn("fact_amt_19",substring(col("value"),233,10)).withColumn("fact_amt_20",substring(col("value"),243,10)).withColumn("other_fct_data",substring(col("value"),253,10)).drop("value")
    logger.info("[extract_fact_ffs] Trimming all columns")
    df_fact=df_fact.select(*[ltrim(rtrim(col)).alias(col)for col in df_fact.columns])
    
    df_srce_mfact = df_fact.withColumn("mkt_extrn_code",when(col("mkt_extrn_code") == 'N', concat(col("mkt_extrn_code2"), col("mkt_extrn_code"))).otherwise(col("mkt_extrn_code")))
    logger.info("[extract_fact_ffs] Extraction completed successfully")

    return df_srce_mfact

def extract_fact_sff(df_raw1,run_id,cntrt_id,srce_sys_id,spark):
    logger = get_logger()
    raw_path = materialise_path(spark)
    
    logger.info("[extract_fact_sff] Renaming columns")
    cls = df_raw1.columns
    j= 1
    for i in cls:
        df_raw1 = df_raw1.withColumnRenamed(i, i.replace('_c', 'col')[:3]+str(j))
        j = j+1
    lst_excld_cols=[] 
    lst_check_cols = [col for col in df_raw1.columns if col not in lst_excld_cols]
    str_cond = ''
    for c in lst_check_cols:
        str_cond += f'{c} is not null or '
    str_cond = str_cond.rstrip('or ')

    df_raw1.createOrReplaceTempView('fcts')
    df_srce_fct = df_raw1.filter(expr(str_cond))

    materialize(df_srce_fct,'Load_Fact_df_srce_fct',run_id)
    logger.info("Successfully materialised df_srce_fct")

    df_srce_fct=spark.read.parquet(f"{raw_path}/{run_id}/Load_Fact_df_srce_fct")

    logger.info("[extract_fact_sff] Filtering out tag rows")
    df_srce_pre_mfct=df_srce_fct.filter((df_srce_fct.col1 != 'MKT_TAG') &(df_srce_fct.col2 != 'PROD_TAG') &(df_srce_fct.col3 != 'PER_TAG'))

    logger.info("[extract_fact_sff] Selecting and renaming columns")

    selected_columns = [
        "col1 as mkt_extrn_code", "col2 as prod_extrn_code", "col3 as time_extrn_code"
    ]
    selected_columns += [f"col{i} as fact_amt_{i-3}" for i in range(4, 104)]
    selected_columns += [f"col{i}" for i in range(1, 104)]
    df_srce_mfct = df_srce_pre_mfct.selectExpr(*selected_columns)

    logger.info("[extract_fact_sff] Adding metadata columns")
    df_srce_mfct = df_srce_mfct.withColumn("line_num", row_number().over(Window.orderBy(monotonically_increasing_id())))
    df_srce_mfct = df_srce_mfct.withColumn("cntrt_id",lit(int(cntrt_id)))
    df_srce_mfct = df_srce_mfct.withColumn("srce_sys_id",lit(int(srce_sys_id)))
    df_srce_mfct = df_srce_mfct.withColumn("run_id",lit(int(run_id)))
    path=f"{raw_path}/{run_id}/Load_Product_df_invld_hier_prod"
    logger.info(f"[extract_fact_sff] df_invld_hier_prod: {path}")
    df_invld_hier_prod = spark.read.parquet(path)
    
    df_srce_mfct = df_srce_mfct.join(df_invld_hier_prod, df_invld_hier_prod.extrn_prod_id==df_srce_mfct.prod_extrn_code, 'left_anti')

    return df_srce_mfct

def extract_fact_tape(df_raw1,run_id,cntrt_id,dbutils,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,postgres_schema):
    logger = get_logger()
    raw_path = materialise_path(spark)
    logger.info("[extract_fact_tape] Filtering rows with valid fact indicators")
    df_raw1= df_raw1.withColumn('first65', trim(col('value')[0:65])).withColumn('count', size(split('first65', ' ')) - 1 ).withColumn('length', length('first65')).withColumn('fact_ind', (col('length')==65) & (col('count')==0)).filter('fact_ind = true')
    logger.info("[extract_fact_tape] Extracting structured columns from raw value")
    df_fact=df_raw1.withColumn("line_num", row_number().over(Window.orderBy(monotonically_increasing_id()))).withColumn("dummy_separator_1",substring(col("value"),1,15)).withColumn("mkt_extrn_code",substring(col("value"),16,6)).withColumn("prod_extrn_code",substring(col("value"),25,37)).withColumn("time_extrn_code",substring(col("value"),7,9)).withColumn("dummy_separator_2",substring(col("value"),16,46)).withColumn("mkt_extrn_code2",substring(col("value"),62,2)).withColumn("dummy_separator_3",substring(col("value"),64,2)).withColumn("fact_amt_1",substring(col("value"),66,11)).withColumn("fact_amt_2",substring(col("value"),77,11)).withColumn("fact_amt_3",substring(col("value"),88,11)).withColumn("fact_amt_4",substring(col("value"),99,11)).withColumn("fact_amt_5",substring(col("value"),110,11)).withColumn("fact_amt_6",substring(col("value"),121,11)).withColumn("fact_amt_7",substring(col("value"),132,11)).withColumn("fact_amt_8",substring(col("value"),143,11)).withColumn("fact_amt_9",substring(col("value"),154,11)).withColumn("fact_amt_10",substring(col("value"),165,11)).withColumn("fact_amt_11",substring(col("value"),176,11)).withColumn("fact_amt_12",substring(col("value"),187,11)).withColumn("fact_amt_13",substring(col("value"),198,11)).withColumn("fact_amt_14",substring(col("value"),209,11)).withColumn("fact_amt_15",substring(col("value"),220,11)).withColumn("fact_amt_16",substring(col("value"),231,11)).withColumn("fact_amt_17",substring(col("value"),242,11)).withColumn("fact_amt_18",substring(col("value"),253,11)).withColumn("fact_amt_19",substring(col("value"),264,11)).withColumn("fact_amt_20",substring(col("value"),275,11)).withColumn("other_fct_data",substring(col("value"),286,11)).drop("value").drop("first65").drop("count").drop("length").drop("fact_ind")
    logger.info("[extract_fact_tape] Trimming all columns")
    df_fact=df_fact.select(*[ltrim(rtrim(col)).alias(col)for col in df_fact.columns])

    df_srce_mfact = df_fact.withColumn("mkt_extrn_code",when(col("mkt_extrn_code") == 'N', concat(col("mkt_extrn_code2"), col("mkt_extrn_code"))).otherwise(col("mkt_extrn_code")))
        
    return df_srce_mfact

def extract_market_tape(df_raw1,run_id,cntrt_id,categ_id,srce_sys_id,strct_code):
    logger = get_logger()
    df_1=df_raw1.withColumn("ccc",col("value")[7:2]).filter("ccc=02")

    df_mark=df_1.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id()))).withColumn("dummy_separator",substring(  col("value"),0,10)).withColumn("extrn_code",substring(col("value"),11,6)).withColumn("extrn_name",substring(col("value"),49,80)).withColumn( "extrn_code_2",substring(col("value"),129,2)).withColumn("attr_code_list",substring(col("value"),131,120)).withColumn("attr_code_array",split("attr_code_list"," "))
    logger.info("[extract_market_tape] Splitting attr_code_array into individual columns")
    new_attr=[f"attr_code_{i+0}" for i in range(11)]
    df_mark=df_mark.select('*',*[col("attr_code_array").getItem(i).alias(new_attr[i]) for i in range(11)]).drop("attr_code_array").drop("value").drop("ccc")
    logger.info("[extract_market_tape] Concatenating extrn_code_2 and extrn_code")
    df_mark=df_mark.withColumn("extrn_code",concat(df_mark["extrn_code_2"],df_mark["extrn_code"]))
    logger.info("[extract_market_tape] Trimming Whitespaces from all columns")
    df_mark=df_mark.select(*[ltrim(rtrim(col)).alias(col)for col in df_mark.columns])
    logger.info("[extract_market_tape] Enriching final DataFrame with metadata values")
    df_srce_mmkt_tape = df_mark.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1) \
        .withColumn("mkt_match_attr_list", when (col("lvl_num") == 9, (concat_ws(' ',df_mark.attr_code_list,df_mark.extrn_name))).otherwise(df_mark.attr_code_list)) \
        .withColumn("extrn_code", col("extrn_code")) \
        .withColumn("extrn_mkt_id", col("extrn_code")) \
        .withColumn("extrn_mkt_attr_val_list", col("attr_code_list")) \
        .withColumn("extrn_mkt_name", col("extrn_name")) \
        .withColumn('categ_id', lit(categ_id)) \
        .withColumn('cntrt_id', lit(cntrt_id)) \
        .withColumn('srce_sys_id', lit(srce_sys_id)) \
        .withColumn('run_id', lit(run_id)) \
        .withColumn('strct_code', lit(strct_code))     

    return df_srce_mmkt_tape

def extract_market_ffs(df_raw1,run_id,cntrt_id,categ_id,srce_sys_id,strct_code):
    logger = get_logger()
    logger.info("[extract_market_ffs] Filtering rows starting with 'G'")
    df_markt_g=df_raw1.filter(col("value").startswith("G"))
    df_markt_g = df_markt_g.withColumn("line_num", row_number().over(Window.orderBy(monotonically_increasing_id())))
    df_mark=df_markt_g.withColumn("dummy_separator",expr("substring(value,0,1)")).withColumn("extrn_code",expr("substring(value,2,9)")).withColumn("extrn_name",expr("substring(value,26,40)")).withColumn("extrn_code_2",expr("substring(value,65,1)")).withColumn("attr_code_list",expr("substring(value,67,40)"))
    df_mark=df_mark.select(*[ltrim(rtrim(col)).alias(col)for col in df_mark.columns])
    df_mark=df_mark.withColumn("attr_code_array",split("attr_code_list"," "))
    new_attr=[f"attr_code_{i+0}" for i in range(11)]
    df_mark=df_mark.select('*',*[col("attr_code_array").getItem(i).alias(new_attr[i]) for i in range(11)]).drop("value").drop("attr_code_array")
    df_mark=df_mark.withColumn("extrn_code", expr("case 'N' when 'Y' then  extrn_code_2||extrn_code else extrn_code end"))
    logger.info("[extract_market_ffs] Trimming whitespace from all columns")
    df_mark=df_mark.select(*[ltrim(rtrim(col)).alias(col)for col in df_mark.columns])
    logger.info("[extract_market_ffs] Adding metadata columns values")
    df_srce_mmkt_ffs = df_mark.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1) \
        .withColumn("mkt_match_attr_list", when (col("lvl_num") == 9, (concat_ws(' ',df_mark.attr_code_list,df_mark.extrn_name))).otherwise(df_mark.attr_code_list)) \
        .withColumn("extrn_code", col("extrn_code")) \
        .withColumn("extrn_mkt_id", col("extrn_code")) \
        .withColumn("extrn_mkt_attr_val_list", col("attr_code_list")) \
        .withColumn("extrn_mkt_name", col("extrn_name")) \
        .withColumn('categ_id', lit(categ_id)) \
        .withColumn('cntrt_id', lit(cntrt_id)) \
        .withColumn('srce_sys_id', lit(srce_sys_id)) \
        .withColumn('run_id', lit(run_id)) \
        .withColumn('strct_code', lit(strct_code))     
    
    return df_srce_mmkt_ffs

def extract_market_ffs2(df_raw1,run_id,cntrt_id,categ_id,srce_sys_id,strct_code):
    logger = get_logger()
    logger.info("[extract_market_ffs2] Filtering rows starting with 'G'")
    df_markt_g=df_raw1.filter(col("value").startswith("G"))
    df_markt_g = df_markt_g.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id())))
    df_mark=df_markt_g.withColumn("dummy_separator",expr("substring(value,0,1)")).withColumn("extrn_code",expr("substring(value,2,9)")).withColumn("extrn_name",expr("substring(value,26,41)")).withColumn("extrn_code_2",expr("substring(value,66,1)")).withColumn("attr_code_list",expr("substring(value,67,112)"))
    df_mark=df_mark.select(*[ltrim(rtrim(col)).alias(col)for col in df_mark.columns])
    df_mark=df_mark.withColumn("attr_code_array",split("attr_code_list"," "))
    new_attr=[f"attr_code_{i+0}" for i in range(11)]
    df_mark=df_mark.select('*',*[col("attr_code_array").getItem(i).alias(new_attr[i]) for i in range(11)]).drop("value").drop("attr_code_array")
    df_mark=df_mark.withColumn("extrn_code", expr("case 'N' when 'Y' then  extrn_code_2||extrn_code else extrn_code end"))
    logger.info("[extract_market_ffs2] Trimming whitespace from all columns")
    df_mark=df_mark.select(*[ltrim(rtrim(col)).alias(col)for col in df_mark.columns])
    logger.info("[extract_market_ffs2] Adding metadata columns values")
    df_srce_mmkt_ffs2 = df_mark.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1) \
        .withColumn("mkt_match_attr_list", when (col("lvl_num") == 9, (concat_ws(' ',df_mark.attr_code_list,df_mark.extrn_name))).otherwise(df_mark.attr_code_list)) \
        .withColumn("extrn_code", col("extrn_code")) \
        .withColumn("extrn_mkt_id", col("extrn_code")) \
        .withColumn("extrn_mkt_attr_val_list", col("attr_code_list")) \
        .withColumn("extrn_mkt_name", col("extrn_name")) \
        .withColumn('categ_id', lit(categ_id)) \
        .withColumn('cntrt_id', lit(cntrt_id)) \
        .withColumn('srce_sys_id', lit(srce_sys_id)) \
        .withColumn('run_id', lit(run_id)) \
        .withColumn('strct_code', lit(strct_code))     
    
    return df_srce_mmkt_ffs2

def extract_market_sff(df_raw1,run_id,cntrt_id,categ_id,srce_sys_id,strct_code):
    logger = get_logger()
    logger.info("[extract_market_sff] Selecting and renaming relevant columns")
    df_srce_mkt = df_raw1.select(col("TAG"),col("SHORT"),col("LONG"),col("DISPLAY_ORDER"))
    df_srce_mkt = df_srce_mkt.withColumnRenamed("TAG", "extrn_code").withColumnRenamed("SHORT", "attr_code_list").withColumnRenamed("LONG", "extrn_name").withColumnRenamed("DISPLAY_ORDER", "line_num")
    logger.info("[extract_market_sff] Splitting attr_code_list and adding values to metadata columns")
    split_col = split(df_srce_mkt['attr_code_list'], ' ')
    df_srce_mmkt_sff = df_srce_mkt.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1)\
        .withColumn("mkt_match_attr_list", when (col("lvl_num") == 9, (concat_ws(' ',df_srce_mkt.attr_code_list,df_srce_mkt.extrn_name))).otherwise(df_srce_mkt.attr_code_list))\
        .withColumn("extrn_code", col("extrn_code")) \
        .withColumn("extrn_mkt_id", col("extrn_code")) \
        .withColumn("extrn_mkt_attr_val_list", col("attr_code_list")) \
        .withColumn("extrn_mkt_name", col("extrn_name")) \
        .withColumn('attr_code_0', split_col.getItem(0)) \
        .withColumn('attr_code_1', split_col.getItem(1)) \
        .withColumn('attr_code_2', split_col.getItem(2)) \
        .withColumn('attr_code_3', split_col.getItem(3)) \
        .withColumn('attr_code_4', split_col.getItem(4)) \
        .withColumn('attr_code_5', split_col.getItem(5)) \
        .withColumn('attr_code_6', split_col.getItem(6)) \
        .withColumn('attr_code_7', split_col.getItem(7)) \
        .withColumn('attr_code_8', split_col.getItem(8)) \
        .withColumn('attr_code_9', split_col.getItem(9)) \
        .withColumn('attr_code_10', split_col.getItem(10)) \
        .withColumn('categ_id', lit(categ_id)) \
        .withColumn('cntrt_id', lit(cntrt_id)) \
        .withColumn('srce_sys_id', lit(srce_sys_id)) \
        .withColumn('run_id', lit(run_id)) \
        .withColumn('strct_code', lit(strct_code))

    return df_srce_mmkt_sff

def extract_market_sff3(df_raw1,run_id,cntrt_id,categ_id,srce_sys_id,strct_code,spark):
    df_srce_mkt = df_raw1.select(col("market_id"),col("M_MKT_DISP_NAME"),col("MRKT_DSC_LNG"))
    windowspec  = Window.orderBy("market_id")
    df_srce_mkt=df_srce_mkt.withColumn("DISPLAY_ORDER",row_number().over(windowspec))
    df_srce_mkt.createOrReplaceTempView('temp_mkt')
    sql="""select market_id as extrn_code,M_MKT_DISP_NAME as attr_code_list,M_MKT_DISP_NAME as extrn_name,display_order as line_num from temp_mkt"""
    df_srce_mkt=spark.sql(sql)
    split_col = split(df_srce_mkt['attr_code_list'], ' ')
    df_srce_mmkt_sff3 = df_srce_mkt.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1)\
        .withColumn("mkt_match_attr_list", when (col("lvl_num") == 9, (concat_ws(' ',df_srce_mkt.attr_code_list,df_srce_mkt.extrn_name))).otherwise(df_srce_mkt.attr_code_list))\
        .withColumn("extrn_code", col("extrn_code")) \
        .withColumn("extrn_mkt_id", col("extrn_code")) \
        .withColumn("extrn_mkt_attr_val_list", col("attr_code_list")) \
        .withColumn("extrn_mkt_name", col("extrn_name")) \
        .withColumn('attr_code_0', split_col.getItem(0)) \
        .withColumn('attr_code_1', split_col.getItem(1)) \
        .withColumn('attr_code_2', split_col.getItem(2)) \
        .withColumn('attr_code_3', split_col.getItem(3)) \
        .withColumn('attr_code_4', split_col.getItem(4)) \
        .withColumn('attr_code_5', split_col.getItem(5)) \
        .withColumn('attr_code_6', split_col.getItem(6)) \
        .withColumn('attr_code_7', split_col.getItem(7)) \
        .withColumn('attr_code_8', split_col.getItem(8)) \
        .withColumn('attr_code_9', split_col.getItem(9)) \
        .withColumn('attr_code_10', split_col.getItem(10)) \
        .withColumn('categ_id', lit(categ_id)) \
        .withColumn('cntrt_id', lit(cntrt_id)) \
        .withColumn('srce_sys_id', lit(srce_sys_id)) \
        .withColumn('run_id', lit(run_id)) \
        .withColumn('strct_code', lit(strct_code))

    return df_srce_mmkt_sff3

def extract_measure_sff(df_raw1):
    
    df_srce_measr=df_raw1.withColumnRenamed('SHORT','extrn_name').withColumnRenamed('LONG','extrn_long_name').withColumnRenamed('DISPLAY_ORDER','line_num').withColumnRenamed('TAG','extrn_code').withColumnRenamed('PRECISION','precision_val').withColumnRenamed('DENOMINATOR','denominator_val')

    return df_srce_measr

def extract_measure_sff3(df_raw1):

    windowspec  = Window.orderBy("fact_description")
    df_srce_measr=df_raw1.withColumn("DISPLAY_ORDER",row_number().over(windowspec))
    df_srce_measr=df_srce_measr.withColumnRenamed('fact_description','extrn_name',).withColumnRenamed('fact_description','extrn_long_name').withColumnRenamed('DISPLAY_ORDER','line_num').withColumnRenamed('fact_column','TAG')
    df_srce_measr=df_srce_measr.withColumnRenamed('SHORT','extrn_name').withColumnRenamed('LONG','extrn_long_name').withColumnRenamed('DISPLAY_ORDER','line_num').withColumnRenamed('TAG','extrn_code')
    df_srce_measr=df_srce_measr.withColumn('precision_val',lit(0))

    return df_srce_measr

def extract_measure_ffs(df_raw1):

    df_raw1=df_raw1.withColumn("vendor",substring(col("value"),1,10)).withColumn("attribute",substring(col("value"),27,40)).withColumn("descrption",substring(col("value"),67,80)).drop("value")
    df_meas_o=df_raw1.filter(col("vendor").startswith("O"))
    df_meas_o=df_meas_o.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id())))
    df_meas=df_meas_o.withColumn("extrn_code",expr("substring(vendor,2,10)")).withColumnRenamed("descrption","extrn_name").drop("attribute").drop("vendor").withColumn("extrn_code",ltrim("extrn_code")).withColumn("extrn_code",rtrim("extrn_code")).withColumn("extrn_name",ltrim("extrn_name")).withColumn("extrn_name",rtrim("extrn_name"))
    
    return df_meas

def extract_measure_tape(df_raw1):

    df_1=df_raw1.withColumn("ccc",col("value")[7:2]).filter("ccc=04")
    df_meas=df_1.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id()))).withColumn("extrn_code",substring(col("value"),12,37)).withColumn("extrn_name",substring(col("value"),49,80)).drop("value").drop("ccc").withColumn("extrn_code",ltrim("extrn_code")).withColumn("extrn_code",rtrim("extrn_code")).withColumn("extrn_name",ltrim("extrn_name")).withColumn("extrn_name",rtrim("extrn_name"))
    
    return df_meas

def extract_time_tape(df_raw1):
    logger = get_logger()
    logger.info("[extract_time_tape] Function execution started.")
    df_1=df_raw1.withColumn("ccc",col("value")[7:2]).filter("ccc=01")
    logger.info("[extract_time_tape] Filtered df_1 with 'ccc=01'")
    df_time_r=df_1.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id()))).withColumn("extrn_code",substring(col("value"),9,9)).withColumn("extrn_name",substring(col("value"),49,80)).drop("value").drop("ccc")
    logger.info("[extract_time_tape] Extracted extrn_code and extrn_name")
    df_time_r=df_time_r.select(*[ltrim(rtrim(col)).alias(col)for col in df_time_r.columns])
    logger.info("[extract_time_tape] Trimmed all columns.")
    df_srce_mtime = df_time_r.withColumn("mm_time_perd_id", lit(0).cast("integer"))
    logger.info("[extract_time_tape] Final df_srce_mtime created")
    
    return df_srce_mtime

def extract_time_sff(df_raw1):
    logger = get_logger()
    logger.info("[extract_time_sff] Function execution started.")
    df_period_dim=df_raw1.withColumnRenamed('TAG','extrn_code').withColumnRenamed('LONG','extrn_name').withColumnRenamed('DISPLAY_ORDER','line_num')
    logger.info("[extract_time_sff] Renamed columns.")
    df_srce_mtime = df_period_dim.withColumn("mm_time_perd_id", lit(0).cast("integer"))
    logger.info("[extract_time_sff] Added mm_time_perd_id column. ")
    
    return df_srce_mtime

def extract_time_sff3(df_raw1):
    logger = get_logger()
    logger.info("[extract_time_sff3] Function execution started.")
    windowspec  = Window.orderBy("period_id")
    df_period_dim=df_raw1.withColumn("DISPLAY_ORDER",row_number().over(windowspec))
    logger.info("[extract_time_sff3] Assigned DISPLAY_ORDER using row_number")
    df_period_dim=df_period_dim.withColumnRenamed('period_id','extrn_code').withColumnRenamed('period_long_description','extrn_name').withColumnRenamed('DISPLAY_ORDER','line_num')
    logger.info("[extract_time_sff3] Renamed columns to extrn_code, extrn_name, and line_num.")
    df_srce_mtime = df_period_dim.withColumn("mm_time_perd_id", lit(0).cast("integer"))
    logger.info("[extract_time_sff3] Added mm_time_perd_id column.")
    
    return df_srce_mtime

def extract_time_ffs(df_raw1):
    logger = get_logger()
    logger.info("[extract_time_ffs] Function execution started.")
    df_raw1=df_raw1.withColumn("vendor",substring(col("value"),1,9)).withColumn("attribute",substring(col("value"),27,40)).withColumn("descrption",substring(col("value"),67,80)).drop("value")
    logger.info("[extract_time_ffs] Extracted vendor, attribute, and descrption.")
    df_time_r=df_raw1.filter(col("vendor").startswith("R"))
    logger.info("[extract_time_ffs] Filtered rows with vendor starting with 'R'.")
    df_time_r = df_time_r.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id())))
    df_time_r=df_time_r.withColumn("extrn_code",expr("substring(vendor,2,10)")).withColumnRenamed("descrption","extrn_name").drop("vendor").drop("attribute")
    logger.info("[extract_time_ffs] Added line_num and extrn_code.")
    df_time_r=df_time_r.select(*[ltrim(rtrim(col)).alias(col)for col in df_time_r.columns])
    logger.info("[extract_time_ffs] Trimmed all columns.")
    df_srce_mtime = df_time_r.withColumn("mm_time_perd_id", lit(0).cast("integer"))
    logger.info("[extract_time_ffs] Final df_srce_mtime created")
    
    return df_srce_mtime

def tier1_time_map(vendr_id,catalog_name,postgres_schema,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,run_id):
    logger = get_logger()
    raw_path = materialise_path(spark)
    logger.info("[tier1_time_map] Function execution started.")
    query=f""" SELECT * FROM {postgres_schema}.mm_time_perd_id_lkp where vendr_id='{vendr_id}'"""
    logger.info(f"[tier1_time_map] Constructed query: {query}")
    df_time_perd_id_lkp = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
        
    df_time_perd_id_lkp.createOrReplaceTempView('df_time_perd_id_lkp')
    df_srce_mtime = spark.read.parquet(f"{raw_path}/{run_id}/Load_Time_df_srce_mtime")
    df_srce_mtime.createOrReplaceTempView('df_srce_mtime')
    
    logger.info("[tier1_time_map] Executing SQL join to create df_time_map.")
    df_time_map = spark.sql(f"""
                            SELECT  input.*, time_lkp.time_perd_id, time_perd.time_perd_end_date as mm_time_perd_end_date 
                            FROM df_srce_mtime AS input 
                            LEFT OUTER JOIN df_time_perd_id_lkp AS time_lkp ON input.extrn_code = time_lkp.extrn_time_perd_id 
                            LEFT OUTER JOIN {catalog_name}.gold_tp.tp_time_perd_fdim AS time_perd ON time_lkp.time_perd_id = time_perd.time_perd_id
                            """)
    
        
    materialize(df_time_map,'Load_Time_df_time_map',run_id)

def extract_product_ffs(df_raw1,run_id,cntrt_id,categ_id,srce_sys_id,df_max_lvl):
    logger = get_logger() 
    df_raw1=df_raw1.filter(col("value").startswith("L"))
    logger.info("Filtered rows starting with 'L'")
    df_raw1 = df_raw1.withColumn('value', regexp_replace('value', '[\･]', '___'))
    df_raw1 = df_raw1.withColumn('value', regexp_replace('value', '[^a-zA-Z0-9 \\-\\!\\@\\#\\$\\%\\^\\&\\*\\(\\)\\[\\]\\{\\}\\;\\:\\,\\.\\/\\<\\>\\?\\|\\`\\~\\-\\"\'\\=\\_\\+\\n\\.]', '__'))
    logger.info("Special characters replaced in 'value' column")
    df_raw2=df_raw1.withColumn("vendor",substring(col("value"),2,9)).withColumn("attribute",substring(col("value"),27,40)).withColumn("descrption",substring(col("value"),67,40)).drop("value")
    df_split=df_raw2.withColumnRenamed("vendor","extrn_code").withColumnRenamed("descrption","attr_code_list").withColumnRenamed("attribute","extrn_name")
    logger.info("Extracted vendor, attribute, and description columns and Renamed columns")
    df_trim=df_split.select(*[ltrim(rtrim(col)).alias(col)for col in df_split.columns])
    df_trim=df_trim.withColumn("attr_code_array",split("attr_code_list"," "))
    new_attr=[f"attr_code_{i+0}" for i in range(11)]
    df_split=df_trim.select('*',*[col("attr_code_array").getItem(i).alias(new_attr[i]) for i in range(11)]).drop("attr_code_array").drop("vendor")
    df_trim=df_split.select(*[ltrim(rtrim(col)).alias(col)for col in df_split.columns])
    logger.info("Trimmed whitespace from all columns")
    df_split = df_trim.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id())))

    df_max_lvl = df_max_lvl.withColumn("max_lvl",col("max_lvl").cast(IntegerType())).withColumn('srce_sys_id',lit(srce_sys_id))

    df_split_ffs = df_split.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1) \
                           .withColumn("extrn_prod_id", col("extrn_code")) \
						   .withColumn("extrn_prod_attr_val_list", col("attr_code_list")) \
						   .withColumn("extrn_prod_name", col("extrn_name").cast("String")) \
						   .withColumn("extrn_name", col("extrn_name").cast("String")) \
						   .withColumn("prod_name", col("extrn_name")) \
						   .withColumn("prod_desc", col("extrn_name")) \
						   .withColumn('categ_id', lit(categ_id)) \
                           .withColumn('cntrt_id', lit(cntrt_id)) \
                           .withColumn('srce_sys_id', lit(srce_sys_id)) \
                           .withColumn('run_id', lit(run_id))                  

    df_join = df_split_ffs.join(df_max_lvl,df_split_ffs.srce_sys_id ==  df_max_lvl.srce_sys_id,"inner")
    cols = ("srce_sys_id","max_lvl")

    df_srce_mprod = df_join.withColumn('prod_match_attr_list',when(df_join.lvl_num == df_join.max_lvl,concat(df_join.attr_code_list,lit(' '),df_join.extrn_name)).otherwise(df_join.attr_code_list)).drop(*cols).withColumn('srce_sys_id', lit(srce_sys_id))

    return df_srce_mprod

def extract_product_ffs2(df_raw1,run_id,cntrt_id,categ_id,srce_sys_id,df_max_lvl):
    logger = get_logger()
    df_prod_l=df_raw1.filter(col("value").startswith("L"))
    logger.info("Filtered rows starting with 'L'")
    df_raw1 = df_prod_l.withColumn('value', regexp_replace('value', '[\�]', '_'))
    logger.info("Replaced special characters in 'value' column")
    df_raw1=df_raw1.withColumn("vendor",substring(col("value"),2,10)).withColumn("attribute",substring(col("value"),27,40)).withColumn("descrption",substring(col("value"),67,80)).drop("value")
    logger.info("Extracted vendor, attribute, and description columns")
    df_split=df_raw1.withColumnRenamed("vendor","extrn_code").withColumnRenamed("attribute","attr_code_list").withColumnRenamed("descrption","extrn_name")
    df_trim=df_split.select(*[ltrim(rtrim(col)).alias(col)for col in df_split.columns])
    logger.info("Trimmed whitespace from all columns")
    df_trim=df_trim.withColumn("attr_code_array",split("attr_code_list"," "))
    new_attr=[f"attr_code_{i+0}" for i in range(11)]
    df_split=df_trim.select('*',*[col("attr_code_array").getItem(i).alias(new_attr[i]) for i in range(11)]).drop("attr_code_array").drop("vendor")
    df_trim=df_split.select(*[ltrim(rtrim(col)).alias(col)for col in df_split.columns])
    df_split = df_trim.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id())))

    df_max_lvl = df_max_lvl.withColumn("max_lvl",col("max_lvl").cast(IntegerType())).withColumn('srce_sys_id',lit(srce_sys_id))
    df_split_ffs2 = df_split.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1) \
                            .withColumn("extrn_prod_id", col("extrn_code")) \
                            .withColumn("extrn_prod_attr_val_list", col("attr_code_list")) \
                            .withColumn("extrn_prod_name", col("extrn_name").cast("String")) \
                            .withColumn("extrn_name", col("extrn_name").cast("String")) \
                            .withColumn("prod_name", col("extrn_name")) \
                            .withColumn("prod_desc", col("extrn_name")) \
                            .withColumn('categ_id', lit(categ_id)) \
                            .withColumn('cntrt_id', lit(cntrt_id)) \
                            .withColumn('srce_sys_id', lit(srce_sys_id)) \
                            .withColumn('run_id', lit(run_id))
    df_join = df_split_ffs2.join(df_max_lvl,df_split_ffs2.srce_sys_id ==  df_max_lvl.srce_sys_id,"inner")
    cols = ("srce_sys_id","max_lvl")
    df_srce_mprod = df_join.withColumn('prod_match_attr_list',when(df_join.lvl_num == df_join.max_lvl,concat(df_join.attr_code_list,lit(' '),df_join.extrn_name)).otherwise(df_join.attr_code_list)).drop(*cols).withColumn('srce_sys_id', lit(srce_sys_id))

    return df_srce_mprod

def extract_product_sff(df_raw1, run_id, cntrt_id, categ_id, srce_sys_id, df_max_lvl):
    logger=get_logger()
    df_max_lvl = df_max_lvl.withColumn("max_lvl",col("max_lvl").cast(IntegerType())).withColumn('srce_sys_id',lit(srce_sys_id))
    df_srce_prod = df_raw1.select(col("TAG"),col("SHORT"),col("LONG"),col("DISPLAY_ORDER"))
    df_srce_prod = df_srce_prod.withColumnRenamed("TAG", "extrn_code").withColumnRenamed("LONG", "attr_code_list").withColumnRenamed("SHORT", "extrn_name").withColumnRenamed("DISPLAY_ORDER", "line_num")
    logger.info("Selected relevant columns from df_raw1")
    split_col = split(df_srce_prod['attr_code_list'], ' ')
    df_srce_prod = df_srce_prod.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1).withColumn("extrn_prod_id", col("extrn_code")).withColumn("extrn_prod_attr_val_list", col("attr_code_list")).withColumn("extrn_prod_name", col("extrn_name").cast("String")).withColumn("extrn_name", col("extrn_name").cast("String")).withColumn("prod_name", col("extrn_name")).withColumn("prod_desc", col("extrn_name")).withColumn('attr_code_0', split_col.getItem(0)).withColumn('attr_code_1', split_col.getItem(1)).withColumn('attr_code_2', split_col.getItem(2)).withColumn('attr_code_3', split_col.getItem(3)).withColumn('attr_code_4', split_col.getItem(4)).withColumn('attr_code_5', split_col.getItem(5)).withColumn('attr_code_6', split_col.getItem(6)).withColumn('attr_code_7', split_col.getItem(7)).withColumn('attr_code_8', split_col.getItem(8)).withColumn('attr_code_9', split_col.getItem(9)).withColumn('attr_code_10', split_col.getItem(10)).withColumn('categ_id', lit(categ_id)).withColumn('cntrt_id', lit(cntrt_id)).withColumn('srce_sys_id', lit(srce_sys_id)).withColumn('run_id', lit(run_id))
    
    df_join = df_srce_prod.join(df_max_lvl,df_srce_prod.srce_sys_id ==  df_max_lvl.srce_sys_id,"inner")
    cols = ("srce_sys_id","max_lvl")
    df_srce_mprod = df_join.withColumn('prod_match_attr_list',when(df_join.lvl_num == df_join.max_lvl,concat(df_join.attr_code_list,lit(' '),df_join.extrn_name)).otherwise(df_join.attr_code_list)).drop(*cols).withColumn('srce_sys_id', lit(srce_sys_id))

    return df_srce_mprod

def extract_product_sff2(df_raw1, run_id, cntrt_id, categ_id, srce_sys_id, df_max_lvl):
    logger = get_logger()
    df_max_lvl = df_max_lvl.withColumn("max_lvl",col("max_lvl").cast(IntegerType())).withColumn('srce_sys_id',lit(srce_sys_id))
    df_srce_prod = df_raw1.select(col("TAG"),col("SHORT"),col("LONG"),col("DISPLAY_ORDER"))
    logger.info("Selected columns: TAG, SHORT, LONG, DISPLAY_ORDER")
    df_srce_prod = df_srce_prod.withColumnRenamed("TAG", "extrn_code").withColumnRenamed("SHORT", "attr_code_list").withColumnRenamed("LONG", "extrn_name").withColumnRenamed("DISPLAY_ORDER", "line_num")
    logger.info("Renamed columns to standardized names")
    split_col = split(df_srce_prod['attr_code_list'], ' ')
    df_srce_prod = df_srce_prod.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1).withColumn("extrn_prod_id", col("extrn_code")).withColumn("extrn_prod_attr_val_list", col("attr_code_list")).withColumn("extrn_prod_name", col("extrn_name").cast("String")).withColumn("extrn_name", col("extrn_name").cast("String")).withColumn("prod_name", col("extrn_name")).withColumn("prod_desc", col("extrn_name")).withColumn('attr_code_0', split_col.getItem(0)).withColumn('attr_code_1', split_col.getItem(1)).withColumn('attr_code_2', split_col.getItem(2)).withColumn('attr_code_3', split_col.getItem(3)).withColumn('attr_code_4', split_col.getItem(4)).withColumn('attr_code_5', split_col.getItem(5)).withColumn('attr_code_6', split_col.getItem(6)).withColumn('attr_code_7', split_col.getItem(7)).withColumn('attr_code_8', split_col.getItem(8)).withColumn('attr_code_9', split_col.getItem(9)).withColumn('attr_code_10', split_col.getItem(10)).withColumn('categ_id', lit(categ_id)).withColumn('cntrt_id', lit(cntrt_id)).withColumn('srce_sys_id', lit(srce_sys_id)).withColumn('run_id', lit(run_id))

    df_join = df_srce_prod.join(df_max_lvl,df_srce_prod.srce_sys_id ==  df_max_lvl.srce_sys_id,"inner")
    cols = ("srce_sys_id","max_lvl")
    df_srce_mprod = df_join.withColumn('prod_match_attr_list',when(df_join.lvl_num == df_join.max_lvl,concat(df_join.attr_code_list,lit(' '),df_join.extrn_name)).otherwise(df_join.attr_code_list)).drop(*cols).withColumn('srce_sys_id', lit(srce_sys_id))

    return df_srce_mprod

def extract_product_sff3(df_raw1, run_id, cntrt_id, categ_id, srce_sys_id, df_max_lvl,spark):
    logger = get_logger()
    mft_path = f"{derive_base_path(spark)}/WORK"
    df_max_lvl=df_max_lvl.withColumn("max_lvl",col("max_lvl").cast(IntegerType())).withColumn('srce_sys_id',lit(srce_sys_id))
    df_max_lvl.show()
    hier_path=f'{mft_path}/*{run_id}*/*hierarchies*.txt'
    logger.info(f"Loaded hierarchy file from: {hier_path}")
    df_hier=spark.read.format('csv').option('header', True).option('delimiter', ',').load(hier_path)
    df_hier.show()
    df_hier.createOrReplaceTempView('hier_srce')
    df_invld_hier = spark.sql("""select distinct hierarchy_number from hier_srce 
                                group by hierarchy_number 
                                having count(*)=1 """)
    df_srce_prod=df_raw1.dropDuplicates()
    df_srce_prod.createOrReplaceTempView('temp_srce')
    for i in df_srce_prod.columns:
        if i=="PRDC_CD":
            sql="""select * except(PRDC_LNG_DSC),trim(concat({}," ",PRDC_LNG_DSC)) AS PRDC_LNG_DSC  from temp_srce""".format(i)
            df_srce_prod=spark.sql(sql)                            
    df_srce_prod.createOrReplaceTempView('temp_prod')
    df_srce_prod.show()
    logger.info(f" df_srce_prod Count: {df_srce_prod.count()}")
    df_hier.createOrReplaceTempView('temp_hier')
    df_hier=df_hier.select(("hierarchy_number")).distinct().collect()
    result_list = [row.hierarchy_number for row in df_hier]
    logger.info(f" Result_list: {result_list}")
    schema = StructType([
    StructField('TAG', StringType(), True),
    StructField('SHORT', StringType(), True),
    StructField('LONG', StringType(), True),
    StructField('DISPLAY_ORDER', StringType(), True),
    StructField('hierarchy_number', StringType(), True),
    StructField('hierarchy_level', StringType(), True)
    ])
    df = spark.createDataFrame([],schema)
    def srce_mprod(result_list):
        logger.info(f"result_list: {result_list}")
        arr=[1,2,3,4,5,6,7,8,9]
        l=[]
        for j  in arr:
            query="""select   column_name  from temp_hier where hierarchy_number= {} and hierarchy_level= {} AND column_name!='PRDC_CD_ITM_RNK' """.format(i,j)
            df_hier1 = [row[0] for row in spark.sql(query).collect()]
            l.extend(df_hier1)
        b = [str(l) for l in l]
        query_prod_cols="select"+" "+",".join(b)+" "+"from temp_prod"
        df_prod=spark.sql(query_prod_cols)
        df_prod=df_prod.columns
        nvl_concat=["nvl("+j+","+'""'+")" for j in df_prod]
        query_hier_levl="select product_id, hierarchy_number,hierarchy_level"+","+ "("+"nvl("+ i+ ","+"' '"+")"+"||"+"||".join(nvl_concat)+")"+"SHORT"+","+"("+"PRDC_LNG_DSC"+")"+"LONG"+" "+"from temp_prod WHERE hierarchy_number= {} ".format(i)
        add_space="||"+"' '"+"||"
        query_hier_levl=query_hier_levl.replace("||",add_space)
        df_fnl_hier_levl=spark.sql(query_hier_levl)
        df_fnl_hier_levl=df_fnl_hier_levl.distinct()
        df_fnl_hier_levl.createOrReplaceTempView('temp_fnl')
        query_fnl="""select nvl(product_id,"") as TAG,replace(trim(regexp_replace(SHORT,"  "," ")),'  ',' ') as SHORT,nvl(LONG,"")as LONG,row_number() over(order by product_id asc) as DISPLAY_ORDER,hierarchy_number,hierarchy_level from temp_fnl"""
        df_srce_prod=spark.sql(query_fnl)
        return df_srce_prod
    query_parts=[]
    for i in result_list:
        df_srce_prod=srce_mprod(i)
        view_name = f"df_union_{int(i)}"
        df_srce_prod.createOrReplaceTempView(view_name)  
        query_parts.append(f"SELECT * FROM {view_name}")

    df_union_query = " UNION ".join(query_parts)
    df_srce_prod=spark.sql(df_union_query)
    logger.info(f" df_srce_prod Count: {df_srce_prod.count()}")
    df_srce_prod.createOrReplaceTempView('temp_srce_prod')
    df_srce_prod=spark.sql("""select *except(SHORT),case when trim(LONG)<>""  then concat(SHORT," ","X") else SHORT end as SHORT  from temp_srce_prod""")
    df_srce_prod = df_srce_prod.withColumnRenamed("TAG", "extrn_code").withColumnRenamed("SHORT", "attr_code_list").withColumnRenamed("LONG", "extrn_name").withColumnRenamed("DISPLAY_ORDER", "line_num")
    split_col = split(df_srce_prod['attr_code_list'], ' ')
    df_srce_prod = df_srce_prod.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1).withColumn("extrn_prod_id", col("extrn_code")).withColumn("extrn_prod_attr_val_list", col("attr_code_list")).withColumn("extrn_prod_name", col("extrn_name").cast("String")).withColumn("extrn_name", col("extrn_name").cast("String")).withColumn("prod_name", col("extrn_name")).withColumn("prod_desc", col("extrn_name")).withColumn('attr_code_0', split_col.getItem(0)).withColumn('attr_code_1', split_col.getItem(1)).withColumn('attr_code_2', split_col.getItem(2)).withColumn('attr_code_3', split_col.getItem(3)).withColumn('attr_code_4', split_col.getItem(4)).withColumn('attr_code_5', split_col.getItem(5)).withColumn('attr_code_6', split_col.getItem(6)).withColumn('attr_code_7', split_col.getItem(7)).withColumn('attr_code_8', split_col.getItem(8)).withColumn('attr_code_9', split_col.getItem(9)).withColumn('attr_code_10', split_col.getItem(10)).withColumn('categ_id', lit(categ_id)).withColumn('cntrt_id', lit(cntrt_id)).withColumn('srce_sys_id', lit(srce_sys_id)).withColumn('run_id', lit(run_id))
    df_join = df_srce_prod.join(df_max_lvl,df_srce_prod.srce_sys_id ==  df_max_lvl.srce_sys_id,"inner")
    cols = ("srce_sys_id","max_lvl")
    df_join.createOrReplaceTempView("df_join_view")
    df_srce_mprod = spark.sql(""" SELECT *, 
                              CASE WHEN lvl_num = max_lvl THEN CONCAT(attr_code_list, ' ', extrn_name) 
                              ELSE attr_code_list END AS prod_match_attr_list 
                              FROM df_join_view """).drop(*cols).withColumn('srce_sys_id', lit(srce_sys_id))
    
    df_invld_hier_prod = df_srce_mprod.join(df_invld_hier, ["hierarchy_number"], 'inner').select('extrn_prod_id')
    logger.info("Displaying df_invld_hier_prod:")
    df_invld_hier_prod.show()
    df_srce_mprod = df_srce_mprod.join(df_invld_hier, ["hierarchy_number"], 'left_anti') 

    return df_srce_mprod,df_invld_hier_prod

def extract_product_tape(df_raw1,run_id,cntrt_id,categ_id,srce_sys_id,df_max_lvl):
    
    df_1=df_raw1.withColumn("ccc",col("value")[7:2]).filter("ccc=06")
    df_prod=df_1.withColumn("line_num",row_number().over(Window.orderBy(monotonically_increasing_id()))).withColumn("extrn_code",substring(col("value"),12,37)).withColumn("extrn_name",substring(col("value"),49,80)).withColumn( "attr_code_list",substring(col("value"),129,120)).withColumn("attr_code_array",split("attr_code_list"," "))
    new_attr=[f"attr_code_{i+0}" for i in range(11)]
    df_split=df_prod.select('*',*[col("attr_code_array").getItem(i).alias(new_attr[i]) for i in range(11)]).drop("attr_code_array").drop("value").drop("ccc").withColumn("extrn_name",ltrim("extrn_name")).withColumn("extrn_name",rtrim("extrn_name")).withColumn("extrn_code",ltrim("extrn_code")).withColumn("extrn_code",rtrim("extrn_code"))

    df_split=df_split.select(*[ltrim(rtrim(col)).alias(col)for col in df_split.columns])
    df_max_lvl = df_max_lvl.withColumn("max_lvl",col("max_lvl").cast(IntegerType())).withColumn('srce_sys_id',lit(srce_sys_id))

    df_split_tape = df_split.withColumn('lvl_num', size(split(col("attr_code_list"), r"\ ")) - 1) \
                           .withColumn("extrn_prod_id", col("extrn_code")) \
						   .withColumn("extrn_prod_attr_val_list", col("attr_code_list")) \
						   .withColumn("extrn_prod_name", col("extrn_name").cast("String")) \
						   .withColumn("extrn_name", col("extrn_name").cast("String")) \
						   .withColumn("prod_name", col("extrn_name")) \
						   .withColumn("prod_desc", col("extrn_name")) \
						   .withColumn('categ_id', lit(categ_id)) \
                           .withColumn('cntrt_id', lit(cntrt_id)) \
                           .withColumn('srce_sys_id', lit(srce_sys_id)) \
                           .withColumn('run_id', lit(run_id))	
    
    df_join = df_split_tape.join(df_max_lvl,df_split_tape.srce_sys_id ==  df_max_lvl.srce_sys_id,"inner")
    cols = ("srce_sys_id","max_lvl")
    df_srce_mprod = df_join.withColumn('prod_match_attr_list',when(df_join.lvl_num == df_join.max_lvl,concat(df_join.attr_code_list,lit(' '),df_join.extrn_name)).otherwise(df_join.attr_code_list)).drop(*cols).withColumn('srce_sys_id', lit(srce_sys_id))

    return df_srce_mprod

def derive_fact_sff(df_srce_mfct,spark):
    logger = get_logger()
    logger.info("[derive_fact_sff] Starting SFF fact derivation")
    df_srce_mfct.createOrReplaceTempView('srce_mfct')
    where_clause = " OR ".join([f"fact_amt_{i} is not null" for i in range(1, 101)])
    fact_cols = ", ".join([f"CASE WHEN substr(fact_amt_{i}, 1, 1) in ('.', ',') THEN 0 || fact_amt_{i} ELSE fact_amt_{i} END as fact_amt_{i}"for i in range(1, 101)])
    
    query = f"""
    SELECT line_num,cntrt_id,srce_sys_id,run_id,mkt_extrn_code,prod_extrn_code,time_extrn_code,
    {fact_cols} 
    FROM srce_mfct
    WHERE {where_clause}"""
    logger.info(f"[derive_fact_sff] Executing SQL query:\n{query}")
    return spark.sql(query)

def derive_fact_non_sff(df_srce_mfct,df_srce_mmeasr,run_id,cntrt_id,srce_sys_id,spark):
    logger = get_logger()
    logger.info("[derive_fact_non_sff] Started fact derivation")
    df_srce_mfct.createOrReplaceTempView('srce_mfct')
    df_srce_mmeasr.createOrReplaceTempView('srce_mmeasr')
    fact_columns = ", ".join([f"CAST(TRIM(REPLACE(REPLACE(REPLACE(UPPER(fact_amt_{i}), 'NA', ''), 'N/A', ''), 'N.A.', '')) AS DECIMAL(30,10)) AS fact_amt_{i}" for i in range(1, 21)])

    logger.info("Executing SQL query to transform source data")
    df_fct_conv = spark.sql(f"""SELECT :cntrt_id AS cntrt_id, :srce_sys_id AS srce_sys_id, :run_id AS run_id, 
    {fact_columns}, mkt_extrn_code, CAST(line_num AS INT) AS line_num, dummy_separator_1, prod_extrn_code, time_extrn_code, dummy_separator_2, mkt_extrn_code2, dummy_separator_3, other_fct_data FROM srce_mfct """,{'cntrt_id':cntrt_id,'srce_sys_id':srce_sys_id,'run_id':run_id})

    lst_all_cols = df_fct_conv.columns
    lst_fct_cols = [c for c in lst_all_cols if 'fact_amt_' in c]
    lst_line_num = [c.replace('fact_amt_','') for c in lst_fct_cols]
    df_missing_measrs = df_srce_mmeasr.filter(~col("line_num").isin(lst_line_num)).select("line_num")
    lst_missing_line_num = [row['line_num'] for row in df_missing_measrs.select('line_num').collect()]
    lst_missing_measrs = ['fact_amt_'+str(c) for c in lst_missing_line_num]
    logger.info(f"Missing measures identified: {lst_missing_measrs}")
    lst_partition_cols = ["cntrt_id","srce_sys_id","run_id","mkt_extrn_code","other_fct_data","dummy_separator_1","prod_extrn_code","time_extrn_code","dummy_separator_2","mkt_extrn_code2","dummy_separator_3"]
    n = 1
    for i in lst_missing_measrs:
        fct_n = 'fact_amt_'+str(n)
        if n==20:
            n=1
        else:
            n = n+1
        fact_amt_i = int(i.replace('fact_amt_',''))
        line_num_n = np.ceil(fact_amt_i / 20)
        df_fct_conv = df_fct_conv.withColumn(i, nth_value(fct_n,line_num_n).over(Window.partitionBy(lst_partition_cols).orderBy("line_num").rowsBetween(Window.unboundedPreceding, Window.unboundedFollowing)))

    df_fct_conv.show()
    df_fct_conv.createOrReplaceTempView("df_fct_conv")
    
    df_cfl=spark.sql(""" 
        select *, 
        COUNT( line_num) OVER (PARTITION BY cntrt_id,srce_sys_id,run_id,mkt_extrn_code,dummy_separator_1,prod_extrn_code,time_extrn_code,dummy_separator_2,mkt_extrn_code2,dummy_separator_3,other_fct_data ORDER BY int(line_num) ASC ) row_mod 
        from df_fct_conv """)
    
    df_cfl.createOrReplaceTempView("df_fct_conv1")
    df_cfl.show()

    df_cfl=spark.sql(''' select * from df_fct_conv1 where row_mod =1 ''')
    df_cfl=df_cfl.na.replace('',None)
    fact_amt_col=[col_name for col_name in df_cfl.columns if 'fact_amt_' in col_name]
    expres=" OR ".join([f"{col_name} IS NOT NULL" for col_name in fact_amt_col])
    condition=expr(expres)
    df_fct_cfl=df_cfl.filter(condition)
    i=1
    for _ in range(100):
        if (f'fact_amt_{i}' not in df_fct_cfl.columns):
            df_fct_cfl = df_fct_cfl.withColumn(f'fact_amt_{i}', lit(None).cast(DecimalType(38,10)))
        i=i+1
    logger.info("[derive_fact_non_sff] Completed")
    return df_fct_cfl

def t1_get_attribute_codes(df_srce_mprod,postgres_schema,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    logger = get_logger()
    logger.info("[t1_get_attribute_codes] function initiated")
    # Load Attribute Metadata from PostgreSQL
    df_categ_strct_attr_assoc_vw = read_query_from_postgres(f"SELECT * FROM {postgres_schema}.mm_categ_strct_attr_assoc_vw",spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    logger.info("[t1_get_attribute_codes] Loaded mm_categ_strct_attr_assoc_vw.")
    df_categ_strct_attr_assoc_vw.show()
    df_srce_mprod.createOrReplaceTempView("df_srce_mprod")

    # Create Level-Specific Attribute Views (1 to 9)
    df_categ_strct_attr_assoc_vw_1 = (df_categ_strct_attr_assoc_vw.filter(df_categ_strct_attr_assoc_vw["lvl_num"] == 1).withColumnRenamed("attr_phys_name", "attr_phys_name_1").withColumnRenamed("attr_id", "attr_id_1").withColumnRenamed("attr_name", "attr_name_1"))
    df_categ_strct_attr_assoc_vw_2 = (df_categ_strct_attr_assoc_vw.filter(df_categ_strct_attr_assoc_vw["lvl_num"] == 2).withColumnRenamed("attr_phys_name", "attr_phys_name_2").withColumnRenamed("attr_id", "attr_id_2").withColumnRenamed("attr_name", "attr_name_2"))
    df_categ_strct_attr_assoc_vw_3 = (df_categ_strct_attr_assoc_vw.filter(df_categ_strct_attr_assoc_vw["lvl_num"] == 3).withColumnRenamed("attr_phys_name", "attr_phys_name_3").withColumnRenamed("attr_id", "attr_id_3").withColumnRenamed("attr_name", "attr_name_3"))
    df_categ_strct_attr_assoc_vw_4 = (df_categ_strct_attr_assoc_vw.filter(df_categ_strct_attr_assoc_vw["lvl_num"] == 4).withColumnRenamed("attr_phys_name", "attr_phys_name_4").withColumnRenamed("attr_id", "attr_id_4").withColumnRenamed("attr_name", "attr_name_4"))
    df_categ_strct_attr_assoc_vw_5 = (df_categ_strct_attr_assoc_vw.filter(df_categ_strct_attr_assoc_vw["lvl_num"] == 5).withColumnRenamed("attr_phys_name", "attr_phys_name_5").withColumnRenamed("attr_id", "attr_id_5").withColumnRenamed("attr_name", "attr_name_5"))
    df_categ_strct_attr_assoc_vw_6 = (df_categ_strct_attr_assoc_vw.filter(df_categ_strct_attr_assoc_vw["lvl_num"] == 6).withColumnRenamed("attr_phys_name", "attr_phys_name_6").withColumnRenamed("attr_id", "attr_id_6").withColumnRenamed("attr_name", "attr_name_6"))
    df_categ_strct_attr_assoc_vw_7 = (df_categ_strct_attr_assoc_vw.filter(df_categ_strct_attr_assoc_vw["lvl_num"] == 7).withColumnRenamed("attr_phys_name", "attr_phys_name_7").withColumnRenamed("attr_id", "attr_id_7").withColumnRenamed("attr_name", "attr_name_7"))
    df_categ_strct_attr_assoc_vw_8 = (df_categ_strct_attr_assoc_vw.filter(df_categ_strct_attr_assoc_vw["lvl_num"] == 8).withColumnRenamed("attr_phys_name", "attr_phys_name_8").withColumnRenamed("attr_id", "attr_id_8").withColumnRenamed("attr_name", "attr_name_8"))
    df_categ_strct_attr_assoc_vw_9 = (df_categ_strct_attr_assoc_vw.filter(df_categ_strct_attr_assoc_vw["lvl_num"] == 9).withColumnRenamed("attr_phys_name", "attr_phys_name_9").withColumnRenamed("attr_id", "attr_id_9").withColumnRenamed("attr_name", "attr_name_9"))
    # Prepare Product-Level Metadata View
    df_categ_strct_attr_assoc_prod_lvl_vw = (df_categ_strct_attr_assoc_vw.withColumnRenamed("strct_lvl_id", "prod_lvl_id").withColumnRenamed("attr_name", "prod_lvl_name").withColumn("row_chng_desc", F.lit(None).cast("string")))

    df_categ_strct_attr_assoc_vw_1.createOrReplaceTempView("df_categ_strct_attr_assoc_vw_1")
    df_categ_strct_attr_assoc_vw_2.createOrReplaceTempView("df_categ_strct_attr_assoc_vw_2")
    df_categ_strct_attr_assoc_vw_3.createOrReplaceTempView("df_categ_strct_attr_assoc_vw_3")
    df_categ_strct_attr_assoc_vw_4.createOrReplaceTempView("df_categ_strct_attr_assoc_vw_4")
    df_categ_strct_attr_assoc_vw_5.createOrReplaceTempView("df_categ_strct_attr_assoc_vw_5")
    df_categ_strct_attr_assoc_vw_6.createOrReplaceTempView("df_categ_strct_attr_assoc_vw_6")
    df_categ_strct_attr_assoc_vw_7.createOrReplaceTempView("df_categ_strct_attr_assoc_vw_7")
    df_categ_strct_attr_assoc_vw_8.createOrReplaceTempView("df_categ_strct_attr_assoc_vw_8")
    df_categ_strct_attr_assoc_vw_9.createOrReplaceTempView("df_categ_strct_attr_assoc_vw_9")
    df_categ_strct_attr_assoc_prod_lvl_vw.createOrReplaceTempView("df_categ_strct_attr_assoc_prod_lvl_vw")

    # Map Product Data with Attribute Metadata
    df_prod_gan = spark.sql("""
        SELECT  input.*, 
        attr1.attr_phys_name_1 as attr_phys_name_1, attr1.attr_id_1 as attr_id_1,attr1.attr_name_1 as attr_name_1,
        attr2.attr_phys_name_2 as attr_phys_name_2, attr2.attr_id_2 as attr_id_2, attr2.attr_name_2 as attr_name_2, 
        attr3.attr_phys_name_3 as attr_phys_name_3, attr3.attr_id_3 as attr_id_3, attr3.attr_name_3 as attr_name_3, 
        attr4.attr_phys_name_4 as attr_phys_name_4, attr4.attr_id_4 as attr_id_4, attr4.attr_name_4 as attr_name_4, 
        attr5.attr_phys_name_5 as attr_phys_name_5, attr5.attr_id_5 as attr_id_5, attr5.attr_name_5 as attr_name_5, 
        attr6.attr_phys_name_6 as attr_phys_name_6, attr6.attr_id_6 as attr_id_6, attr6.attr_name_6 as attr_name_6, 
        attr7.attr_phys_name_7 as attr_phys_name_7, attr7.attr_id_7 as attr_id_7, attr7.attr_name_7 as attr_name_7, 
        attr8.attr_phys_name_8 as attr_phys_name_8, attr8.attr_id_8 as attr_id_8, attr8.attr_name_8 as attr_name_8, 
        attr9.attr_phys_name_9 as attr_phys_name_9, attr9.attr_id_9 as attr_id_9, attr9.attr_name_9 as attr_name_9, 
        prod_lvl.prod_lvl_id as prod_lvl_id, prod_lvl.prod_lvl_name as prod_lvl_name, prod_lvl.row_chng_desc as row_chng_desc 
        FROM df_srce_mprod AS input 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_vw_1 AS attr1 ON input.attr_code_1 = attr1.categ_id and input.attr_code_0 = attr1.strct_num 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_vw_2 AS attr2 ON input.attr_code_1 = attr2.categ_id and input.attr_code_0 = attr2.strct_num 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_vw_3 AS attr3 ON input.attr_code_1 = attr3.categ_id and input.attr_code_0 = attr3.strct_num 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_vw_4 AS attr4 ON input.attr_code_1 = attr4.categ_id and input.attr_code_0 = attr4.strct_num 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_vw_5 AS attr5 ON input.attr_code_1 = attr5.categ_id and input.attr_code_0 = attr5.strct_num 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_vw_6 AS attr6 ON input.attr_code_1 = attr6.categ_id and input.attr_code_0 = attr6.strct_num 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_vw_7 AS attr7 ON input.attr_code_1 = attr7.categ_id and input.attr_code_0 = attr7.strct_num 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_vw_8 AS attr8 ON input.attr_code_1 = attr8.categ_id and input.attr_code_0 = attr8.strct_num 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_vw_9 AS attr9 ON input.attr_code_1 = attr9.categ_id and input.attr_code_0 = attr9.strct_num 
        LEFT OUTER JOIN df_categ_strct_attr_assoc_prod_lvl_vw AS prod_lvl ON input.attr_code_1 = prod_lvl.categ_id and input.attr_code_0 = prod_lvl.strct_num and input.lvl_num = prod_lvl.lvl_num
        """
    )
    
    logger.info(f"[t1_get_attribute_codes] Mapped product count: {df_prod_gan.count()}")
    logger.info("[t1_get_attribute_codes] Preview of df_prod_gan:")
    df_prod_gan.show(5)
    logger.info("[t1_get_attribute_codes] Extracting unique attribute column names")
    df_prod_gan.createOrReplaceTempView("df_assoc")

    str1 = """SELECT * , """

    # Extract all unique attribute names from the 9 attribute name columns
    lst_cols = spark.sql(
        """SELECT DISTINCT column_name FROM (
                            SELECT ATTR_PHYS_NAME_1 column_name FROM df_assoc
                            UNION ALL
                            SELECT ATTR_PHYS_NAME_2 column_name FROM df_assoc
                            UNION ALL
                            SELECT ATTR_PHYS_NAME_3 column_name FROM df_assoc
                            UNION ALL
                            SELECT ATTR_PHYS_NAME_4 column_name FROM df_assoc
                            UNION ALL
                            SELECT ATTR_PHYS_NAME_5 column_name FROM df_assoc
                            UNION ALL
                            SELECT ATTR_PHYS_NAME_6 column_name FROM df_assoc
                            UNION ALL
                            SELECT ATTR_PHYS_NAME_7 column_name FROM df_assoc
                            UNION ALL
                            SELECT ATTR_PHYS_NAME_8 column_name FROM df_assoc
                            UNION ALL
                            SELECT ATTR_PHYS_NAME_9 column_name FROM df_assoc
                            ) WHERE column_name IS NOT NULL"""
    )
    # Convert the result into a Python list of attribute names
    lst_cols = [row["column_name"] for row in lst_cols.collect()]
    str2 = ""
    logger.info(f"[t1_get_attribute_codes] Unique attributes: {lst_cols}")
    # Step 2: Build the dynamic SQL CASE statements for each unique attribute
    # Create a new column named after the attribute
    for i in lst_cols:
        str2 += f""" CASE WHEN attr_phys_name_1='{i}' THEN attr_code_1 
        WHEN attr_phys_name_2='{i}' THEN attr_code_2 
        WHEN attr_phys_name_3='{i}' THEN attr_code_3 
        WHEN attr_phys_name_4='{i}' THEN attr_code_4 
        WHEN attr_phys_name_5='{i}' THEN attr_code_5 
        WHEN attr_phys_name_6='{i}' THEN attr_code_6 
        WHEN attr_phys_name_7='{i}' THEN attr_code_7 
        WHEN attr_phys_name_8='{i}' THEN attr_code_8 
        WHEN attr_phys_name_9='{i}' THEN attr_code_9 end as {i} ,"""
    str2 = str2.rstrip(",")

    str3 = """FROM df_assoc"""
    query = str1 + str2 + str3
    logger.info(f"[t1_get_attribute_codes] dynamic SQL query for pivoting attribute codes:\n {query}")
    df_prod_nactc = spark.sql(query)
    
    logger.info(f"[t1_get_attribute_codes] Pivoted product count: {df_prod_nactc.count()}")
    logger.info("[t1_get_attribute_codes] Preview of df_prod_nactc")
    df_prod_nactc.show(5)
    logger.info("[t1_get_attribute_codes] Completed successfully")
    return df_prod_gan,df_prod_nactc

def t1_get_attribute_values(df_mm_prod_csdim,df_prod_gan,postgres_schema,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,fileformat,run_id):
    logger = get_logger()
    logger.info("[t1_get_attribute_values] Started")
    query = f"SELECT * FROM {postgres_schema}.mm_prod_attr_val_lkp WHERE use_ind='Y' "
    df_prod_attr_val_lkp = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    df_prod_attr_val_lkp.createOrReplaceTempView("df_prod_attr_val_lkp")
    df_prod_gan.createOrReplaceTempView("df_prod_gan")
    df_mm_prod_csdim.createOrReplaceTempView("df_mm_prod_csdim")
    logger.info("[t1_get_attribute_values] Loaded df_prod_attr_val_lkp")
    df_prod_attr_val_lkp.show()
    logger.info("[t1_get_attribute_values] Executing product input and attribute value mapping SQL query")
    df_prod_gav = spark.sql(""" 
        SELECT input.*, 
        attr1.prod_attr_val_code as prod_attr_val_code_1, attr1.categ_id as categ_id_1, attr1.prod_attr_val_name as prod_attr_val_name_1, 
        attr2.prod_attr_val_code as prod_attr_val_code_2, attr2.categ_id as categ_id_2, attr2.prod_attr_val_name as prod_attr_val_name_2, 
        attr3.prod_attr_val_code as prod_attr_val_code_3, attr3.categ_id as categ_id_3, attr3.prod_attr_val_name as prod_attr_val_name_3, 
        attr4.prod_attr_val_code as prod_attr_val_code_4, attr4.categ_id as categ_id_4, attr4.prod_attr_val_name as prod_attr_val_name_4, 
        attr5.prod_attr_val_code as prod_attr_val_code_5, attr5.categ_id as categ_id_5, attr5.prod_attr_val_name as prod_attr_val_name_5, 
        attr6.prod_attr_val_code as prod_attr_val_code_6, attr6.categ_id as categ_id_6, attr6.prod_attr_val_name as prod_attr_val_name_6, 
        attr7.prod_attr_val_code as prod_attr_val_code_7, attr7.categ_id as categ_id_7, attr7.prod_attr_val_name as prod_attr_val_name_7, 
        attr8.prod_attr_val_code as prod_attr_val_code_8, attr8.categ_id as categ_id_8, attr8.prod_attr_val_name as prod_attr_val_name_8, 
        csdim.prod_skid as prod_skid
        FROM df_prod_gan AS input
        LEFT OUTER JOIN df_prod_attr_val_lkp AS attr1 ON input.attr_id_1 = attr1.attr_id and input.attr_code_1 = attr1.categ_id and input.attr_code_1 like attr1.prod_attr_val_code
        LEFT OUTER JOIN df_prod_attr_val_lkp AS attr2 ON input.attr_id_2 = attr2.attr_id and input.attr_code_1 = attr2.categ_id and input.attr_code_2 like attr2.prod_attr_val_code
        LEFT OUTER JOIN df_prod_attr_val_lkp AS attr3 ON input.attr_id_3 = attr3.attr_id and input.attr_code_1 = attr3.categ_id and input.attr_code_3 like attr3.prod_attr_val_code
        LEFT OUTER JOIN df_prod_attr_val_lkp AS attr4 ON input.attr_id_4 = attr4.attr_id and input.attr_code_1 = attr4.categ_id and input.attr_code_4 like attr4.prod_attr_val_code
        LEFT OUTER JOIN df_prod_attr_val_lkp AS attr5 ON input.attr_id_5 = attr5.attr_id and input.attr_code_1 = attr5.categ_id and input.attr_code_5 like attr5.prod_attr_val_code
        LEFT OUTER JOIN df_prod_attr_val_lkp AS attr6 ON input.attr_id_6 = attr6.attr_id and input.attr_code_1 = attr6.categ_id and input.attr_code_6 like attr6.prod_attr_val_code
        LEFT OUTER JOIN df_prod_attr_val_lkp AS attr7 ON input.attr_id_7 = attr7.attr_id and input.attr_code_1 = attr7.categ_id and input.attr_code_7 like attr7.prod_attr_val_code
        LEFT OUTER JOIN df_prod_attr_val_lkp AS attr8 ON input.attr_id_8 = attr8.attr_id and input.attr_code_1 = attr8.categ_id and input.attr_code_8 like attr8.prod_attr_val_code
        LEFT OUTER JOIN df_mm_prod_csdim AS csdim ON input.extrn_prod_id = csdim.extrn_prod_id 
    """)
    if fileformat in ('SFF','SFF2','SFF3'):
        df_prod_gan = df_prod_gav
        materialize(df_prod_gan,'Product_Derivation_df_prod_gan',run_id)

    df_prod_gav = df_prod_gav.withColumn("attr_val_1", expr("case when instr(prod_attr_val_code_1, '%') > 0 then attr_code_1 else prod_attr_val_name_1 end")) \
                            .withColumn("attr_val_2", expr("case when instr(prod_attr_val_code_2, '%') > 0 then attr_code_2 else prod_attr_val_name_2 end")) \
                            .withColumn("attr_val_3", expr("case when instr(prod_attr_val_code_3, '%') > 0 then attr_code_3 else prod_attr_val_name_3 end")) \
                            .withColumn("attr_val_4", expr("case when instr(prod_attr_val_code_4, '%') > 0 then attr_code_4 else prod_attr_val_name_4 end")) \
                            .withColumn("attr_val_5", expr("case when instr(prod_attr_val_code_5, '%') > 0 then attr_code_5 else prod_attr_val_name_5 end")) \
                            .withColumn("attr_val_6", expr("case when instr(prod_attr_val_code_6, '%') > 0 then attr_code_6 else prod_attr_val_name_6 end")) \
                            .withColumn("attr_val_7", expr("case when instr(prod_attr_val_code_7, '%') > 0 then attr_code_7 else prod_attr_val_name_7 end")) \
                            .withColumn("attr_val_8", expr("case when instr(prod_attr_val_code_8, '%') > 0 then attr_code_8 else prod_attr_val_name_8 end")) \
                            .withColumn("attr_val_9", col("attr_code_9")) \
                            .withColumn("prod_skid", col("prod_skid"))
    logger.info("Preview of df_prod_gav")
    df_prod_gav.show()
    logger.info(f"[t1_get_attribute_values] Mapped product count: {df_prod_gav.count()}")
    materialize(df_prod_gav,'Product_Derivation_df_prod_gav',run_id)
    str1 = """SELECT * , """
    if fileformat in ('SFF','SFF2','SFF3'):
        lst_cols = ['pg_type_txt','pg_categ_txt','pg_mfgr_txt','pg_base_size_txt','pg_varnt','pg_brand_txt','pg_size_txt','pg_pkg_size','item']
    else:
        lst_cols = ['pg_categ_txt','pg_mfgr_txt','pg_base_size_txt','pg_varnt','pg_brand_txt','pg_boutq','pg_sub_brand_txt','item','pg_form_txt']

    str2=""
    for i in lst_cols:
        str2+=f""" 
        case 
        when lower(attr_phys_name_1)='{i}' then attr_val_1 
        when lower(attr_phys_name_2)='{i}' then attr_val_2 
        when lower(attr_phys_name_3)='{i}' then attr_val_3 
        when lower(attr_phys_name_4)='{i}' then attr_val_4 
        when lower(attr_phys_name_5)='{i}' then attr_val_5 
        when lower(attr_phys_name_6)='{i}' then attr_val_6 
        when lower(attr_phys_name_7)='{i}' then attr_val_7 
        when lower(attr_phys_name_8)='{i}' then attr_val_8 
        when lower(attr_phys_name_9)='{i}' then attr_val_9 
        end as {i} ,"""
    str2=str2.rstrip(",")

    df_prod_gav.createOrReplaceTempView('df_assoc1')
    str3 = """FROM df_assoc1"""
    query = str1+str2+str3
    logger.info(f"[t1_get_attribute_values] final pivot SQL Query: {query}")
    df_out=spark.sql(query)
    for i in df_out.columns:
        df_out = df_out.withColumnRenamed(i,i.lower())

    df_prod_navtc = df_out
    logger.info("[t1_get_attribute_values] Preview of df_prod_navtc:")
    df_prod_navtc.show()
    materialize(df_prod_navtc,'Product_Derivation_df_prod_navtc',run_id)
    logger.info("[t1_get_attribute_values] Completed successfully")
    return df_prod_gav

def t1_add_row_change_description(df_prod_nactc,df_mm_prod_csdim,spark):
    logger = get_logger()
    logger.info("[t1_add_row_change_description] Started")
    df_prod_nactc.createOrReplaceTempView('nactc')
    df_mm_prod_csdim.createOrReplaceTempView('csdim')

    lst_nactc_cols = [c for c in df_prod_nactc.columns if c not in ('row_chng_desc')]
    str_nactc_cols = ','.join(lst_nactc_cols)
    logger.info("[t1_add_row_change_description] Executing SQL query to derive row change description")
    # join input with csdim and update row change description based on matching logic
    df_prod_nactc = spark.sql(f"""
    WITH joined_data AS (
    SELECT
        n.*,
        c.extrn_prod_id AS extrn_prod_id_csdim,
        c.prod_match_attr_list AS prod_match_attr_list_csdim,
        c.extrn_prod_attr_val_list AS extrn_prod_attr_val_list_csdim,
        c.prod_name AS extrn_prod_name_csdim,
        c.prod_skid AS prod_skid_csdim
    FROM nactc n
    LEFT JOIN csdim c
        ON c.extrn_prod_id = n.extrn_prod_id
    )
    SELECT
    CASE
        WHEN n.row_chng_desc IS NOT NULL THEN n.row_chng_desc
        WHEN n.row_chng_desc IS NULL
            AND n.extrn_prod_id IS NOT NULL
            AND n.prod_match_attr_list IS NOT NULL
            AND extrn_prod_id_csdim IS NULL THEN 'NEW PRODUCT'
        WHEN n.extrn_prod_id = extrn_prod_id_csdim
            AND n.extrn_prod_attr_val_list = extrn_prod_attr_val_list_csdim
            AND n.extrn_prod_name != extrn_prod_name_csdim
            AND n.prod_lvl_name = 'ITEM' THEN 'ITEM NAME CHANGE'
        WHEN n.extrn_prod_id = extrn_prod_id_csdim AND n.prod_match_attr_list != prod_match_attr_list_csdim THEN 'EXTERNAL CODE MATCH'
        WHEN n.extrn_prod_id = extrn_prod_id_csdim AND n.prod_match_attr_list = prod_match_attr_list_csdim THEN 'FULL MATCH'
        ELSE NULL
    END AS row_chng_desc,
    prod_skid_csdim AS prod_skid,
    {str_nactc_cols}
    FROM joined_data n
    """)
    
    logger.info(f"[t1_add_row_change_description] Row change description column added to df_prod_nactc with {df_prod_nactc.count()} records")
    df_prod_nactc.show()

    return df_prod_nactc

def t1_normalise_product_attributes(df_prod_dsdim,df_prod_dsdim_cols,postgres_schema,run_id,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    logger = get_logger()
    logger.info("[t1_normalise_product_attributes] Started")
    logger.info("[t1_normalise_product_attributes] Converting column names to uppercase")
    df_prod_dsdim = convert_cols_to_upper(df_prod_dsdim)
    df_prod_dsdim.createOrReplaceTempView("PROD_CSDIM")
    logger.info("[t1_normalise_product_attributes] Loading lookup tables from PostgreSQL")
    query = f"SELECT * FROM {postgres_schema}.mm_prod_attr_val_lkp WHERE use_ind='Y' "
    perd_df = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    perd_df.createOrReplaceTempView("MM_PROD_ATTR_VAL_LKP")
    logger.info("Loaded mm_prod_attr_val_lkp:")
    perd_df.show()
    query = f"SELECT * FROM {postgres_schema}.mm_attr_lkp "
    attr_df = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    attr_df.createOrReplaceTempView("MM_ATTR_LKP")
    logger.info("Loaded mm_attr_lkp:")
    attr_df.show()
    query = f"SELECT * FROM {postgres_schema}.mm_strct_lvl_lkp"
    df_mm_strct_lvl_lkp = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    df_mm_strct_lvl_lkp.createOrReplaceTempView('mm_strct_lvl_lkp')
    logger.info("Loaded mm_strct_lvl_lkp:")
    df_mm_strct_lvl_lkp.show()
    df_lvls = spark.createDataFrame([(1, "1"),(2, "2"),(3, "3"),(4, "4"),(5, "5"),(6, "6"),(7, "7"),(8, "8"),(9, "9"),],["lvl_num", "label"] )
    df_lvls.createOrReplaceTempView('lvls')
    df_prod_dsdim.createOrReplaceTempView('input_nav')
    spark.sql('''create or replace temp view prod_lvls as
    select distinct prod_lvl_id, prod_lvl_name from input_nav''')
    spark.sql("""create or replace temp view lvl_mapping as
    WITH lvl
            AS ( SELECT * FROM lvls),
            mapping
            AS (SELECT NVL2 (a.attr_name,
                        '''' || a.attr_name || ''' ',
                        'NULL')
                        || ' as at'
                        || lvl.lvl_num
                        || ', '
                        || NVL (attr_phys_name, 'NULL')
                        || ' as av'
                        || lvl.lvl_num
                            AS map_stmt,
                        a.attr_name,
                        a.attr_phys_name,
                        lvl.lvl_num,
                        s2.lvl_num AS curr_lvl_num,
                        pl.prod_lvl_id,
                        pl.prod_lvl_name
                FROM mm_strct_lvl_lkp s1
                JOIN mm_attr_lkp a
                    ON (a.attr_id = s1.attr_id AND a.attr_name != 'ITEM')
                JOIN mm_strct_lvl_lkp s2
                    ON (    s1.strct_id = s2.strct_id)
                JOIN prod_lvls pl
                    ON (    pl.prod_lvl_id = s2.strct_lvl_id)
                RIGHT JOIN lvl
                    ON (lvl.lvl_num = cast(s1.lvl_num as INT)))
        SELECT    m.*
            FROM mapping m
            where lvl_num <= curr_lvl_num
            order by prod_lvl_id,lvl_num""")

    df_nav_output = spark.sql('''select distinct attr_phys_name from input_nav nav
    join lvl_mapping map
    on nav.prod_lvl_id = map.prod_lvl_id''')
    logger.info("Preview of df_nav_output")
    df_nav_output.show()
    dsdim_cols_lst = [i.upper() for i in df_prod_dsdim_cols.columns]
    attr_cols_lst = df_nav_output.select('attr_phys_name').toPandas()['attr_phys_name'].tolist()
    logger.info(f"[t1_normalise_product_attributes] Attribute columns: {attr_cols_lst}")
    diff_cols_lst = [col for col in dsdim_cols_lst if col not in attr_cols_lst]
    cols_lst = attr_cols_lst.copy()
    for col in diff_cols_lst:
        cols_lst.append(col)  
    sel_cols = []
    for i in cols_lst:
        if i in df_prod_dsdim.columns:
            sel_cols.append(i.upper())
    df_prod_dsdim = df_prod_dsdim.select(*sel_cols)
    # Select the attribute physical name in attribute lookup
    csdim_cols = df_prod_dsdim.columns
    attr_collect = attr_df.select('ATTR_PHYS_NAME').collect()
    # Find List of attribute columns
    attr_cols = []
    for i in attr_collect:
        if i[0] in csdim_cols:
            attr_cols.append(i[0])

    list_of_attr = attr_cols
    [i.upper() for i in list_of_attr]
    # Find List of non attribute columns
    non_attr_cols = list(set(csdim_cols)-set(attr_cols))
    list_non_attr_cols = non_attr_cols
    [i.upper() for i in list_non_attr_cols]
    logger.info("[t1_normalise_product_attributes] Building dynamic SQL query for attribute normalization")
    # Building Query
    st1 = ''
    for i in list_non_attr_cols:
        st1 = st1 + 'PROD_CSDIM.'+i + ','
    st2 = ''
    itr = 1
    for i in list_of_attr:
        if i!='ATTR_VAL':
            st2 = st2 + f"""CASE WHEN mavl{itr}.PROD_ATTR_VAL_NAME NOT LIKE '%/%%' ESCAPE '/' THEN NVL(mavl{itr}.PROD_ATTR_VAL_NAME, PROD_CSDIM.{i}) ELSE PROD_CSDIM.{i} END AS {i},  \n"""
        itr = itr+1
    st3 = ''
    itr2=1
    for i in list_of_attr:
        if i!='ATTR_VAL':
            st3 = st3 + f"""WHEN m{itr2}.ATTR_NAME = PROD_LVL_NAME THEN  CASE WHEN mavl{itr2}.PROD_ATTR_VAL_NAME NOT LIKE '%/%%' ESCAPE '/' THEN NVL(mavl{itr2}.PROD_ATTR_VAL_NAME, PROD_CSDIM.{i})  ELSE PROD_CSDIM.{i} END\n """
        itr2 = itr2+1
    st4 = ''
    itr3 = 1
    for i in list_of_attr:
        if i!='ATTR_VAL':
            st4 = st4 + f"""JOIN MM_ATTR_LKP m{itr3} ON m{itr3}.ATTR_PHYS_NAME='{i}'
                        LEFT JOIN MM_PROD_ATTR_VAL_LKP mavl{itr3}
                        ON
                        PROD_CSDIM.PG_CATEG_TXT = mavl{itr3}.CATEG_ID
                        AND m{itr3}.ATTR_ID = mavl{itr3}.ATTR_ID
                        AND PROD_CSDIM.{i} LIKE  mavl{itr3}.PROD_ATTR_VAL_CODE \n """
        itr3 = itr3+1
    query = f""" 
    select 
    {st1}\n
    {st2}
    CASE
    {st3} END ATTR_VAL FROM PROD_CSDIM
    {st4}
    """
    logger.info(f"[t1_normalise_product_attributes] SQL Query:\n{query}")
    df_prod_nav = spark.sql(query)
    df_prod_nav = convert_cols_to_lower(df_prod_nav)
    logger.info(f"[t1_normalise_product_attributes] Normalized product attributes for {df_prod_nav.count()} records")
    logger.info("[t1_normalise_product_attributes] Preview of df_prod_nav")
    df_prod_nav.show()
    materialize(df_prod_nav,'Product_Derivation_df_prod_nav',run_id)
    logger.info("[t1_normalise_product_attributes] Completed Succesfully")

def t1_generate_product_description(df_mm_prod_csdim,postgres_schema,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,run_id):
    logger = get_logger()
    raw_path = materialise_path(spark)
    logger.info("[t1_generate_product_description] Started")
    df_prod_nav = spark.read.parquet(f"{raw_path}/{run_id}/Product_Derivation_df_prod_nav")
    logger.info("[t1_generate_product_description] Loaded df_prod_nav")
    df_prod_nav.show()
    df_lvls = spark.createDataFrame([(1, "1"),(2, "2"),(3, "3"),(4, "4"),(5, "5"),(6, "6"),(7, "7"),(8, "8"),(9, "9"),],["lvl_num", "label"] )
    df_lvls.createOrReplaceTempView('lvls')
    df_prod_nav.createOrReplaceTempView('input_nav')
    query = f"SELECT * FROM {postgres_schema}.mm_attr_lkp "
    MM_ATTR_LKP = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    MM_ATTR_LKP.createOrReplaceTempView("MM_ATTR_LKP")
    query = f"SELECT * FROM {postgres_schema}.mm_strct_lvl_lkp"
    mm_strct_lvl_lkp = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_strct_lvl_lkp.createOrReplaceTempView('mm_strct_lvl_lkp')
    logger.info("[t1_generate_product_description] Loaded attribute and structure level lookups")
    spark.sql('''create or replace temp view prod_lvls as
    select distinct prod_lvl_id, prod_lvl_name from input_nav''')
    spark.sql("""create or replace temp view lvl_mapping as
    WITH lvl
            AS ( SELECT * FROM lvls),
            mapping
            AS (SELECT NVL2 (a.attr_name,
                        '''' || a.attr_name || ''' ',
                        'NULL')
                        || ' as at'
                        || lvl.lvl_num
                        || ', '
                        || NVL (attr_phys_name, 'NULL')
                        || ' as av'
                        || lvl.lvl_num
                            AS map_stmt,
                        a.attr_name,
                        a.attr_phys_name,
                        lvl.lvl_num,
                        s2.lvl_num AS curr_lvl_num,
                        pl.prod_lvl_id,
                        pl.prod_lvl_name
                FROM mm_strct_lvl_lkp s1
                JOIN MM_ATTR_LKP a
                    ON (a.attr_id = s1.attr_id AND a.attr_name != 'ITEM')
                JOIN mm_strct_lvl_lkp s2
                    ON (    s1.strct_id = s2.strct_id)
                JOIN prod_lvls pl
                    ON (    pl.prod_lvl_id = s2.strct_lvl_id)
                RIGHT JOIN lvl
                    ON (lvl.lvl_num = cast(s1.lvl_num as INT)))
        SELECT    m.*
            FROM mapping m
            where lvl_num <= curr_lvl_num
            order by prod_lvl_id,lvl_num""")
    spark.sql("""
    create or replace temp view input_nav_with_cols as
    with base_data as (
        select case when nav.prod_lvl_name != 'CATEGORY' and map.attr_name = 'CATEGORY' then null 
                    when nav.prod_lvl_name != 'COMPANY' and map.attr_name = 'COMPANY' then null
                    else attr_phys_name 
                    end attr_phys_name1,
        case when nav.prod_lvl_name != 'CATEGORY' and map.attr_name = 'CATEGORY' then 'categ_id' else attr_phys_name end attr_phys_name2,
        extrn_prod_id,
        lvl_num
        from input_nav nav
        join lvl_mapping map
        on nav.prod_lvl_id = map.prod_lvl_id
    ),
    cols_extrn_prod_id as (
        select 
            extrn_prod_id,
            concat_ws(',', 
                transform(
                    array_sort(
                        collect_list(struct(lvl_num, attr_phys_name1))
                    ),
                    x -> x.attr_phys_name1
                )
            ) as bus_name_1,
            concat_ws(',', 
                transform(
                    array_sort(
                        collect_list(struct(lvl_num, attr_phys_name2))
                    ),
                    x -> x.attr_phys_name2
                )
            ) as bus_name_2
        from base_data
        group by extrn_prod_id
    )
    select bus_name_1,ltrim(bus_name_1,',') bus_name_111,bus_name_2,nav1.*
    from input_nav nav1
    left join cols_extrn_prod_id c
    on c.extrn_prod_id = nav1.extrn_prod_id
    where nav1.prod_lvl_name != 'ITEM'""")

    df_input_nav_with_cols = spark.sql('''select * from input_nav_with_cols''')  
    logger.info("[t1_generate_product_description] Preview of input_nav_with_cols")
    df_input_nav_with_cols.show()
    df_tmp = df_input_nav_with_cols.select(col('prod_Desc').alias('prod_name_desc'),'extrn_prod_id').limit(0)

    lst_prod_lvls = spark.sql('''select distinct prod_lvl_id from input_nav_with_cols''').collect()

    for i in lst_prod_lvls:
        cols1 = spark.sql(f'''select distinct bus_name_1 from input_nav_with_cols where prod_lvl_id = {i[0]}''').collect()[0][0]
        cols2 = spark.sql(f'''select distinct bus_name_2 from input_nav_with_cols where prod_lvl_id = {i[0]}''').collect()[0][0]
        lst_part1_cols = cols1.split(',')
        lst_part2_cols = cols2.split(',')
        
        df_tmp = df_tmp.union(df_input_nav_with_cols.where(df_input_nav_with_cols["prod_lvl_id"]==i[0]).select(concat_ws('|',concat_ws(' ',*lst_part1_cols),concat_ws(' ',*lst_part2_cols)).alias('prod_name_desc_test'),'extrn_prod_id'))
    
    logger.info("[t1_generate_product_description] Preview of df_tmp")
    df_tmp.show()

    df_prod_gpnd = df_prod_nav.join(df_tmp,'extrn_prod_id','left')  
    logger.info("[t1_generate_product_description] Preview of df_prod_gpnd")
    df_prod_gpnd.show()
    materialize(df_prod_gpnd,'Product_Derivation_df_prod_gpnd',run_id)

    logger.info("[t1_generate_product_description] Adding product name and description")
    df_prod_gpnd = spark.read.parquet(f"{raw_path}/{run_id}/Product_Derivation_df_prod_gpnd")
    df_prod_gpnd.createOrReplaceTempView('df_prod_gpnd')
    df_prod_apnd=spark.sql(""" SELECT * EXCEPT (prod_name,prod_desc),attr_val AS all_prod_name, 
                       CASE prod_lvl_name WHEN 'ITEM' THEN prod_name ELSE split(prod_name_desc, '\\\\|')[0] 
                       END AS prod_name,
                       CASE prod_lvl_name WHEN 'ITEM' THEN prod_desc ELSE split(prod_name_desc, '\\\\|')[1] 
                       END AS prod_desc
                    FROM df_prod_gpnd """)
    df_prod_apnd = df_prod_apnd.distinct()
    logger.info("Preview of df_prod_apnd after adding product description:")
    df_prod_apnd.show()
    materialize(df_prod_apnd,'Product_Derivation_df_prod_apnd',run_id)

    logger.info("[t1_generate_product_description] Adding partition columns and staging view")
    df_prod_apnd = df_prod_apnd.withColumn("part_srce_sys_id", df_prod_apnd["srce_sys_id"]).withColumn("part_cntrt_id", df_prod_apnd["cntrt_id"])
    materialize(df_prod_apnd,'product_transformation_df_prod_stgng_vw',run_id)
    logger.info("[t1_generate_product_description] Applying column complementation")
    df_prod_apnd = column_complementer(df_prod_apnd, df_mm_prod_csdim)
    logger.info("Preview of df_prod_apnd after column complimenting:")
    df_prod_apnd.show()
    materialize(df_prod_apnd,'Product_Derivation_cc_df_prod_apnd',run_id)
    
    logger.info("[t1_generate_product_description] Completed successfully")

@retry_with_backoff()
def t1_product_publish(cntrt_id, run_id, srce_sys_id, spark, catalog_name):
    logger = get_logger()
    raw_path = materialise_path(spark)
    logger.info("[t1_product_publish] Started")
    df_prod_apnd = spark.read.parquet(f"{raw_path}/{run_id}/Product_Derivation_cc_df_prod_apnd")
    df_prod_apnd.createOrReplaceTempView('input')
    logger.info(f"[t1_product_publish] Products row count before publish: {df_prod_apnd.count()}")
    
    dim_type='TP_PROD_SDIM'
    prod_path = semaphore_generate_path(dim_type, srce_sys_id, cntrt_id)

    prod_check_path = semaphore_acquisition(run_id, prod_path, catalog_name, spark)
    logger.info(f"[t1_product_publish] Semaphore acquired for path: {prod_check_path}")
    try:
        logger.info("[t1_product_publish] Starting MERGE into tp_prod_sdim")
        query = f"""
        MERGE INTO {catalog_name}.internal_tp.tp_prod_sdim  sdim
        USING input temp
        ON sdim.prod_skid = temp.prod_skid
        AND sdim.part_srce_sys_id = {srce_sys_id}
        AND sdim.part_cntrt_id = {cntrt_id}
        WHEN MATCHED THEN
        UPDATE SET *
        WHEN NOT MATCHED THEN
        INSERT *
        """
        merge_df = spark.sql(query)
        logger.info("Updated records:")
        merge_df.show()
        logger.info("[t1_product_publish] MERGE operation completed")

    finally:
        logger.info("[t1_product_publish] Releasing semaphore lock")
        # Release the semaphore for the current run
        release_semaphore(catalog_name,run_id,prod_check_path, spark )
        logger.info("[t1_product_publish] Semaphore released and product published completed successfully")

def skid_service(df,input_col,table_name,data_provider_code,spark,dbutils,SkidClientv2,tenant_id,env,publisher_name,application_key):
    logger = get_logger()
    logger.info("[skid_service] Started")
    skid = SkidClientv2(
    spark, 
    publisher_name= publisher_name, 
    application_key= application_key,
    client_id= dbutils.secrets.get(scope="KeyVault_Scope" , key=f'az-sp-cdl-flexflow-tp-01-{env}-ID'),
    client_secret= dbutils.secrets.get(scope="KeyVault_Scope", key=f'az-sp-cdl-flexflow-tp-01-{env}-KEY'),
    tenant_id= tenant_id, 
    env= env,
    use_lookup_table= False
    )
    result_dataframe = skid.assign_surrogate_keys(
    input_df= df,
    natural_col_list= input_col, # Example: 'cust_id', 'project_id','extrn_mkt_id','extrn_prod_id', 'cntrt_id'
    table_name = table_name, #market_dim,prod_dim
    data_provider_code = data_provider_code)  #'TP_140_MKT','TP_20_PROD, T1: 'TP_EU_{CNTRY_CODE}' for MKT, 'TP_EU_{CATEGORY_ID}' for PROD
    logger.info("[skid_service] Skid Assignment completed")
    return result_dataframe

def derive_product_assoc(df_prod_dim, df_strct_lkp,spark):
    logger = get_logger()
    logger.info("Started deriving product assoc")
    # Register temp views
    df_prod_dim.createOrReplaceTempView("prod_dim")
    df_strct_lkp.createOrReplaceTempView("strct_lvl_lkp")

    # Derive prod_assoc
    query = """
        WITH deduped_prod_dim AS (
            SELECT * FROM (
            SELECT
                *,
                ROW_NUMBER() OVER (
                    PARTITION BY  srce_sys_id , prod_match_attr_list
                    ORDER BY run_id DESC
                ) AS Row_Number_Deduplicator_Column
            FROM prod_dim
        )
        WHERE Row_Number_Deduplicator_Column=1
        ),
        prod_dim_strct_lvl AS (
            SELECT  input.*, sll.lvl_num AS lvl_num 
            FROM deduped_prod_dim AS input  
            INNER JOIN strct_lvl_lkp AS sll 
            ON input.prod_lvl_id = sll.strct_lvl_id
        ),
        input_filtered AS (
            SELECT prod_match_attr_list, prod_skid
            FROM prod_dim
        ) ,
        derived_assoc AS (
        SELECT 
            1 AS seq_num,
            input_filtered.prod_match_attr_list AS child_prod_match_attr_list,
            SUBSTRING_INDEX(
                input_filtered.prod_match_attr_list, ' ', 
                CASE 
                    WHEN prod_dim_strct_lvl.lvl_num = 1 THEN 2 
                    ELSE prod_dim_strct_lvl.lvl_num 
                END
            ) AS parnt_prod_match_attr_list,
            prod_dim_strct_lvl.secure_group_key
        FROM input_filtered
        INNER JOIN prod_dim_strct_lvl 
        ON input_filtered.prod_skid = prod_dim_strct_lvl.prod_skid
        )
        SELECT derived_assoc.* FROM derived_assoc
        INNER JOIN (SELECT prod_match_attr_list FROM deduped_prod_dim) AS parent_check 
            ON derived_assoc.parnt_prod_match_attr_list = parent_check.prod_match_attr_list

    """
    # Execute and return result
    return spark.sql(query)

def convert_cols_to_lower(df):
  cols = df.columns
  cols_lower = [col.lower() for col in cols]
  df = df.select(*cols_lower)
  return df

def convert_cols_to_upper(df):
  cols = df.columns
  cols_upper = [col.upper() for col in cols]
  df = df.select(*cols_upper)
  return df

def load_measr_vendr_factr_lkp(postgres_schema,spark,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    df_measr_vendr_factr_lkp= read_query_from_postgres(f"SELECT * FROM {postgres_schema}.MM_MEASR_VENDR_FACTR_LKP",spark,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    return df_measr_vendr_factr_lkp

def load_measr_cntrt_factr_lkp(postgres_schema,spark,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    df_measr_cntrt_factr_lkp = read_query_from_postgres(f"SELECT * FROM {postgres_schema}.MM_MEASR_CNTRT_FACTR_LKP",spark,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    return df_measr_cntrt_factr_lkp

def load_cntrt_categ_cntry_assoc(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    query = f"""
    SELECT 
        ca.categ_id, 
        c.vendr_id, 
        c.srce_sys_id, 
        c.file_formt, 
        c.cntrt_code, 
        c.crncy_id, 
        cn.cntry_id
    FROM {postgres_schema}.mm_cntrt_lkp c
    JOIN {postgres_schema}.mm_cntrt_categ_assoc ca ON c.cntrt_id = ca.cntrt_id
    JOIN {postgres_schema}.mm_cntry_lkp cn ON c.cntry_name = cn.cntry_name
    WHERE c.cntrt_id = '{cntrt_id}'
    """
    result = read_query_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).collect()[0]
    return result

def load_measr_id_lkp(vendr_id,postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    query = f""" SELECT * FROM {postgres_schema}.MM_MEASR_ID_LKP WHERE vendr_id='{vendr_id}' """
    df_measr_id_lkp = read_query_from_postgres(query,spark, ref_db_jdbc_url,ref_db_name, ref_db_user, ref_db_pwd)
    return df_measr_id_lkp

def load_max_lvl(categ_id,postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    query = f"""
    SELECT ROUND(MAX(sll.lvl_num)) AS max_lvl
    FROM {postgres_schema}.mm_categ_strct_assoc csa
    JOIN {postgres_schema}.mm_strct_lvl_lkp sll ON sll.strct_id = csa.strct_id
    JOIN {postgres_schema}.mm_attr_lkp al ON al.attr_id = sll.attr_id
    WHERE csa.categ_id = '{categ_id}' AND csa.STRCT_NUM = '1' AND al.attr_name = 'ITEM'
    """
    df_max_lvl = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    return df_max_lvl

def load_mkt_skid_lkp(cntry_id,vendr_id,srce_sys_id,postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    query =f""" 
    SELECT * FROM {postgres_schema}.mm_mkt_skid_lkp WHERE srce_sys_id={srce_sys_id} AND vendr_id={vendr_id} AND cntry_id='{cntry_id}'"""
    df_mkt_skid_lkp = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    return df_mkt_skid_lkp
    
def exec_missing_measures(spark,df_srce_mmeasr,str_line_num):
    df_srce_mmeasr.createOrReplaceTempView('srce_mmeasr')
    return spark.sql(f"""select line_num from srce_mmeasr where line_num not in ({str_line_num})""")

@retry_with_backoff()
def merge_tbl(df_new,catalog_name, tgt_table, merge_expr, spark):
    logger = get_logger()
    df_new.createOrReplaceTempView("df_new")
    merge_sql = f"""
            MERGE INTO {catalog_name}.{tgt_table} tgt
            USING df_new src
            ON {merge_expr}
            WHEN MATCHED THEN UPDATE SET *
            WHEN NOT MATCHED THEN INSERT *
            WHEN NOT MATCHED BY SOURCE THEN DELETE
            """

    spark.sql(merge_sql)
    logger.info(f"Successfully merged the dataframe into {tgt_table}")

def mm_time_perd_assoc_tier1_vw(spark, mm_time_perd_assoc,mm_time_perd_fdim,mm_time_perd_assoc_type):
    query = """with CAL_TYPE_2 as (
    SELECT   2 CAL_TYPE_ID,
            assoc.TIME_PERD_ID_A,
            assoc.TIME_PERD_ID_B,
            assoc.TIME_PERD_ASSOC_TYPE_ID,
            TB.TIME_PERD_CLASS_CODE TIME_PERD_CLASS_CODE_B,
            TB.TIME_PERD_TYPE_CODE TIME_PERD_type_CODE_B,
            tb.time_perd_end_date time_perd_end_date_b,
            tb.time_perd_start_date time_perd_start_date_b,
            ta.TIME_PERD_CLASS_CODE TIME_PERD_CLASS_CODE_a,
            ta.TIME_PERD_TYPE_CODE TIME_PERD_type_CODE_a,
            ta.time_perd_end_date time_perd_end_date_a,
            ta.time_perd_start_date time_perd_start_date_a
        FROM mm_time_perd_assoc assoc
            JOIN mm_time_perd_fdim ta ON ta.time_perd_id = assoc.time_perd_id_a
            JOIN mm_time_perd_fdim tb ON tb.time_perd_id = assoc.time_perd_id_b
            JOIN mm_time_perd_assoc_type asty
                ON asty.time_perd_assoc_type_id = assoc.time_perd_assoc_type_id
        WHERE assoc.CAL_TYPE_ID=2),
    BIMTH_2_MTH AS (
    SELECT  2 CAL_TYPE_ID,
            assoc.TIME_PERD_ID_B TIME_PERD_ID_A,
            assoc.TIME_PERD_ID_A TIME_PERD_ID_B,
            assoc.TIME_PERD_ASSOC_TYPE_ID ,
            ta.TIME_PERD_CLASS_CODE TIME_PERD_CLASS_CODE_B,
            ta.TIME_PERD_TYPE_CODE TIME_PERD_type_CODE_B,
            ta.time_perd_end_date time_perd_end_date_B,
            ta.time_perd_start_date time_perd_start_date_B,
            TB.TIME_PERD_CLASS_CODE TIME_PERD_CLASS_CODE_A,
            TB.TIME_PERD_TYPE_CODE TIME_PERD_type_CODE_a,
            tb.time_perd_end_date time_perd_end_date_A,
            tb.time_perd_start_date time_perd_start_date_A
        FROM mm_time_perd_assoc assoc
            JOIN mm_time_perd_fdim ta ON ta.time_perd_id = assoc.time_perd_id_a
            JOIN mm_time_perd_fdim tb ON tb.time_perd_id = assoc.time_perd_id_b
            JOIN mm_time_perd_assoc_type asty
                ON asty.time_perd_assoc_type_id = assoc.time_perd_assoc_type_id
        WHERE
        (TA.TIME_PERD_TYPE_CODE in ('EB','OB')   AND TB.TIME_PERD_CLASS_CODE = 'MTH' AND TB.TIME_PERD_TYPE_CODE = 'MH')
    AND assoc.CAL_TYPE_ID =2 AND assoc.TIME_PERD_ASSOC_TYPE_ID=1
    ),
    MTH_2_BIMTH AS (
    SELECT  2 CAL_TYPE_ID,
            assoc.TIME_PERD_ID_A ,
            assoc.TIME_PERD_ID_B ,
            assoc.TIME_PERD_ASSOC_TYPE_ID ,
            ta.TIME_PERD_CLASS_CODE TIME_PERD_CLASS_CODE_A,
            ta.TIME_PERD_TYPE_CODE TIME_PERD_type_CODE_A,
            ta.time_perd_end_date time_perd_end_date_A,
            ta.time_perd_start_date time_perd_start_date_A,
            TB.TIME_PERD_CLASS_CODE TIME_PERD_CLASS_CODE_B,
            TB.TIME_PERD_TYPE_CODE TIME_PERD_type_CODE_B,
            tb.time_perd_end_date time_perd_end_date_B,
            tb.time_perd_start_date time_perd_start_date_B
        FROM mm_time_perd_assoc assoc
            JOIN mm_time_perd_fdim ta ON ta.time_perd_id = assoc.time_perd_id_a
            JOIN mm_time_perd_fdim tb ON tb.time_perd_id = assoc.time_perd_id_b
            JOIN mm_time_perd_assoc_type asty
                ON asty.time_perd_assoc_type_id = assoc.time_perd_assoc_type_id
        WHERE
        (TA.TIME_PERD_TYPE_CODE in ('EB','OB')   AND TB.TIME_PERD_CLASS_CODE = 'MTH' AND TB.TIME_PERD_TYPE_CODE = 'MH')
    AND assoc.CAL_TYPE_ID =2 AND assoc.TIME_PERD_ASSOC_TYPE_ID=1
    ),

    BW_2_EB AS (SELECT  2 CAL_TYPE_ID,
            ta.EVEN_BIMTH_TIME_PERD_ID TIME_PERD_ID_A,
            ta.TIME_PERD_ID TIME_PERD_ID_B,
            1 TIME_PERD_ASSOC_TYPE_ID ,
            ta.TIME_PERD_CLASS_CODE TIME_PERD_CLASS_CODE_B,
            ta.TIME_PERD_TYPE_CODE TIME_PERD_type_CODE_B,
            ta.time_perd_end_date time_perd_end_date_B,
            ta.time_perd_start_date time_perd_start_date_B,
            ta.EVEN_BIMTH_TIME_PERD_CLASS_COD TIME_PERD_CLASS_CODE_A,
            ta.EVEN_BIMTH_TIME_PERD_TYPE_CODE TIME_PERD_TYPE_CODE_A,
            ta.EVEN_BIMTH_END_DATE TIME_PERD_END_DATE_A,
            ta.EVEN_BIMTH_START_DATE TIME_PERD_START_DATE_A
        FROM mm_time_perd_fdim ta
        WHERE
        TA.TIME_PERD_TYPE_CODE in ('BW') AND  TA.EVEN_BIMTH_TIME_PERD_TYPE_CODE ='EB'
        ),

    BW_2_MTH AS (SELECT  2 CAL_TYPE_ID,
            ta.ODD_BIMTH_TIME_PERD_ID TIME_PERD_ID_A,
            ta.TIME_PERD_ID TIME_PERD_ID_B,
            1 TIME_PERD_ASSOC_TYPE_ID ,
            ta.TIME_PERD_CLASS_CODE TIME_PERD_CLASS_CODE_B,
            ta.TIME_PERD_TYPE_CODE TIME_PERD_type_CODE_B,
            ta.time_perd_end_date time_perd_end_date_B,
            ta.time_perd_start_date time_perd_start_date_B,
            ta.MTH_TIME_PERD_CLASS_CODE TIME_PERD_CLASS_CODE_A,
            ta.MTH_TIME_PERD_TYPE_CODE TIME_PERD_TYPE_CODE_A,
            ta.MTH_END_DATE TIME_PERD_END_DATE_A,
            ta.MTH_START_DATE TIME_PERD_START_DATE_A
                FROM mm_time_perd_fdim ta
        WHERE
        TA.TIME_PERD_TYPE_CODE in ('BW') AND  TA.EVEN_BIMTH_TIME_PERD_TYPE_CODE ='MTH'
        )
    SELECT TIME_PERD_ASSOC_TYPE_ID,CAL_TYPE_ID ,TIME_PERD_ID_A,TIME_PERD_CLASS_CODE_A,TIME_PERD_TYPE_CODE_A,TIME_PERD_START_DATE_A,TIME_PERD_END_DATE_A,
    TIME_PERD_ID_B,TIME_PERD_CLASS_CODE_B,TIME_PERD_TYPE_CODE_B,TIME_PERD_START_DATE_B,TIME_PERD_END_DATE_B
    FROM CAL_TYPE_2
    UNION
    SELECT TIME_PERD_ASSOC_TYPE_ID,CAL_TYPE_ID ,TIME_PERD_ID_A,TIME_PERD_CLASS_CODE_A,TIME_PERD_TYPE_CODE_A,TIME_PERD_START_DATE_A,TIME_PERD_END_DATE_A,
    TIME_PERD_ID_B,TIME_PERD_CLASS_CODE_B,TIME_PERD_TYPE_CODE_B,TIME_PERD_START_DATE_B,TIME_PERD_END_DATE_B
    FROM BIMTH_2_MTH
    UNION
    SELECT TIME_PERD_ASSOC_TYPE_ID,CAL_TYPE_ID ,TIME_PERD_ID_A,TIME_PERD_CLASS_CODE_A,TIME_PERD_TYPE_CODE_A,TIME_PERD_START_DATE_A,TIME_PERD_END_DATE_A,
    TIME_PERD_ID_B,TIME_PERD_CLASS_CODE_B,TIME_PERD_TYPE_CODE_B,TIME_PERD_START_DATE_B,TIME_PERD_END_DATE_B
    FROM BW_2_EB
    UNION
    SELECT TIME_PERD_ASSOC_TYPE_ID, CAL_TYPE_ID ,TIME_PERD_ID_A,TIME_PERD_CLASS_CODE_A,TIME_PERD_TYPE_CODE_A,TIME_PERD_START_DATE_A,TIME_PERD_END_DATE_A,
    TIME_PERD_ID_B,TIME_PERD_CLASS_CODE_B,TIME_PERD_TYPE_CODE_B,TIME_PERD_START_DATE_B,TIME_PERD_END_DATE_B
    FROM BW_2_MTH
    UNION
    SELECT TIME_PERD_ASSOC_TYPE_ID,CAL_TYPE_ID ,TIME_PERD_ID_A,TIME_PERD_CLASS_CODE_A,TIME_PERD_TYPE_CODE_A,TIME_PERD_START_DATE_A,TIME_PERD_END_DATE_A,
    TIME_PERD_ID_B,TIME_PERD_CLASS_CODE_B,TIME_PERD_TYPE_CODE_B,TIME_PERD_START_DATE_B,TIME_PERD_END_DATE_B
    FROM MTH_2_BIMTH"""
    df_mm_time_perd_assoc_tier1_vw= spark.sql(query)
    
    return df_mm_time_perd_assoc_tier1_vw

def find_bad_ascii(df,column):
    df_bad = df.filter(~col(column).rlike("^[\\x00-\\x7F]*$")).drop(column)
    return df_bad


def bad_kpi_check(spark,dq,df,kpi):

    schema = StructType([
        StructField("DQ",StringType(),True), 
        StructField("DIMENTIONTAG",StringType(),True),
        StructField("IMPACTED_COLUMN",StringType(),True),
        StructField("COLUMN_VALUE",StringType(),True)
    ])
    
    df.createOrReplaceTempView('df_materialized_tbl')
    cols = df.columns
    concat_expr = "CONCAT_WS(' ', " + ", ".join(cols) + ") AS concat_str"
    query = f"SELECT *, {concat_expr} FROM df_materialized_tbl"
    df_bad = spark.sql(query)

    # Created concat_str to check if there are any ascii characters present 
    df_bad = find_bad_ascii(df_bad,'concat_str')
    
    cl = df_bad.collect()
    
    data = []
    for i in cl:
        for key, value in i.asDict().items():
            if str(value).isascii()== False:
                data.append((dq,kpi,key,value))
    print(data)
    bad_kpi = spark.createDataFrame(data,schema)

    return bad_kpi



# ====================================Tier 1 File Structure Validation====================================
   

def tier1_file_structure_validation(validation_name,file_name,cntrt_id,run_id,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,postgres_schema,catalog_name):

    valdn_grp_id = 5
    logger = get_logger()
    
    # Gathering all the required variables
    df_cntrt_lkp=load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    srce_sys_id = df_cntrt_lkp.collect()[0].srce_sys_id
    time_perd_type_code = df_cntrt_lkp.collect()[0].time_perd_type_code
    cntry_name=df_cntrt_lkp.collect()[0].cntry_name
    tier1_vendor_id = df_cntrt_lkp.collect()[0].vendr_id
    file_format = df_cntrt_lkp.collect()[0].file_formt

    tier1_prod_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Product_df_srce_mprod")
    tier1_measr_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Measure_df_srce_mmeasr")
    tier1_mkt_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Market_df_srce_mmkt")
    tier1_time_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Time_df_srce_mtime")
    tier1_fct_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Fact_df_srce_mfact")

    # Create views
    tier1_prod_mtrlz_tbl.createOrReplaceTempView('tier1_prod_mtrlz_tbl')
    tier1_measr_mtrlz_tbl.createOrReplaceTempView('tier1_measr_mtrlz_tbl')
    tier1_mkt_mtrlz_tbl.createOrReplaceTempView('tier1_mkt_mtrlz_tbl')
    tier1_time_mtrlz_tbl.createOrReplaceTempView('tier1_time_mtrlz_tbl')
    tier1_fct_mtrlz_tbl.createOrReplaceTempView('tier1_fct_mtrlz_tbl')
    tier1_fct_mtrlz_tbl.createOrReplaceTempView('tier1_fact_mtrlz_tbl')

    mm_categ_strct_attr_assoc_vw_query = f'select * from {postgres_schema}.mm_categ_strct_attr_assoc_vw'
    mm_categ_strct_attr_assoc_vw = read_query_from_postgres(mm_categ_strct_attr_assoc_vw_query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_categ_strct_attr_assoc_vw = mm_categ_strct_attr_assoc_vw.withColumn("igrtd_layer_attr_phys_name", F.regexp_replace(mm_categ_strct_attr_assoc_vw["attr_phys_name"], "_TXT", ""))
    mm_categ_strct_attr_assoc_vw.createOrReplaceTempView('MM_CATEG_STRCT_ATTR_ASSOC_VW')

    mm_time_perd_fdim_vw = spark.sql(f'''select * from {catalog_name}.gold_tp.tp_time_perd_fdim''')
    mm_time_perd_fdim_vw.createOrReplaceTempView('mm_time_perd_fdim_vw')
    MM_CNTRT_TIME_PERD_TYPE_ASSOC = load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    
    MM_CNTRT_TIME_PERD_TYPE_ASSOC.createOrReplaceTempView("MM_CNTRT_TIME_PERD_TYPE_ASSOC")
    
    MM_CNTRT_TIER_EXTND_VW = MM_CNTRT_TIME_PERD_TYPE_ASSOC
    MM_CNTRT_TIER_EXTND_VW.createOrReplaceTempView("MM_CNTRT_TIER_EXTND_VW")

    mm_cntrt_tier1_atomic_vw = spark.sql("""select * from MM_CNTRT_TIER_EXTND_VW WHERE SRCE_SYS_ID=:srce_sys_id""", {"srce_sys_id":srce_sys_id})
    mm_cntrt_tier1_atomic_vw.createOrReplaceTempView('mm_cntrt_tier1_atomic_vw')

    mm_cntrt_categ_assoc_query = f"select * from {postgres_schema}.mm_cntrt_categ_assoc"
    mm_cntrt_categ_assoc = read_query_from_postgres(mm_cntrt_categ_assoc_query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_cntrt_categ_assoc.createOrReplaceTempView("mm_cntrt_categ_assoc")

    MM_TIME_PERD_ID_LKP_QUERY = f"select * from {postgres_schema}.mm_time_perd_id_lkp"
    MM_TIME_PERD_ID_LKP = read_query_from_postgres(MM_TIME_PERD_ID_LKP_QUERY,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    MM_TIME_PERD_ID_LKP.createOrReplaceTempView('MM_TIME_PERD_ID_LKP')


    # Validation Starts
    
    data = []
    
    # ---------------Bad fact data format Validation-------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Bad fact data format',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Bad fact data format',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ---------------Duplicated market code in the input file Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Duplicated market code in the input file',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Duplicated market code in the input file',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ---------------Duplicated measure in the input file Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Duplicated measure in the input file',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Duplicated measure in the input file',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ----------------Duplicated product attributes in the input file--------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Duplicated product attributes in the input file',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Duplicated product attributes in the input file',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # -----------------Duplicated product code in the input file Validation--------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Duplicated product code in the input file',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Duplicated product code in the input file',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Duplicated time period in the input file Validation------------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Duplicated time period in the input file',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Duplicated time period in the input file',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)


    # ------------------Fact data existence Validation------------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Fact data existence',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Fact data existence',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Fact uses supplier market tag not defined in the reference part Validation------------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Fact uses supplier market tag not defined in the reference part',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Fact uses supplier market tag not defined in the reference part',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Fact uses supplier product tag not defined in the reference part Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Fact uses supplier product tag not defined in the reference part',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Fact uses supplier product tag not defined in the reference part',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Fact uses supplier time tag not defined in the reference part Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Fact uses supplier time tag not defined in the reference part',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Fact uses supplier time tag not defined in the reference part',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Incorrect file format Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Incorrect file format',spark)
    query = df.collect()[0]['qry_txt']
    df_val = spark.sql(query)
    dq_val = validation(catalog_name,df_val,'Incorrect file format',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # # ------------------Incorrect time period type Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Incorrect time period type',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    df_val.show()
    dq_val = validation(catalog_name,df_val,'Incorrect time period type',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Market data existence Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Market data existence',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Market data existence',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Measure data existence Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Measure data existence',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Measure data existence',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    # ------------------Number of fact lines does not match number of measures Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Number of fact lines does not match number of measures',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Number of fact lines does not match number of measures',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # -------------------Product data existence Validation-----------------------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Product data existence',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Product data existence',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    # -------------------Time data existence Validation---------------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Time data existence',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Time data existence',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # -------------------Too many measures in the row Validation---------------------------
    if file_format not in ("SFF3","SFF2","SFF"):
        df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Too many measures in the row',spark)
        query = df.collect()[0]['qry_txt']
        dq_val = validation(catalog_name,query,'Too many measures in the row',run_id,spark,validation_name)
        print(dq_val)
        data.append(dq_val)

    # -------------------Category in the file is matching contract specification Validation----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Category in the file is matching contract specification',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Category in the file is matching contract specification',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    #  Create Dataframe Schema and Merge into tp_data_vldtn_rprt table
    merge_into_tp_data_vldtn_rprt_tbl(run_id,spark,data,valdn_grp_id,catalog_name,postgres_schema, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    # # Report Formation
    generate_report(run_id,cntrt_id,file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
    
    # # # Report Generation
    generate_formatting_report(run_id,catalog_name)

    # Email Trigger after File Structure Validation
    send_emails(spark,run_id, cntrt_id,catalog_name,postgres_schema, file_name,validation_name,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)


def tier1_reference_vendor_validation(validation_name,file_name,cntrt_id,run_id,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,postgres_schema,catalog_name):

    logger = get_logger()
    valdn_grp_id = 9
    # Reading mm_cntrt_lkp 
    mm_cntrt_lkp = load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    mm_cntrt_lkp.createOrReplaceTempView('mm_cntrt_lkp')
    tier1_srce_sys_id = mm_cntrt_lkp.collect()[0].srce_sys_id

    logger.info(f"tier1_srce_sys_id-------{tier1_srce_sys_id}")
    

    tier1_prod_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Product_df_srce_mprod")
    tier1_measr_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Measure_df_srce_mmeasr")
    tier1_mkt_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Market_df_srce_mmkt")
    tier1_time_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Time_df_srce_mtime")

    tier1_prod_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Product_df_srce_mprod")
    tier1_prod_mtrlz_tbl.createOrReplaceTempView('tier1_prod_mtrlz_tbl')

    tier1_fact_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Fact_df_srce_mfact")
    tier1_fact_mtrlz_tbl.createOrReplaceTempView('tier1_fact_mtrlz_tbl')

    tier1_prod_gav = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Product_Derivation_df_prod_gav")
    tier1_prod_gav.createOrReplaceTempView('tier1_prod_gav')

    mm_prod_sdim = spark.sql(f'''select * from {catalog_name}.internal_tp.tp_prod_sdim''')
    mm_prod_sdim.createOrReplaceTempView('mm_prod_sdim')

    mm_categ_strct_attr_assoc_vw = read_query_from_postgres(f'select * from {postgres_schema}.mm_categ_strct_attr_assoc_vw',spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_categ_strct_attr_assoc_vw = mm_categ_strct_attr_assoc_vw.withColumn("igrtd_layer_attr_phys_name", F.regexp_replace(mm_categ_strct_attr_assoc_vw["attr_phys_name"], "_TXT", ""))
    mm_categ_strct_attr_assoc_vw.createOrReplaceTempView('mm_categ_strct_attr_assoc_vw')
    print("------mm_categ_strct_attr_assoc_vw------")

    mm_categ_strct_assoc = read_query_from_postgres(f"select * from {postgres_schema}.mm_categ_strct_assoc",spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_categ_strct_assoc.createOrReplaceTempView('mm_categ_strct_assoc')

    mm_strct_lvl_lkp = read_query_from_postgres(f"select * from {postgres_schema}.mm_strct_lvl_lkp",spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_strct_lvl_lkp.createOrReplaceTempView('mm_strct_lvl_lkp')

    # Creating mm_strct_lkp
    mm_strct_lkp_query = f"select * from {postgres_schema}.mm_strct_lkp"
    mm_strct_lkp = read_query_from_postgres(mm_strct_lkp_query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_strct_lkp.createOrReplaceTempView('mm_strct_lkp')

    data = []

    # ------------------Unknown hierarchy for product-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Unknown hierarchy for product',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Unknown hierarchy for product',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    

    # ------------------Too many attributes for product hierarchy-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Too many attributes for product hierarchy',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Too many attributes for product hierarchy',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    # ------------------No parent in product hierarchy-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','No parent in product hierarchy',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'No parent in product hierarchy',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    # ------------------Duplicated product after mapping-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Duplicated product after mapping',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Duplicated product after mapping',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    

    # ------------------External code change Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','External code change',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'External code change',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    #  Create Dataframe Schema and Merge into tp_data_vldtn_rprt table
    merge_into_tp_data_vldtn_rprt_tbl(run_id,spark,data,valdn_grp_id,catalog_name,postgres_schema, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    # # Report Formation
    generate_report(run_id,cntrt_id,file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
    
    # # # Report Generation
    generate_formatting_report(run_id,catalog_name)

    # Calling send_email function to send emails to vendors
    send_emails(spark,run_id, cntrt_id,catalog_name,postgres_schema, file_name,validation_name,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    

def tier1_fyi_validation(validation_name, file_name, cntrt_id, run_id, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd, postgres_schema, catalog_name):

    logger = get_logger()

    # Validate inputs
    postgres_schema = sanitize_variable(postgres_schema)
    catalog_name = sanitize_variable(catalog_name)
    run_id = sanitize_variable(run_id)
    cntrt_id = sanitize_variable(cntrt_id)

    valdn_grp_id = 7
    mm_cntrt_lkp = load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_cntrt_lkp.createOrReplaceTempView('mm_cntrt_lkp')
    
    row = mm_cntrt_lkp.collect()[0]
    tier1_srce_sys_id = row.srce_sys_id
    tier1_vendr_id = row.vendr_id
    cntry_name_cntrct_id = row.cntry_name
    time_perd_type_code = row.time_perd_type_code
    logger.info(f"time_perd_type_code is {time_perd_type_code}")
    
    mm_cntrt_lkp.createOrReplaceTempView('MM_CNTRT_TIER_EXTND_VW')

    df_level = spark.createDataFrame([[i] for i in range(1, 9)], ['unpivot_row'])
    df_level.createOrReplaceTempView("df_level_vw")

    time_perd_class_code = time_perd_class_codes_new(time_perd_type_code,catalog_name,spark)
    time_prd = sanitize_variable(time_perd_class_code.lower())
    
    query = f"""SELECT c.categ_id, lkp.srce_sys_id, lkp.file_formt, lkp.cntrt_code,vendr_id, lkp.crncy_id FROM {postgres_schema}.mm_cntrt_lkp lkp JOIN {postgres_schema}.mm_cntrt_categ_assoc c ON lkp.cntrt_id = c.cntrt_id"""
    mm_cntrt_categ_assoc = read_query_from_postgres(query,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).collect()[0]
    
    tier1_categ_id = mm_cntrt_categ_assoc.categ_id
    logger.info(f"tier1_categ_id------{tier1_categ_id}")

    # Safe SQL query
    mm_cntry_lkp_query = f"SELECT cntry_id, cntry_name FROM {postgres_schema}.mm_cntry_lkp"
    tier_1_cntry_id_df = read_query_from_postgres(mm_cntry_lkp_query, spark, ref_db_jdbc_url,
                                                  ref_db_name, ref_db_user, ref_db_pwd)
    tier1_cntry_id = tier_1_cntry_id_df.filter(tier_1_cntry_id_df.cntry_name == cntry_name_cntrct_id).collect()[0].cntry_id
    logger.info(f"tier1_cntry_id----{tier1_cntry_id}")

    # Load and register views
    for table in ["tp_run_prttn_plc", "tp_time_perd_fdim", "tp_prod_sdim"]:
        if table =='tp_prod_sdim':
            df = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.{table}")
        else:    
            df = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.{table}")
        df.createOrReplaceTempView(table)

    # Load parquet files safely
    tier1_prod_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Product_df_srce_mprod")
    tier1_measr_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Measure_df_srce_mmeasr")
    tier1_mkt_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Market_df_srce_mmkt")
    tier1_time_mtrlz_tbl = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Load_Time_df_srce_mtime")
    tier1_prod_dsdim = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Product_Derivation_df_prod_dsdim")
    tier1_mkt_dsdim = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Market_Derivation_df_mkt_as")
    tier1_prod_gav = spark.read.parquet(f"{materialise_path(spark)}/{run_id}/Product_Derivation_df_prod_gav")
    tier1_prod_gav.createOrReplaceTempView('tier1_prod_gav')

    tier1_prod_mtrlz_tbl.createOrReplaceTempView('tier1_prod_mtrlz_tbl')
    tier1_measr_mtrlz_tbl.createOrReplaceTempView('tier1_measr_mtrlz_tbl')
    tier1_mkt_mtrlz_tbl.createOrReplaceTempView('tier1_mkt_mtrlz_tbl')
    tier1_time_mtrlz_tbl.createOrReplaceTempView('tier1_time_mtrlz_tbl')
    tier1_mkt_dsdim.createOrReplaceTempView('tier1_mkt_dsdim')
    tier1_prod_dsdim.createOrReplaceTempView('tier1_prod_dsdim')
    # Parameterized Spark SQL
    spark.sql("SELECT * FROM MM_CNTRT_TIER_EXTND_VW WHERE SRCE_SYS_ID = :srce_sys_id",
              {'srce_sys_id': tier1_srce_sys_id})

    table_path = f"{catalog_name}.gold_tp.tp_{time_prd}_fct"
    mm_tp_fct = (spark.table(table_path).filter((F.col("srce_sys_id") == tier1_srce_sys_id) & (F.col("cntrt_id") == cntrt_id)))
    mm_tp_fct.createOrReplaceTempView(f"mm_tp_{time_prd}_fct")

    for table in ["tp_run_measr_plc", "tp_time_perd_assoc_type", "tp_time_perd_assoc"]:
        spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.{table}").createOrReplaceTempView(table)

    spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_run_mkt_plc where cntrt_id=:cntrt_id",{"cntrt_id":cntrt_id}).createOrReplaceTempView("mm_run_mkt_plc")

    # Load reference tables
    reference_tables = [
        "mm_comb_rule_prc_valdn_vw", "mm_measr_id_lkp", "mm_measr_lkp",
        "mm_time_perd_id_lkp", "mm_mkt_dim", "mm_run_plc"
    ]
    for table in reference_tables:
        query = f"SELECT * FROM {postgres_schema}.{table}"
        df = read_query_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
        df.createOrReplaceTempView(table)

    dpf_all_run_vw_query = f"SELECT run_id, run_sttus_name AS run_sttus_id, cntrt_id AS prcsg_id FROM {postgres_schema}.mm_run_plc"
    dpf_all_run_vw = read_query_from_postgres(dpf_all_run_vw_query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    dpf_all_run_vw.createOrReplaceTempView("dpf_all_run_vw")

    mm_time_perd_assoc=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_assoc')
    mm_time_perd_assoc.createOrReplaceTempView('mm_time_perd_assoc')

    mm_time_perd_assoc_type=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_assoc_type')
    mm_time_perd_assoc_type.createOrReplaceTempView('mm_time_perd_assoc_type')
    
    mm_time_perd_fdim=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_fdim')
    mm_time_perd_fdim.createOrReplaceTempView('mm_time_perd_fdim')

    mm_run_measr_plc=spark.sql(f"SELECT * from {catalog_name}.gold_tp.tp_run_measr_plc where cntrt_id=:cntrt_id",{"cntrt_id":cntrt_id})
    mm_run_measr_plc.createOrReplaceTempView('mm_run_measr_plc')

    mm_run_prttn_plc = spark.sql(f"select * from {catalog_name}.gold_tp.tp_run_prttn_plc where cntrt_id=:cntrt_id",{"cntrt_id":cntrt_id})
    mm_run_prttn_plc.createOrReplaceTempView('mm_run_prttn_plc')

    df_mm_time_perd_assoc_tier1_vw = mm_time_perd_assoc_tier1_vw(spark, mm_time_perd_assoc, mm_time_perd_fdim, mm_time_perd_assoc_type)
    df_mm_time_perd_assoc_tier1_vw.createOrReplaceTempView("mm_time_perd_assoc_tier1_vw")

    logger.info("Validations are starting")
    data = []
    # ------------------Measure description from the input file is different from expected Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Measure description from the input file is different from expected',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Measure description from the input file is different from expected',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Modified market description Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Modified market description',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Modified market description',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    # ------------------Missing product Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Missing product',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Missing product',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------The same brand placed under two companies Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','The same brand placed under two companies',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'The same brand placed under two companies',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Skipped measures Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Skipped measures',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Skipped measures',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ----------------------bad_product Validation---------------------
    bad_product = bad_kpi_check(spark,'Bad character in the input of the product file',tier1_prod_mtrlz_tbl,"Product")
    dq_val = validation(catalog_name,bad_product,'Bad character in the input of the product file',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # bad_market
    bad_market = bad_kpi_check(spark,'Bad character in the input of the market file',tier1_mkt_mtrlz_tbl,"Market")
    dq_val = validation(catalog_name,bad_market,'Bad character in the input of the market file',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # bad_measure
    bad_measure = bad_kpi_check(spark,'Bad character in the input of the measure file',tier1_measr_mtrlz_tbl,"Measure")
    dq_val = validation(catalog_name,bad_measure,'Bad character in the input of the measure file',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # bad_time
    bad_time = bad_kpi_check(spark,'Bad character in the input of the time file',tier1_time_mtrlz_tbl,"Time")
    dq_val = validation(catalog_name,bad_time,'Bad character in the input of the time file',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Invalid combination for attribute for product Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Invalid combination for attribute for product',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Invalid combination for attribute for product',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # -------------------New product list Validation------------------------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','New product list',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'New product list',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    logger.info("For Your Information Validations ------ DONE")

    #  Create Dataframe Schema and Merge into tp_data_vldtn_rprt table
    merge_into_tp_data_vldtn_rprt_tbl(run_id,spark,data,valdn_grp_id,catalog_name,postgres_schema, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    # # Report Formation
    generate_report(run_id,cntrt_id,file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
    
    # # # Report Generation
    generate_formatting_report(run_id,catalog_name)
    
def send_emails(spark,run_id, cntrt_id,catalog_name,postgres_schema, file_name,validation_name,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    
    logger = get_logger()
    # Send email 
    result = 'FAILED'
    query = f"""SELECT * FROM {catalog_name}.internal_tp.tp_valdn_rprt WHERE run_id=:run_id AND valdn_type =:validation_name AND reslt =:result"""
    df_tap_report_creation = spark.sql(query,{'run_id':run_id,'validation_name':validation_name,'result':result}
    )
    dbutils = get_dbutils(spark)

    df_cntrt_lkp=load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    cntrt_code=df_cntrt_lkp.collect()[0].cntrt_code

    if df_tap_report_creation.count()>0:
        try:
            # SMTP configuration
            smtp_port = 587
            smtp_server = dbutils.secrets.get('tp_dpf2cdl', 'cp-notification-smtp-server')
            email_from = dbutils.secrets.get('tp_dpf2cdl', 'cp-notification-email')
            pswd = dbutils.secrets.get('tp_dpf2cdl', 'cp-notification-password')


            # Fetch vendor contacts for File structure validation, reference data validation
            if validation_name in ['File Structure Validation', 'Reference Data Validations']:
                df = spark.sql(f"""select * from {catalog_name}.internal_tp.tp_notif_lkp""")
                emails = df.collect()[0]['email_id']
                
            else:
                df = read_query_from_postgres(f"""select vendr_email_addr_list from {postgres_schema}_consol.mmc_cntrt_detl_vw where cntrt_id = {cntrt_id}""",spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
                emails = df.collect()[0]['vendr_email_addr_list']
            
            # Fetch vendor contact emails (if present)
            flag_emails_present = True
            if emails:
                email_list = emails.split(';')
            else:
                flag_emails_present = False
                logger.info(f"no emails are found for contract code {cntrt_code}")

            # Fetch Run details
            mmc_dlvry_lkp_vw = read_from_postgres(f"{postgres_schema}.MM_RUN_PLC",spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
            mmc_dlvry_lkp_vw.createOrReplaceTempView("mmc_dlvry_lkp_vw")

            sql="""select dlvry_date_new as end_time from  
                            (select *,concat(date_format(left(end_time,10),'MMMM')," ",substr(end_time,9,2),",",left(end_time,4)) as dlvry_date_new from mmc_dlvry_lkp_vw where run_id =: run_id)"""
            mmc_dlvry_lkp_vw=spark.sql(sql, {'run_id':run_id})
            run_details=mmc_dlvry_lkp_vw.collect()
            end_time=run_details[0]['end_time']
            # Email content
            report_name = f"tp_dvm_rprt_{run_id}.xlsx"
            subject = f"Validation Report for run {run_id} and contract {cntrt_id} and contract code {cntrt_code}"

            # File path for attachment
            SAFE_DIR = f"/Volumes/{catalog_name}/internal_tp/tp-source-data"  # Define a safe base directory
            filename_path = f"/Volumes/{catalog_name}/internal_tp/tp-source-data/Validation_Reports/tp_dvm_rprt_{run_id}.xlsx"
            file_size = os.path.getsize(filename_path)
            file_size_mb=file_size//1048576
            logger.info(f" File Size in MB is {file_size_mb}")
            valdn_name_list = ['Reference Data Vendors Validations','File Structure Validation','Reference Data Validations','Business Validations']
            file_size_check_flag = True
            if validation_name in valdn_name_list and file_size_mb<=10:
                body = f"""
                <html>
                 <body>
                     <p>Validation summary:</p>
                     <table border="1" style="border-collapse: collapse; text-align: left;">
                         <tr><td>Source File Name</td><td>{file_name}</td></tr>
                         <tr><td>RUN ID</td><td>{run_id}</td></tr>
                         <tr><td>CONTRACT ID</td><td>{cntrt_id}</td></tr>
                         <tr><td>Contract Code</td><td>{cntrt_code}</td></tr>
                         <tr><td>End time</td><td>{end_time}</td></tr>
                     </table>
                     <p>Please find the attachment for the detailed validation report.</p>
                     <p>Best regards,<br>Tradepanel Team</p>
                 </body>
                 </html>
                 """
            elif file_size_mb>10:
                file_size_check_flag = False
                body = f"""
                <html>
                 <body>
                     <p>Validation summary:</p>
                     <table border="1" style="border-collapse: collapse; text-align: left;">
                         <tr><td>Source File Name</td><td>{file_name}</td></tr>
                         <tr><td>RUN ID</td><td>{run_id}</td></tr>
                         <tr><td>CONTRACT ID</td><td>{cntrt_id}</td></tr>
                         <tr><td>Contract Code</td><td>{cntrt_code}</td></tr>
                         <tr><td>End time</td><td>{end_time}</td></tr>
                     </table>
                     <p>Please  download attachments for detailed validation report  from dpf console delivery screen</p>
                     <p>Regards,<br>Tradepanel Team</p>
                 </body>
                 </html>
                 """
            if flag_emails_present:
                # Connect to SMTP server
                with smtplib.SMTP(smtp_server, smtp_port) as server:
                    server.starttls()
                    server.login(email_from, pswd)
                    print("Successfully connected to SMTP server.")
                    if emails:
                        for recipient in email_list:
                            msg = MIMEMultipart()
                            msg['From'] = email_from
                            msg['To'] = recipient
                            msg['Subject'] = subject
                            msg.attach(MIMEText(body, "html"))

                            # Resolve absolute path
                            abs_path = os.path.abspath(filename_path)
                            logger.info(f"---abs_path is {abs_path}")

                            # Check if the file is within the safe directory and is having size less than 10 mb
                            if file_size_check_flag:
                                if abs_path.startswith(os.path.abspath(SAFE_DIR)) and os.path.basename(abs_path) == report_name:
                                    with open(abs_path, 'rb') as attachment:
                                        part = MIMEBase('application', 'octet-stream')
                                        part.set_payload(attachment.read())
                                        encoders.encode_base64(part)
                                        part.add_header(
                                            'Content-Disposition',
                                            f'attachment; filename="{os.path.basename(report_name)}"'
                                        )
                                        msg.attach(part)
                                else:
                                    raise ValueError("Unsafe or invalid file path.")
                            else:
                                logger.info("File size exceeds limit, file attachment is not added")
                            # Send the email
                            server.sendmail(email_from, recipient, msg.as_string())
                            logger.info(f"Email sent to: {recipient}")
            else:
                logger.info("Email is not attached with this particular contract")
        except Exception as e:
            logger.info(f"Error sending email: {e}")

def tier1_business_validation(validation_name,file_name,cntrt_id,run_id,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,postgres_schema,catalog_name):

    logger = get_logger()
    valdn_grp_id = 8
    run_id = sanitize_variable(run_id)
    cntrt_id = sanitize_variable(cntrt_id)
    postgres_schema = sanitize_variable(postgres_schema)
    catalog_name = sanitize_variable(catalog_name)
    df_cntrt_lkp=load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd )
    df_cntrt_lkp.createOrReplaceTempView('df_cntrt_lkp')
    df_cntrt_categ_assoc =load_mm_cntrt_categ_assoc(spark,postgres_schema,cntrt_id,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    
    srce_sys_id=df_cntrt_lkp.collect()[0].srce_sys_id   
    cntrt_code=df_cntrt_lkp.collect()[0].cntrt_code   
    time_perd_type_code=df_cntrt_lkp.collect()[0].time_perd_type_code
    cntry_name=df_cntrt_lkp.collect()[0].cntry_name
    tier1_vendr_id=df_cntrt_lkp.collect()[0].vendr_id
    
    tier1_categ_id=df_cntrt_categ_assoc.collect()[0].categ_id
    
    time_perd_class_code=time_perd_class_codes_new(time_perd_type_code,catalog_name,spark)
    
    time_perd_class_code = sanitize_variable(time_perd_class_code)
    df = load_cntry_lkp(spark,postgres_schema,cntry_name,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    
    cntry_id=df.collect()[0].cntry_id
    
    path3 = 'tp-publish-data/'
    
    tier1_prod_gav=spark.read.parquet(f'{materialise_path(spark)}/{run_id}/Product_Derivation_df_prod_gav')
    tier1_prod_gav.createOrReplaceTempView('tier1_prod_gav')
    
    tier1_fct_skid=spark.read.parquet(f'{materialise_path(spark)}/{run_id}/Generate_Fact_Image_df_fct_skid')
    tier1_fct_skid.createOrReplaceTempView('tier1_fct_skid')      
    
    tier1_fct_smn_m=spark.read.parquet(f'{materialise_path(spark)}/{run_id}/Atomic_Measure_Calculations_df_fct_smn_m')
    tier1_fct_smn_m.createOrReplaceTempView('tier1_fct_smn_m')
    
    tier1_fct_dvm_100_measr=spark.read.parquet(f'{materialise_path(spark)}/{run_id}/Fact_Derivation_df_fct_dvm_100_measr')
    tier1_fct_dvm_100_measr.createOrReplaceTempView('tier1_fct_dvm_100_measr')
    
    tier1_measr_mtrlz_tbl=spark.read.parquet(f'{materialise_path(spark)}/{run_id}/Load_Measure_df_srce_mmeasr')
    tier1_measr_mtrlz_tbl.createOrReplaceTempView('tier1_measr_mtrlz_tbl')
    
    tier1_time_mtrlz_tbl=spark.read.parquet(f'{materialise_path(spark)}/{run_id}/Load_Time_df_srce_mtime')
    tier1_time_mtrlz_tbl.createOrReplaceTempView('tier1_time_mtrlz_tbl')
    
    tier1_prod_cxref=spark.sql(f'''SELECT extrn_prod_id,  extrn_prod_attr_val_list, prod_match_attr_list, prod_skid, srce_sys_id, cntrt_id, part_srce_sys_id FROM {catalog_name}.internal_tp.tp_prod_sdim WHERE cntrt_id=:cntrt_id''',{'cntrt_id':cntrt_id})
    tier1_prod_cxref.createOrReplaceTempView('tier1_prod_cxref')
    
    
    tier1_mkt_dsdim=spark.read.parquet(f'{materialise_path(spark)}/{run_id}/Market_Derivation_df_mkt_as')
    tier1_mkt_dsdim.createOrReplaceTempView('tier1_mkt_dsdim')
    
    # Dataframes from Pre-Business Validation Step

    tier1_dqm_cntrt_mkt = df_cntrt_lkp
    tier1_dqm_cntrt_mkt.createOrReplaceTempView('tier1_dqm_cntrt_mkt')
        
    tier1_dqm_calc_index=spark.read.parquet(f'{materialise_path(spark)}/{run_id}/t1_dq_pre_business_validation_df_dqm_calc_index')
    tier1_dqm_calc_index.createOrReplaceTempView('tier1_dqm_calc_index')
    

    mm_time_perd_assoc_type=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_assoc_type')
    mm_time_perd_assoc_type.createOrReplaceTempView('mm_time_perd_assoc_type')

    mm_time_perd_assoc=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_assoc')
    mm_time_perd_assoc.createOrReplaceTempView('mm_time_perd_assoc')

    mm_time_perd_fdim=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_fdim')
    mm_time_perd_fdim.createOrReplaceTempView('mm_time_perd_fdim')

    mm_prod_dim=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_prod_dim')


    mm_dvm_run_strct_lvl_plc= spark.sql(f'SELECT * FROM {catalog_name}.internal_tp.tp_valdn_run_strct_lvl_plc where run_id=:run_id',{"run_id":run_id})
    mm_dvm_run_strct_lvl_plc.createOrReplaceTempView('mm_dvm_run_strct_lvl_plc')
    
    mm_time_perd=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_fdim')
    mm_time_perd.createOrReplaceTempView('mm_time_perd')
    
    
    mm_dvm_run_strct_plc=spark.sql(f'SELECT * FROM {catalog_name}.internal_tp.tp_valdn_run_strct_plc where run_id=:run_id',{"run_id":run_id})
    mm_dvm_run_strct_plc.createOrReplaceTempView('mm_dvm_run_strct_plc')
    
    mm_run_prttn_plc=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_run_prttn_plc where cntrt_id=:cntrt_id',{"cntrt_id":cntrt_id})
    mm_run_prttn_plc.createOrReplaceTempView('mm_run_prttn_plc')
    
    table_path = f"{catalog_name}.gold_tp.tp_{time_perd_class_code}_fct"
    mm_tp_fct = (
        spark.table(table_path)
        .filter(
            (F.col("part_srce_sys_id") == srce_sys_id) &
            (F.col("part_cntrt_id") == cntrt_id)
        )
    )
    
    mm_tp_fct.createOrReplaceTempView('mm_tp_mth_fct')
    
    mm_prod_dim =  spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_prod_dim')
    mm_prod_dim.createOrReplaceTempView('mm_prod_dim')


    mm_mkt_dim=spark.sql(f'SELECT * FROM {catalog_name}.gold_tp.tp_mkt_dim')
    mm_mkt_dim.createOrReplaceTempView('mm_mkt_dim')
    
    mm_strct_lkp=read_from_postgres(f'{postgres_schema}.mm_strct_lkp', spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_strct_lkp.createOrReplaceTempView('mm_strct_lkp')
    
    mm_strct_lvl_lkp=read_from_postgres(f'{postgres_schema}.mm_strct_lvl_lkp', spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_strct_lvl_lkp.createOrReplaceTempView('mm_strct_lvl_lkp')
    
    mm_attr_lkp=read_from_postgres(f'{postgres_schema}.mm_attr_lkp', spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_attr_lkp.createOrReplaceTempView('mm_attr_lkp')
    
    dpf_all_run_vw=read_from_postgres(f'{postgres_schema}.mm_run_plc', spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    dpf_all_run_vw.createOrReplaceTempView('dpf_all_run_vw')
    
    mm_time_perd_id_lkp=read_from_postgres(f'{postgres_schema}.mm_time_perd_id_lkp', spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    mm_time_perd_id_lkp.createOrReplaceTempView('mm_time_perd_id_lkp')

    df_mm_time_perd_assoc_tier1_vw = mm_time_perd_assoc_tier1_vw(spark, mm_time_perd_assoc,mm_time_perd_fdim,mm_time_perd_assoc_type)
    df_mm_time_perd_assoc_tier1_vw.createOrReplaceTempView('mm_time_perd_assoc_tier1_vw')
    
    query = f"{postgres_schema}.mm_cntrt_valdn_assoc"
    mm_cntrt_checks_assoc = read_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd).filter(f"cntrt_id = {cntrt_id} and use_ind = 'Y'")
    checks = mm_cntrt_checks_assoc.select('valdn_id').collect()

    lst_of_checks=[]
    for i in checks:
        lst_of_checks.append(i[0].upper())
    
    chk_dq1 = 'CHK_DQ001' in lst_of_checks
    chk_dq2 = 'CHK_DQ002' in lst_of_checks
    chk_dq3 = 'CHK_DQ003' in lst_of_checks
    chk_dq4 = 'CHK_DQ004' in lst_of_checks
    chk_dq5 = 'CHK_DQ005' in lst_of_checks
    chk_dq6 = 'CHK_DQ006' in lst_of_checks
    chk_dq7 = 'CHK_DQ007' in lst_of_checks

    logger.info(f"Flags for business validation is {lst_of_checks}")  # True or False
    # Validations
    
    # #1 Unexpected change vs. previous period or year ago
    data = []
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Unexpected change vs. previous period or year ago',spark)
    query = df.collect()[0]['qry_txt']
    if chk_dq2:    
        dq_val = validation(catalog_name,query,'Unexpected change vs. previous period or year ago',run_id,spark,validation_name)
        print(dq_val)
        data.append(dq_val)

    #2 Unexpected backdata difference
    if chk_dq3:
        df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Unexpected backdata difference',spark)
        query = df.collect()[0]['qry_txt']
        dq_val = validation(catalog_name,query,'Unexpected backdata difference',run_id,spark,validation_name)
        print(dq_val)
        data.append(dq_val)

    # 3 Product Parent-Children mismatch on Volume Sales (MSU)
    if chk_dq4:
        df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Product Parent-Children mismatch on Volume Sales (MSU)',spark)
        query = df.collect()[0]['qry_txt']
        query2 = f"f'''{query}'''"
        e=eval
        query2 = e(query2)
        df_val = spark.sql(query2)
        dq_val = validation(catalog_name,df_val,'Product Parent-Children mismatch on Volume Sales (MSU)',run_id,spark,validation_name)
        print(dq_val)
        data.append(dq_val)

    # 4 - Top level Volume/Value Sales mismatch between hierarchies
    if chk_dq5:
        df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Top level Volume/Value Sales mismatch between hierarchies',spark)
        query = df.collect()[0]['qry_txt']
        dq_val = validation(catalog_name,query,'Top level Volume/Value Sales mismatch between hierarchies',run_id,spark,validation_name)
        print(dq_val)
        data.append(dq_val)

    #5 - Negative fact values
    if chk_dq7:
        df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Negative fact values',spark)
        query = df.collect()[0]['qry_txt']
        query2 = f"f'''{query}'''"
        e=eval
        query2 = e(query2)
        df_val = spark.sql(query2)
        dq_val = validation(catalog_name,df_val,'Negative fact values',run_id,spark,validation_name)
        print(dq_val)
        data.append(dq_val)

    # 6 - Missing fact data for parent product level
    if chk_dq6:
        df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Missing fact data for parent product level',spark)
        query = df.collect()[0]['qry_txt']
        query2 = f"f'''{query}'''"
        e=eval
        query2 = e(query2)
        df_val = spark.sql(query2)
        dq_val = validation(catalog_name,df_val,'Missing fact data for parent product level',run_id,spark,validation_name)
        print(dq_val)
        data.append(dq_val)

    # 7 - Missing/delivered product level
    if chk_dq1:
        df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Missing/delivered product level',spark)
        query = df.collect()[0]['qry_txt']
        query2 = f"f'''{query}'''"
        e=eval
        query2 = e(query2)
        dq_val = validation(catalog_name,query2,'Missing/delivered product level',run_id,spark,validation_name)
        print(dq_val)
        data.append(dq_val)


    logger.info("Business Validations ------ DONE")

    #  Create Dataframe Schema and Merge into tp_data_vldtn_rprt table
    merge_into_tp_data_vldtn_rprt_tbl(run_id,spark,data,valdn_grp_id,catalog_name,postgres_schema, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    # # Report Formation
    generate_report(run_id,cntrt_id,file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
    
    # # # Report Generation
    generate_formatting_report(run_id,catalog_name)

    # Final Report Send
    send_emails(spark,run_id, cntrt_id,catalog_name,postgres_schema, file_name,validation_name,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    

# Load postgres tables
def load_and_register_postgres_view(spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,query_template, view_name, params=None):
    logger = get_logger()
    try:
        if params:
            query = query_template.format(**params)
            
        else:
            query = query_template
        logger.info(f"Query currently running---- {query}")
        df = read_query_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
        
    except Exception as e:
        logger.error(f"Failed to load or register view {view_name}: {e}")
    return df

# Load and register materialized views safely
def safe_read_and_register(spark,path, view_name):
    logger = get_logger()
    try:
        df = spark.read.parquet(path)
    except AnalysisException as e:
        logger.error(f"Failed to read or register view {view_name}: {e}")
    return df

def load_and_register_catalog_view(spark,query_template, view_name, params):
    logger = get_logger()
    try:
        if params:
            df = spark.sql(query_template, params)
        else:
            df = spark.sql(query_template)
    except Exception as e:
        logger.error(f"Failed to load or register catalog view {view_name}: {e}")
    return df

def match_time_perd_class(time_perd):
    match time_perd:
        case 'mth':
            class_cd = 'mth'
        case 'wk':
            class_cd = 'wk'
        case 'bimth':
            class_cd = 'bimth'
        case 'qtr':
            class_cd = 'qtr'
        case 'custm':
            class_cd = 'custm'
    return class_cd

def read_fact(catalog_name,spark,time_perd, srce_sys_id, cntrt_id):
    class_cd = match_time_perd_class(time_perd)
    table_name = f"{catalog_name}.gold_tp.tp_{class_cd}_fct"
    srce_sys_id = sanitize_variable(srce_sys_id)
    cntrt_id = sanitize_variable(cntrt_id)
    df = spark.sql(f"select * from {table_name} where part_srce_sys_id = :srce_sys_id and part_cntrt_id = :cntrt_id", {"srce_sys_id": srce_sys_id,"cntrt_id": cntrt_id})
    return df

#======================= Reference Data Validation=====================
def tier1_reference_data_validation(validation_name, file_name, cntrt_id, run_id, spark,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd, postgres_schema, catalog_name):
    logger = get_logger()
    valdn_grp_id = 6
    # Validate inputs
    postgres_schema = sanitize_variable(postgres_schema)
    catalog_name = sanitize_variable(catalog_name)
    run_id = sanitize_variable(run_id)
    cntrt_id = sanitize_variable(cntrt_id)
    
    try:
        mm_cntrt_lkp = load_cntrt_lkp(cntrt_id, postgres_schema, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
        mm_cntrt_lkp.createOrReplaceTempView('mm_cntrt_lkp')
        cntrt_row = mm_cntrt_lkp.collect()[0]

        tier1_srce_sys_id = cntrt_row.srce_sys_id
        tier1_vendr_id = cntrt_row.vendr_id
        cntry_name_cntrct_id = cntrt_row.cntry_name
        time_perd_type_code = cntrt_row.time_perd_type_code
        tier1_vendr_id = sanitize_variable(tier1_vendr_id)
        
        # Load Postgres views
        postgres_views = {
            "mm_measr_id_lkp": f"SELECT * FROM {postgres_schema}.mm_measr_id_lkp",
            "mm_measr_lkp": f"SELECT * FROM {postgres_schema}.mm_measr_lkp",
            "mm_time_perd_id_lkp": f"SELECT * FROM {postgres_schema}.mm_time_perd_id_lkp",
            "mm_categ_strct_attr_assoc_vw": f"SELECT * FROM {postgres_schema}.mm_categ_strct_attr_assoc_vw",
            "MM_PROD_ATTR_VAL_LKP": f"SELECT * FROM {postgres_schema}.MM_PROD_ATTR_VAL_LKP",
            "mm_strct_lvl_lkp": f"SELECT * FROM {postgres_schema}.mm_strct_lvl_lkp",
            "MM_STRCT_LKP": f"SELECT * FROM {postgres_schema}.mm_strct_lkp",
            "MM_MEASR_VENDR_FACTR_LKP": f"SELECT * FROM {postgres_schema}.MM_MEASR_VENDR_FACTR_LKP WHERE VENDR_ID = {tier1_vendr_id}",
            "MM_MEASR_CNTRT_FACTR_LKP": f"SELECT * FROM {postgres_schema}.MM_MEASR_CNTRT_FACTR_LKP WHERE cntrt_id = {cntrt_id}",
            "mm_run_plc": f"SELECT * FROM {postgres_schema}.mm_run_plc WHERE cntrt_id = {cntrt_id}",
            "mm_mkt_xref": f"SELECT * FROM {postgres_schema}.mm_mkt_skid_lkp",
            "mm_time_perd": f"SELECT * FROM {postgres_schema}.mm_time_perd",
            "dpf_all_run_vw":f"SELECT * FROM {postgres_schema}.mm_run_plc where cntrt_id = {cntrt_id}"
        }

        for view, query in postgres_views.items():
            df = load_and_register_postgres_view(
                spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,query,view,{"postgres_schema": postgres_schema, "cntrt_id": cntrt_id, "tier1_vendr_id": tier1_vendr_id}
            )
            df.createOrReplaceTempView(view)

        # Load lookup tables securely
        def load_lookup_table(query, view_name):
            try:
                df = read_query_from_postgres(query, spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
                df.createOrReplaceTempView(view)
            except Exception as e:
                logger.error(f"Error loading {view_name}: {e}")
        
        # Secure query using parameterized logic
        tier_1_cntry_id_df = read_query_from_postgres(
            "SELECT cntry_id, cntry_name FROM {}.mm_cntry_lkp".format(postgres_schema),
            spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd
        )
        filtered_row = tier_1_cntry_id_df.filter(tier_1_cntry_id_df.cntry_name == cntry_name_cntrct_id).collect()[0]

        tier1_cntry_id = filtered_row.cntry_id
        print("tier1_cntry_id----", tier1_cntry_id)

        # Validate time_perd_type_code before using
        if time_perd_type_code:
            time_perd_class_code = time_perd_class_codes_new(time_perd_type_code,catalog_name,spark)
            time_prd = sanitize_variable(time_perd_class_code.lower())
            print("time_perd_class_code-----", time_prd)
        else:
            logger.warning("Missing time_perd_type_code")

        base_path = "{}/{}".format(materialise_path(spark),run_id)
        parquet_views = {
            "Load_Measure_df_srce_mmeasr": "tier1_measr_mtrlz_tbl",
            "Load_Product_df_srce_mprod": "tier1_prod_mtrlz_tbl",
            "Load_Market_df_srce_mmkt": "tier1_mkt_mtrlz_tbl",
            "Load_Time_df_srce_mtime": "tier1_time_mtrlz_tbl",
            "Market_Derivation_df_mkt_as": "derv_mkt_as",
            "Product_Derivation_df_prod_gav": "tier1_prod_gav",
            "Fact_Derivation_df_fct_dvm_100_measr": "tier1_fct_dvm_100_measr",
            "Load_Fact_df_srce_mfact": ["tier1_fct_mtrlz_tbl", "tier1_fact_mtrlz_tbl", "tier1_fct_mtrlz_tbl2"],
            "Fact_Derivation_df_measr_map": "tier1_measr_map"
        }

        for file, views in parquet_views.items():
            path = f"{base_path}/{file}"
            if isinstance(views, list):
                df = spark.read.parquet(path)
                for view in views:
                    df.createOrReplaceTempView(view)
            else:
                df = safe_read_and_register(spark,path, views)
                df.createOrReplaceTempView(views)

        # Handle missing columns in tier1_fct_conv
        tier1_fct_conv = spark.read.parquet(f"{base_path}/Fact_Derivation_df_fct_cfl")
        existing_cols = tier1_fct_conv.columns
        lst_fct_cols = [f"fact_amt_{i}" for i in range(1, 101)]
        missing_exprs = [f"CAST(NULL AS DECIMAL(38,10)) AS {col}" for col in lst_fct_cols if col not in existing_cols]

        if missing_exprs:
            sql_query = f"SELECT *, {', '.join(missing_exprs)} FROM tier1_fct_conv"
            tier1_fct_conv = spark.sql(sql_query)
            tier1_fct_conv.createOrReplaceTempView('tier1_fct_conv')
            
        else:
            tier1_fct_conv.createOrReplaceTempView('tier1_fct_conv')

        
        lookup_queries = {
            "mm_mkt_skid_lkp": "MM_MKT_SKID_LKP",
            "mm_mkt_dim": "mm_mkt_dim",
            "mm_measr_id_lkp": "mm_measr_id_lkp",
            "mm_measr_lkp": "mm_measr_lkp",
            "mm_time_perd_id_lkp": "mm_time_perd_id_lkp"
        }

        for table, view in lookup_queries.items():
            query = f"SELECT * FROM {postgres_schema}.{table}"
            load_lookup_table(query, view)

        mm_cntrt_lkp.createOrReplaceTempView('mm_cntrt_tier_extnd_vw')

        mm_run_prttn_plc = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_run_prttn_plc where cntrt_id=:cntrt_id",{"cntrt_id":cntrt_id})
        mm_run_prttn_plc.createOrReplaceTempView('mm_run_prttn_plc')

    except Exception as e:
        logger.error(f"Error in tier1_reference_data_validation: {e}")
        raise        
    
    mm_prod_sdim = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_prod_sdim WHERE srce_sys_id = :tier1_srce_sys_id and cntrt_id = :cntrt_id",{"tier1_srce_sys_id": tier1_srce_sys_id,"cntrt_id": cntrt_id})
    mm_prod_sdim.createOrReplaceTempView('mm_prod_sdim')

    df_fct = read_fact(catalog_name,spark,time_prd,tier1_srce_sys_id, cntrt_id)
    df_fct.createOrReplaceTempView(f'mm_tp_{time_prd}_fct')

    tp_run_prttn_plc = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_run_prttn_plc where cntrt_id=:cntrt_id",{"cntrt_id":cntrt_id})
    tp_run_prttn_plc.createOrReplaceTempView('mm_run_prttn_plc')

    tp_run_measr_plc = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_run_measr_plc where cntrt_id=:cntrt_id",{"cntrt_id":cntrt_id})
    tp_run_measr_plc.createOrReplaceTempView('mm_run_measr_plc')

    tp_prod_sdim = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_prod_sdim")
    tp_prod_sdim.createOrReplaceTempView('tp_prod_sdim')

    tp_time_perd_assoc = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_assoc")
    tp_time_perd_assoc.createOrReplaceTempView('mm_time_perd_assoc')

    tp_time_perd_fdim = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_fdim")
    tp_time_perd_fdim.createOrReplaceTempView('mm_time_perd_fdim')

    tp_time_perd_assoc_type = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_time_perd_assoc_type")
    tp_time_perd_assoc_type.createOrReplaceTempView('mm_time_perd_assoc_type')

    tp_run_mkt_plc = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_run_mkt_plc where cntrt_id=:cntrt_id",{"cntrt_id":cntrt_id})
    tp_run_mkt_plc.createOrReplaceTempView('MM_RUN_MKT_PLC')

    tp_run_time_perd_plc = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_run_time_perd_plc where cntrt_id=:cntrt_id",{"cntrt_id":cntrt_id})
    tp_run_time_perd_plc.createOrReplaceTempView('mm_run_time_perd_plc')

    tp_run_prod_plc = spark.sql(f"SELECT * FROM {catalog_name}.internal_tp.tp_run_prod_plc where cntrt_id=:cntrt_id",{"cntrt_id":cntrt_id})
    tp_run_prod_plc.createOrReplaceTempView('MM_RUN_PROD_PLC')

    tp_prod_dim = spark.sql(f"SELECT * FROM {catalog_name}.gold_tp.tp_prod_dim")
    tp_prod_dim.createOrReplaceTempView('mm_prod_dim')

    # Register mm_cntrt_lkp again for aliasing
    mm_cntrt_lkp.createOrReplaceTempView('MM_CNTRT_TIME_PERD_TYPE_ASSOC')

    # Create derived view
    try:
        df_mm_time_perd_assoc_tier1_vw = mm_time_perd_assoc_tier1_vw(
            spark,
            spark.table("mm_time_perd_assoc"),
            spark.table("mm_time_perd_fdim"),
            spark.table("mm_time_perd_assoc_type")
        )
        df_mm_time_perd_assoc_tier1_vw.createOrReplaceTempView('mm_time_perd_assoc_tier1_vw')
    except Exception as e:
        logger.error(f"Failed to create mm_time_perd_assoc_tier1_vw: {e}")

    
    logger.info("Validations are starting")
    data = []

    # ------------------Unknown measure Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Unknown measure',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Unknown measure',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Duplicated measure after mapping-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Duplicated measure after mapping',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Duplicated measure after mapping',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------N/A or zero in every input row for Measure-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','N/A or zero in every input row for Measure',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'N/A or zero in every input row for Measure',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Unknown attribute value for product Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Unknown attribute value for product',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Unknown attribute value for product',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------N/A or zero in every input row for time period Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','N/A or zero in every input row for time period',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'N/A or zero in every input row for time period',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------N/A or zero in every input row for time period Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','N/A or zero in every input row, for market',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'N/A or zero in every input row, for market',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    # ------------------Time period of weeks not properly generated Validation---------------
    if time_perd_class_code == 'WK':
        df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Time period of weeks not properly generated',spark)
        
    else:
        df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Time period of weeks not properly generated not week',spark)

    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    dq_val = validation(catalog_name,query2,'Time period of weeks not properly generated',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    

    # ------------------Duplicated time period after mapping Validation -----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Duplicated time period after mapping',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Duplicated time period after mapping',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Missing/delivered areas Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Missing/delivered areas',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    dq_val = validation(catalog_name,query2,'Missing/delivered areas',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    
    # ------------------Unknown time period Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Unknown time period',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Unknown time period',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Unknown Market Validation -----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Unknown Market',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Unknown Market',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)


    # ------------------Duplicated market after mapping Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Duplicated market after mapping',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Duplicated market after mapping',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    # ------------------Missing/delivered hierarchies Product Validation -----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Missing/delivered hierarchies Product',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    dq_val = validation(catalog_name,query2,'Missing/delivered hierarchies Product',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    # ------------------Missing/delivered measures Validation -----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Missing/delivered measures',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    dq_val = validation(catalog_name,query2,'Missing/delivered measures',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

        
    # ------------------Modified product attributes Validation-----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Modified product attributes',spark)
    query = df.collect()[0]['qry_txt']
    query2 = f"f'''{query}'''"
    e=eval
    query2 = e(query2)
    df_val = spark.sql(query2)
    dq_val = validation(catalog_name,df_val,'Modified product attributes',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    # ------------------No parent in market hierarchy Validation -----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','No parent in market hierarchy',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'No parent in market hierarchy',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Unknown hierarchy for market Validation -----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Unknown hierarchy for market',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Unknown hierarchy for market',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)

    # ------------------Too many attributes for market hierachy Validation -----------------------
    df = dq_query_retrieval(catalog_name,validation_name,'Tier1','Too many attributes for market hierachy',spark)
    query = df.collect()[0]['qry_txt']
    dq_val = validation(catalog_name,query,'Too many attributes for market hierachy',run_id,spark,validation_name)
    print(dq_val)
    data.append(dq_val)
    
    #  Create Dataframe Schema and Merge into tp_data_vldtn_rprt table
    merge_into_tp_data_vldtn_rprt_tbl(run_id,spark,data,valdn_grp_id,catalog_name,postgres_schema, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

    # Report Formation
    generate_report(run_id,cntrt_id,file_name,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd,catalog_name,postgres_schema)
    
    # # # Report Generation
    generate_formatting_report(run_id,catalog_name)

    # Calling send_email function to send emails to vendors
    send_emails(spark,run_id, cntrt_id,catalog_name,postgres_schema, file_name,validation_name,ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    

def add_row_count(count_of_df,postgres_schema,run_id,cntrt_id,spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd):
    #Add Row_count to mm_run_detl_plc
    run_id = int(run_id)
    cntrt_id = int(cntrt_id)
    df_run_detl_plc = read_query_from_postgres(f"SELECT run_id from {postgres_schema}.mm_run_detl_plc where run_id={int(run_id)}", spark, ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)
    
    if df_run_detl_plc.count() == 0:
        
        # Define schema
        schema = StructType([
            StructField("row_cnt", LongType(), True),
            StructField("run_id", LongType(), True),
            StructField("cntrt_id", LongType(), True)
        ])

        # Create DataFrame
        data = [(count_of_df, run_id, cntrt_id)]
        df_row_cnt = spark.createDataFrame(data, schema)

        write_to_postgres(df_row_cnt, f"{postgres_schema}.mm_run_detl_plc", ref_db_jdbc_url, ref_db_name, ref_db_user, ref_db_pwd)

def add_latest_time_perd(df,postgres_schema,run_id,ref_db_name, ref_db_user, ref_db_pwd,ref_db_hostname):
    #Add latest time_perd into mm_run_detl_plc table
    max_date = df.select(max("mm_time_perd_end_date")).first()[0]
    latest_perd_date = df.filter(col("mm_time_perd_end_date") == max_date).select("time_perd_id").first()[0]
    params=()
    query=f"UPDATE {postgres_schema}.mm_run_detl_plc SET latst_time_perd_id = '{latest_perd_date}' WHERE run_id = {int(run_id)} "
    update_to_postgres(query,params, ref_db_name, ref_db_user, ref_db_pwd,ref_db_hostname)